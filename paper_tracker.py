"""
经济学论文追踪系统 - 单文件完整版
运行方式: python paper_tracker.py
打包方式: pyinstaller --onefile --windowed --name PaperTracker paper_tracker.py
"""

# =============================================================================
# 标准库导入
# =============================================================================
import sys
import os
import re
import csv
import json
import time
import logging
import subprocess
import urllib.request
import urllib.parse
import urllib.error
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

# =============================================================================
# 配置区（用户可在 GUI 中修改，运行时由 settings.json 持久化）
# =============================================================================

# ── API 密钥（可在 GUI 中填写，无需修改此处）──────────────────────────────────
ANTHROPIC_API_KEY   = ""
ANTHROPIC_BASE_URL  = ""   # 原生 Anthropic 留空；中转站填入如 https://xxx.com
MODEL_FAST          = ""   # 翻译用轻量模型，留空默认 claude-haiku-4-5-20251001
MODEL_STRONG        = ""   # 综述用强力模型，留空默认 claude-sonnet-4-6
API_TYPE            = "anthropic"  # "anthropic" 或 "openai"（兼容 OpenAI 格式的接口）
CROSSREF_EMAIL      = ""   # 填入邮箱可获得 CrossRef 更高速率限制

# ── 英文顶刊列表（ISSN → 期刊信息）────────────────────────────────────────────
ENGLISH_JOURNALS = {
    # 综合经济学 Top5
    "0002-8282": {"name": "American Economic Review",              "abbr": "AER",      "tier": "Top5"},
    "0033-5533": {"name": "Quarterly Journal of Economics",        "abbr": "QJE",      "tier": "Top5"},
    "0022-3808": {"name": "Journal of Political Economy",          "abbr": "JPE",      "tier": "Top5"},
    "0034-6527": {"name": "Review of Economic Studies",            "abbr": "REStud",   "tier": "Top5"},
    "0012-9682": {"name": "Econometrica",                          "abbr": "ECMA",     "tier": "Top5"},
    # 综合
    "0034-6535": {"name": "Review of Economics and Statistics",    "abbr": "REStat",   "tier": "Top"},
    "0022-0515": {"name": "Journal of Economic Literature",        "abbr": "JEL",      "tier": "Top"},
    "0895-3309": {"name": "Journal of Economic Perspectives",      "abbr": "JEP",      "tier": "Top"},
    "1759-7323": {"name": "American Economic Journal: Applied Economics", "abbr": "AEJ:AE",  "tier": "Top"},
    "1945-7707": {"name": "American Economic Journal: Economic Policy",   "abbr": "AEJ:EP",  "tier": "Top"},
    "1945-7685": {"name": "American Economic Journal: Macroeconomics",    "abbr": "AEJ:Mac", "tier": "Top"},
    "1945-7669": {"name": "American Economic Journal: Microeconomics",    "abbr": "AEJ:Mic", "tier": "Top"},
    # 金融
    "0022-1082": {"name": "Journal of Finance",                    "abbr": "JF",       "tier": "Top"},
    "0304-405X": {"name": "Journal of Financial Economics",        "abbr": "JFE",      "tier": "Top"},
    "0893-9454": {"name": "Review of Financial Studies",           "abbr": "RFS",      "tier": "Top"},
    "0022-1090": {"name": "Journal of Financial and Quantitative Analysis", "abbr": "JFQA", "tier": "Top"},
    # 劳动/发展
    "0734-306X": {"name": "Journal of Labor Economics",            "abbr": "JOLE",     "tier": "Top"},
    "0095-2583": {"name": "Journal of Human Resources",            "abbr": "JHR",      "tier": "Top"},
    "0304-3878": {"name": "Journal of Development Economics",      "abbr": "JDE",      "tier": "Top"},
    # 国际/计量
    "0022-1996": {"name": "Journal of International Economics",    "abbr": "JIE",      "tier": "Top"},
    "0143-9782": {"name": "Journal of Time Series Analysis",       "abbr": "JTSA",     "tier": "Top"},
    "0304-4076": {"name": "Journal of Econometrics",               "abbr": "JoE",      "tier": "Top"},
    # 工作论文
    "nber-wp":   {"name": "NBER Working Paper",                    "abbr": "NBER WP",  "tier": "Top"},
}

# ── 抓取参数 ──────────────────────────────────────────────────────────────────
MAX_PAPERS_PER_JOURNAL = 50
YEARS_BACK             = 3
MIN_CITATIONS          = 0
DEFAULT_SEARCH_MODE    = "both"
REQUEST_DELAY          = 1.0
TRANSLATE_BATCH_SIZE   = 3

# ── 输出路径（动态，相对于用户桌面）──────────────────────────────────────────
def _get_output_root() -> str:
    """返回输出根目录"""
    root = r"D:\claude\research\results\tables"
    try:
        os.makedirs(root, exist_ok=True)
        return root
    except Exception:
        fallback = os.path.join(os.path.expanduser("~"), "Desktop", "PaperTracker")
        os.makedirs(fallback, exist_ok=True)
        return fallback

def get_output_path(keyword: str, suffix: str = "") -> str:
    """生成 Excel 输出路径"""
    date_str = datetime.now().strftime("%Y%m%d")
    # 替换 Windows 文件名非法字符：\ / : * ? " < > |
    safe_kw = re.sub(r'[\\/:*?"<>|]', "_", keyword).replace(" ", "_")[:40]
    folder = os.path.join(_get_output_root(), date_str)
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, f"papers_{safe_kw}{suffix}_{date_str}.xlsx")

def get_review_path(keyword: str) -> str:
    """生成综述输出路径"""
    date_str = datetime.now().strftime("%Y%m%d")
    safe_kw = re.sub(r'[\\/:*?"<>|]', "_", keyword).replace(" ", "_")[:30]
    folder = r"D:\claude\research\paper"
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, f"review_{safe_kw}_{date_str}.txt")

def get_api_url(endpoint: str = "/v1/messages") -> str:
    base = (ANTHROPIC_BASE_URL or "https://api.anthropic.com").rstrip("/")
    return base + endpoint

def get_model(kind: str = "fast") -> str:
    if kind == "fast":
        return MODEL_FAST or ("claude-haiku-4-5-20251001" if API_TYPE == "anthropic" else "gpt-4o-mini")
    return MODEL_STRONG or ("claude-sonnet-4-6" if API_TYPE == "anthropic" else "gpt-4o")

def _call_llm(prompt: str, max_tokens: int = 8192, kind: str = "fast") -> Optional[str]:
    """统一 LLM 调用：自动根据 API_TYPE 选择 Anthropic 或 OpenAI 格式"""
    if not ANTHROPIC_API_KEY:
        return None
    model = get_model(kind)
    if API_TYPE == "openai":
        # OpenAI 兼容格式（DeepSeek / Gemini / 豆包 / 任意兼容接口）
        base = (ANTHROPIC_BASE_URL or "https://api.openai.com").rstrip("/")
        url = base + "/v1/chat/completions"
        payload = {
            "model": model,
            "max_tokens": max_tokens,
            "messages": [{"role": "user", "content": prompt}],
        }
        headers = {
            "Authorization": f"Bearer {ANTHROPIC_API_KEY}",
            "content-type": "application/json",
        }
        def _parse(result): return result["choices"][0]["message"]["content"]
    else:
        # Anthropic 原生格式（Claude / 兼容中转）
        url = get_api_url("/v1/messages")
        payload = {
            "model": model,
            "max_tokens": max_tokens,
            "messages": [{"role": "user", "content": prompt}],
            "thinking": {"type": "disabled"},  # 禁用 extended thinking，避免中转服务强制开启
        }
        headers = {
            "x-api-key": ANTHROPIC_API_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        }
        def _parse(result):
            # Anthropic 格式（含 extended thinking 兜底）
            if "content" in result and result["content"]:
                for block in result["content"]:
                    if isinstance(block, dict) and block.get("type") == "text":
                        return block["text"]
                return result["content"][0].get("text", "")
            # OpenAI 兼容格式
            if "choices" in result and result["choices"]:
                return result["choices"][0]["message"]["content"]
            raise KeyError(f"未知响应格式: {list(result.keys())}")

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=300) as resp:
            result = json.loads(resp.read().decode())
            try:
                return _parse(result)
            except (KeyError, IndexError, TypeError) as e:
                logger.error(f"LLM API 调用失败 (url={url}, model={model}): {e}，响应={str(result)[:300]}")
                return None
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")[:300]
        logger.error(f"LLM API HTTP错误 {e.code} (url={url}, model={model}): {body}")
        return f"[错误] HTTP {e.code}: {body}"
    except Exception as e:
        logger.error(f"LLM API 调用失败 (url={url}, model={model}): {e}")
        return None

# settings.json 路径（与可执行文件同目录）
_EXE_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))
SETTINGS_PATH = os.path.join(_EXE_DIR, "paper_tracker_settings.json")


# =============================================================================
# Paper 数据结构
# =============================================================================

@dataclass
class Paper:
    """论文数据结构"""
    title:        str = ""
    title_zh:     str = ""
    abstract:     str = ""
    abstract_zh:  str = ""
    authors:      list = field(default_factory=list)
    year:         int = 0
    journal:      str = ""
    journal_abbr: str = ""
    doi:          str = ""
    url:          str = ""
    citations:    int = -1
    source:       str = "crossref"
    match_reason: str = ""
    keywords:     list = field(default_factory=list)
    download_url: str = ""


# =============================================================================
# 布尔检索表达式解析器
# 支持：* (AND)  + (OR)  - (NOT)  '...' (精确短语)  () (括号分组)
# =============================================================================

class _Token:
    """词法单元"""
    PHRASE = "PHRASE"   # 'quoted phrase'
    WORD   = "WORD"     # 普通词
    AND    = "AND"      # *
    OR     = "OR"       # +
    NOT    = "NOT"      # -（一元前缀）
    LPAREN = "LPAREN"   # (
    RPAREN = "RPAREN"   # )
    EOF    = "EOF"

    def __init__(self, ttype: str, value: str = ""):
        self.ttype = ttype
        self.value = value

    def __repr__(self):
        return f"Token({self.ttype}, {self.value!r})"


def _tokenize_query(expr: str) -> list:
    """将检索表达式拆分为 Token 列表"""
    tokens = []
    i = 0
    n = len(expr)
    while i < n:
        c = expr[i]
        if c in (' ', '\t'):
            i += 1
        elif c == '*':
            tokens.append(_Token(_Token.AND))
            i += 1
        elif c == '+':
            tokens.append(_Token(_Token.OR))
            i += 1
        elif c == '-':
            tokens.append(_Token(_Token.NOT))
            i += 1
        elif c == '(':
            tokens.append(_Token(_Token.LPAREN))
            i += 1
        elif c == ')':
            tokens.append(_Token(_Token.RPAREN))
            i += 1
        elif c == "'":
            # 精确短语，直到下一个单引号
            j = i + 1
            while j < n and expr[j] != "'":
                j += 1
            phrase = expr[i+1:j].strip().lower()
            tokens.append(_Token(_Token.PHRASE, phrase))
            i = j + 1
        else:
            # 普通词（连续非空白非运算符字符）
            j = i
            while j < n and expr[j] not in (' ', '\t', '*', '+', '-', '(', ')', "'"):
                j += 1
            word = expr[i:j].strip().lower()
            if word:
                tokens.append(_Token(_Token.WORD, word))
            i = j
    tokens.append(_Token(_Token.EOF))
    return tokens


class _BoolParser:
    """
    递归下降解析器，优先级（低→高）：
      OR (+) < AND (*) < NOT (-前缀) < 原子（词/短语/括号）
    """

    def __init__(self, tokens: list):
        self._tokens = tokens
        self._pos = 0

    def _peek(self):
        return self._tokens[self._pos]

    def _consume(self, ttype=None):
        tok = self._tokens[self._pos]
        if ttype and tok.ttype != ttype:
            raise ValueError(f"期望 {ttype}，得到 {tok}")
        self._pos += 1
        return tok

    def parse(self):
        node = self._parse_or()
        self._consume(_Token.EOF)
        return node

    def _parse_or(self):
        left = self._parse_and()
        while self._peek().ttype == _Token.OR:
            self._consume(_Token.OR)
            right = self._parse_and()
            left = ("OR", left, right)
        return left

    def _parse_and(self):
        left = self._parse_not()
        while self._peek().ttype == _Token.AND:
            self._consume(_Token.AND)
            right = self._parse_not()
            left = ("AND", left, right)
        return left

    def _parse_not(self):
        if self._peek().ttype == _Token.NOT:
            self._consume(_Token.NOT)
            operand = self._parse_atom()
            return ("NOT", operand)
        return self._parse_atom()

    def _parse_atom(self):
        tok = self._peek()
        if tok.ttype == _Token.LPAREN:
            self._consume(_Token.LPAREN)
            node = self._parse_or()
            self._consume(_Token.RPAREN)
            return node
        elif tok.ttype == _Token.PHRASE:
            self._consume(_Token.PHRASE)
            return ("PHRASE", tok.value)
        elif tok.ttype == _Token.WORD:
            self._consume(_Token.WORD)
            return ("WORD", tok.value)
        else:
            raise ValueError(f"意外的 token: {tok}")


def _is_bool_expr(expr: str) -> bool:
    """判断字符串是否含有布尔运算符（非简单关键词）"""
    return bool(re.search(r"[*+()\']", expr))


def _extract_all_terms(expr: str) -> list[str]:
    """从布尔表达式中提取所有词项（用于 CrossRef 搜索字符串）"""
    terms = []
    for tok in _tokenize_query(expr):
        if tok.ttype in (_Token.WORD, _Token.PHRASE) and tok.value:
            terms.append(tok.value)
    return terms


def _eval_bool_node(node, text: str) -> bool:
    """对单段文本 text 递归求值布尔表达式树"""
    t = node[0]
    if t == "WORD":
        return node[1] in text
    elif t == "PHRASE":
        return node[1] in text
    elif t == "AND":
        return _eval_bool_node(node[1], text) and _eval_bool_node(node[2], text)
    elif t == "OR":
        return _eval_bool_node(node[1], text) or _eval_bool_node(node[2], text)
    elif t == "NOT":
        return not _eval_bool_node(node[1], text)
    return False


def _match_bool_expr(expr: str, *texts: str) -> bool:
    """
    对多段文本（如篇名、摘要、关键词拼接）求值布尔表达式。
    文本全部转小写后拼接。
    """
    combined = " ".join((t or "").lower() for t in texts)
    try:
        tokens = _tokenize_query(expr.lower())
        node = _BoolParser(tokens).parse()
        return _eval_bool_node(node, combined)
    except Exception:
        # 解析失败时退化为简单包含匹配
        return expr.lower() in combined


def _match_bool_expr_fields(expr: str, title: str, abstract: str, keywords: list,
                             use_title: bool, use_kw: bool, use_abs: bool) -> tuple[bool, str]:
    """
    在指定字段的拼合文本上求值布尔表达式，同时标注各字段的命中情况。
    返回 (整体是否匹配, 匹配位置描述)

    逻辑：
    1. 先把勾选的字段文本拼合，对整体求值——决定该论文是否通过过滤
    2. 再逐字段单独判断各个词项是否出现（用于"匹配位置"列的标注）
    """
    try:
        tokens = _tokenize_query(expr.lower())
        node = _BoolParser(tokens).parse()
    except Exception:
        node = None

    def _eval(text: str) -> bool:
        if node:
            return _eval_bool_node(node, (text or "").lower())
        return expr.lower() in (text or "").lower()

    # ── 1. 整体匹配：在所有勾选字段的拼合文本上求值一次 ──────────────────────
    combined_parts = []
    if use_title:   combined_parts.append(title or "")
    if use_kw:      combined_parts.append(" ".join(keywords))
    if use_abs:     combined_parts.append(abstract or "")
    combined = " ".join(combined_parts)
    matched = _eval(combined)

    # ── 2. 匹配位置标注：逐字段检查词项出现情况 ──────────────────────────────
    # 提取表达式中所有词项（忽略运算符），检查各字段是否含有至少一个词项
    all_terms = _extract_all_terms(expr)
    parts = []
    if use_title:
        title_l = (title or "").lower()
        if any(t in title_l for t in all_terms):
            parts.append("篇名")
    if use_kw:
        kw_l = " ".join(keywords).lower()
        if any(t in kw_l for t in all_terms):
            parts.append("关键词")
    if use_abs:
        abs_l = (abstract or "").lower()
        if any(t in abs_l for t in all_terms):
            parts.append("摘要")

    return matched, "、".join(parts)


# =============================================================================
# 知网期刊列表（中文顶刊，按级别分组）
# =============================================================================

CNKI_JOURNALS = {
    # ── 经济类 ────────────────────────────────────────────────────────────────
    "经济类·特A/A1": [
        "经济研究", "经济学(季刊)", "世界经济",
    ],
    "经济类·A/A2": [
        "中国工业经济", "金融研究", "数量经济技术经济研究",
        "统计研究", "经济学动态", "中国农村经济",
    ],
    "经济类·A-": [
        "财贸经济", "国际贸易问题", "南开经济研究", "经济科学",
        "财经研究", "经济评论", "经济学家", "国际金融研究",
        "产业经济研究", "中国人口·资源与环境",
    ],
    # ── 管理类 ────────────────────────────────────────────────────────────────
    "管理类·特A/A1": [
        "管理世界",
    ],
    "管理类·A/A2": [
        "会计研究", "管理科学学报", "系统工程理论与实践", "中国行政管理",
    ],
}

# 所有期刊名平铺列表
CNKI_ALL_JOURNALS = [j for grp in CNKI_JOURNALS.values() for j in grp]

# 知网期刊 CNKI code（通过 CNKI 高级检索弹窗查询所得）
CNKI_JOURNAL_CODES = {
    "经济研究":           "CJFD_JJYJ",
    "管理世界":           "CJFD_GLSJ",
    "经济学(季刊)":       "CJFD_JJXU",
    "世界经济":           "CJFD_SJJJ",
    "中国工业经济":       "CJFD_GGYY",
    "数量经济技术经济研究": "CJFD_SLJY",
    "金融研究":           "CJFD_JRYJ",
    "统计研究":           "CJFD_TJYJ",
    "经济学动态":         "CJFD_JJXD",
    "中国农村经济":       "CJFD_ZNJJ",
    "中国人口·资源与环境": "CJFD_ZGRZ",
    "财贸经济":           "CJFD_CMJJ",
    "国际贸易问题":       "CJFD_GJMW",
    "南开经济研究":       "CJFD_NKJJ",
    "经济科学":           "CJFD_JJKX",
    "财经研究":           "CJFD_CJYJ",
    "经济评论":           "CJFD_JJPL",
    "经济学家":           "CJFD_JJXJ",
    "国际金融研究":       "CJFD_GJJR",
    "产业经济研究":       "CJFD_CYJJ",
    # 暨南大学 A2 新增
    "会计研究":           "CJFD_KJYJ",
    "管理科学学报":       "CJFD_GLKX",
    "系统工程理论与实践": "CJFD_XTGL",
    "中国行政管理":       "CJFD_ZGXZ",
}

# =============================================================================
# 知网 JS 常量（供 CnkiManager 注入到 QWebEnginePage）
# =============================================================================

_CNKI_JS_FILL_ADVSEARCH = """
(function(keyword, journalName) {
    // 检查验证码
    var outer = document.querySelector('#tcaptcha_transform_dy');
    if (outer && outer.getBoundingClientRect().top >= 0) { return 'captcha'; }

    var topicInp = document.querySelector('input[data-tipid="gradetxt-1"]');
    var srcInp   = document.querySelector('input[data-tipid="gradetxt-3"]');
    var btn      = document.querySelector('.btn-search, input.btn-search');
    if (!topicInp || !btn) { return 'no_topic'; }

    // 用 React 内部 setter 触发受控组件的 onChange，避免直接赋值被 React 忽略
    function reactSet(el, val) {
        var nativeInputValueSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value');
        if (nativeInputValueSetter && nativeInputValueSetter.set) {
            nativeInputValueSetter.set.call(el, val);
        } else {
            el.value = val;
        }
        el.dispatchEvent(new Event('input', {bubbles: true}));
        el.dispatchEvent(new Event('change', {bubbles: true}));
    }

    // 填主题词
    topicInp.focus();
    reactSet(topicInp, keyword);

    // 填期刊来源（可选）
    if (srcInp && journalName) {
        srcInp.focus();
        reactSet(srcInp, journalName);
        topicInp.focus();
    }

    btn.click();
    return 'submitted';
})("{KEYWORD}", {JOURNAL_NAME_JSON})
"""

_CNKI_JS_CHECK_FILL_DONE = """
(function() {
    if (window.__cnkiFillDone) return window.__cnkiFillResult || 'done';
    return 'pending';
})()
"""

_CNKI_JS_CHECK_RESULTS = """
(function() {
    var outer = document.querySelector('#tcaptcha_transform_dy');
    if (outer && outer.getBoundingClientRect().top >= 0) return JSON.stringify({error:'captcha'});
    var ready = document.body.innerText.includes('\\u6761\\u7ed3\\u679c');
    return JSON.stringify({ready: ready});
})()
"""

_CNKI_JS_EXTRACT_LIST = """
(function() {
    var rows = document.querySelectorAll('.result-table-list tbody tr');
    var results = [];
    rows.forEach(function(row, i) {
        var titleLink = row.querySelector('td.name a.fz14');
        if (!titleLink) return;
        var authors = Array.from(row.querySelectorAll('td.author a.KnowledgeNetLink')||[]).map(function(a){return (a.innerText||'').trim();});
        var journal = (row.querySelector('td.source a')||{}).innerText || '';
        var date = (row.querySelector('td.date')||{}).innerText || '';
        var citations = (row.querySelector('td.quote')||{}).innerText || '';
        var cbItem = row.querySelector('input.cbItem');
        // 优先找 PDF 下载链接（title 含 PDF），否则取第一个 downloadlink
        var dlLinks = Array.from(row.querySelectorAll('a.downloadlink'));
        var pdfLink = dlLinks.find(function(a){ return (a.title||'').toUpperCase().includes('PDF') || (a.getAttribute('data-type')||'').toUpperCase().includes('PDF'); });
        var cajLink = dlLinks.find(function(a){ return !pdfLink || a !== pdfLink; });
        var dlLink = pdfLink || cajLink || null;
        results.push({
            title: (titleLink.innerText||'').trim(),
            href: titleLink.href || '',
            exportId: cbItem ? cbItem.value : '',
            authors: authors.join('; '),
            journal: (journal||'').trim(),
            date: (date||'').trim(),
            citations: (citations||'').trim(),
            download_url: dlLink ? dlLink.href : ''
        });
    });
    var total = '';
    var m = document.body.innerText.match(/\\u5171\\u627e\\u5230\\s*([\\d,]+)\\s*\\u6761\\u7ed3\\u679c/);
    if (m) total = m[1];
    return JSON.stringify({results: results, total: total});
})()
"""

_CNKI_JS_NEXT_PAGE = """
(function() {
    var links = document.querySelectorAll('a');
    for (var i = 0; i < links.length; i++) {
        if (links[i].innerText && links[i].innerText.trim() === '\\u4e0b\\u4e00\\u9875') {
            links[i].click();
            return JSON.stringify({clicked:true});
        }
    }
    return JSON.stringify({error:'no_next'});
})()
"""

_CNKI_JS_EXTRACT_ABSTRACT = """
(function() {
    var abEl = document.querySelector('.abstract-text');
    var abstract = abEl ? (abEl.innerText||'').trim() : '';
    var kwEls = document.querySelectorAll('p.keywords a');
    var keywords = Array.from(kwEls).map(function(a){return (a.innerText||'').replace(/;$/,'').trim();});
    return JSON.stringify({abstract: abstract, keywords: keywords});
})()
"""

_CNKI_JS_APPLY_SORT = """
(function(sortId) {
    // sortId: FFD=相关度, PT=发表时间, CF=被引, DFR=下载, ZH=综合
    var btn = document.getElementById(sortId);
    if (!btn) return JSON.stringify({ok: false, reason: 'not_found'});
    if (btn.className.indexOf('cur') >= 0) return JSON.stringify({ok: true, already: true});
    btn.click();
    return JSON.stringify({ok: true, already: false});
})("{SORT_ID}")
"""

_CNKI_JS_APPLY_YEAR_FILTER = """
(function(cutoffYear) {
    // 点击左侧年度筛选复选框：勾选所有 >= cutoffYear 的年份
    // 选择器：左侧 facet 面板内年份 checkbox，data-val 属性为年份字符串
    var currentYear = new Date().getFullYear();
    var clicked = 0;
    // 找到年度筛选区域（标题含"年"）
    var yearBoxes = document.querySelectorAll('.filter-box');
    var yearSection = null;
    for (var i = 0; i < yearBoxes.length; i++) {
        var titleEl = yearBoxes[i].querySelector('.filter-title, .title, h3, strong');
        if (titleEl && /年/.test(titleEl.innerText||'')) {
            yearSection = yearBoxes[i];
            break;
        }
    }
    // 如果没找到包含"年"标题的区域，尝试直接查找年份 checkbox
    var checkboxes = yearSection
        ? yearSection.querySelectorAll('input[type="checkbox"]')
        : document.querySelectorAll('.filter-box input[type="checkbox"], .facet-box input[type="checkbox"]');

    var yearCheckboxes = [];
    for (var i = 0; i < checkboxes.length; i++) {
        var val = checkboxes[i].getAttribute('data-val') || checkboxes[i].value || '';
        var yr = parseInt(val);
        if (yr >= 2000 && yr <= currentYear) {
            yearCheckboxes.push({el: checkboxes[i], year: yr});
        }
    }

    if (yearCheckboxes.length === 0) {
        return JSON.stringify({applied: false, reason: 'no_year_checkboxes'});
    }

    // 勾选 >= cutoffYear 的复选框，取消 < cutoffYear 的
    for (var i = 0; i < yearCheckboxes.length; i++) {
        var cb = yearCheckboxes[i].el;
        var yr = yearCheckboxes[i].year;
        if (yr >= cutoffYear && !cb.checked) {
            cb.click();
            clicked++;
        } else if (yr < cutoffYear && cb.checked) {
            cb.click();
        }
    }

    return JSON.stringify({applied: true, clicked: clicked, years: yearCheckboxes.map(function(x){return x.year;})});
})({CUTOFF_YEAR})
"""

# =============================================================================
# 英文期刊抓取（CrossRef + Semantic Scholar）
# =============================================================================

logger = logging.getLogger(__name__)


def _http_get(url: str, timeout: int = 20) -> Optional[dict]:
    try:
        headers = {
            "User-Agent": f"PaperTracker/1.0 (mailto:{CROSSREF_EMAIL or 'user@example.com'})",
            "Accept": "application/json",
        }
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        logger.warning(f"GET {url} 失败: {e}")
        return None


def _clean_abstract(raw: str) -> str:
    if not raw:
        return ""
    cleaned = re.sub(r"<[^>]+>", " ", raw)
    return re.sub(r"\s+", " ", cleaned).strip()


def _crossref_search_journal(
    issn: str, keyword: str, max_results: int = 50,
    years_back: int = 3, sort: str = "relevance",
) -> list[dict]:
    year_from = datetime.now().year - years_back if years_back > 0 else 2000
    # NBER Working Paper 没有 ISSN，用 member:1960 + type:report 过滤
    if issn == "nber-wp":
        params = {
            "query": keyword,
            "filter": f"member:1960,type:report,from-pub-date:{year_from}",
            "rows": min(max_results, 100),
            "sort": sort,
            "order": "desc",
            "select": "DOI,title,abstract,author,published,container-title,URL,is-referenced-by-count,subject",
        }
    else:
        params = {
            "query": keyword,
            "filter": f"issn:{issn},from-pub-date:{year_from}",
            "rows": min(max_results, 100),
            "sort": sort,
            "order": "desc",
            "select": "DOI,title,abstract,author,published,container-title,URL,is-referenced-by-count,subject",
        }
    url = "https://api.crossref.org/works?" + urllib.parse.urlencode(params)
    data = _http_get(url)
    if not data:
        return []
    items = data.get("message", {}).get("items", [])
    logger.info(f"  CrossRef [{issn}] sort={sort}: 返回 {len(items)} 篇")
    return items


def _parse_crossref_item(item: dict, journal_info: dict) -> Paper:
    titles = item.get("title", [])
    title = titles[0] if titles else ""
    abstract = _clean_abstract(item.get("abstract", ""))
    authors = []
    for a in item.get("author", []):
        name = f"{a.get('given', '')} {a.get('family', '')}".strip()
        if name:
            authors.append(name)
    pub = item.get("published", {})
    date_parts = pub.get("date-parts", [[0]])[0]
    year = date_parts[0] if date_parts else 0
    doi = item.get("DOI", "")
    url = item.get("URL", f"https://doi.org/{doi}" if doi else "")
    citations = item.get("is-referenced-by-count", -1)
    keywords = [s.lower() for s in item.get("subject", []) if isinstance(s, str)]
    return Paper(
        title=title, abstract=abstract, authors=authors, year=year,
        journal=journal_info["name"], journal_abbr=journal_info["abbr"],
        doi=doi, url=url, citations=citations, source="crossref",
        keywords=keywords,
    )


def _openalex_fetch_batch(batch: list[str]) -> dict:
    """查询一批 DOI 的 OpenAlex 数据，返回 {doi: {cite, abstract}}"""
    result = {}
    filter_str = "doi:" + "|".join(batch)
    url = (
        "https://api.openalex.org/works?filter=" +
        urllib.parse.quote(filter_str, safe=":|/") +
        "&select=doi,cited_by_count,abstract_inverted_index,keywords&per-page=50"
        "&mailto=papertracker@example.com"
    )
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "PaperTracker/1.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode())
        for item in data.get("results", []):
            raw_doi = (item.get("doi") or "").lower()
            doi = raw_doi.replace("https://doi.org/", "").strip()
            if not doi:
                continue
            entry = {}
            if item.get("cited_by_count") is not None:
                entry["cite"] = item["cited_by_count"]
            inv = item.get("abstract_inverted_index")
            if inv:
                max_pos = max(pos for positions in inv.values() for pos in positions)
                words = [""] * (max_pos + 1)
                for word, positions in inv.items():
                    for pos in positions:
                        words[pos] = word
                abstract = " ".join(w for w in words if w).strip()
                if len(abstract) > 80:
                    entry["abstract"] = abstract
            if entry:
                result[doi] = entry
            # 关键词（即使无摘要也记录）
            kws = item.get("keywords") or []
            kw_list = [kw.get("keyword", "").lower() for kw in kws if isinstance(kw, dict) and kw.get("keyword")]
            if not kw_list:
                kw_list = [kw.lower() for kw in kws if isinstance(kw, str)]
            if kw_list:
                if doi in result:
                    result[doi]["keywords"] = kw_list
                else:
                    result[doi] = {"keywords": kw_list}
    except Exception as e:
        logger.warning(f"OpenAlex batch 失败: {e}")
    return result


def _enrich_citations_batch(papers: list[Paper]) -> None:
    """通过 OpenAlex（并发主）+ Semantic Scholar（辅）批量补充引用数和缺失摘要"""

    dois_all = [p.doi for p in papers if p.doi]
    doi_to_cite: dict[str, int] = {}
    doi_to_abstract: dict[str, str] = {}
    doi_to_keywords: dict[str, list] = {}

    # ── 第一步：OpenAlex 并发批量查询 ─────────────────────────────────────────
    logger.info("补全 OpenAlex 摘要 & 引用数...")
    batches = [dois_all[i:i + 50] for i in range(0, len(dois_all), 50)]
    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(_openalex_fetch_batch, b): b for b in batches}
        for future in as_completed(futures):
            batch_result = future.result()
            for doi, entry in batch_result.items():
                if "cite" in entry:
                    doi_to_cite[doi] = entry["cite"]
                if "abstract" in entry:
                    doi_to_abstract[doi] = entry["abstract"]
                if "keywords" in entry:
                    doi_to_keywords[doi] = entry["keywords"]

    # ── 第二步：Semantic Scholar 补充仍缺摘要的论文（小批次+退避）──────────────
    dois_still_missing = [
        p.doi for p in papers if p.doi and p.doi.lower() not in doi_to_abstract
    ]
    if dois_still_missing:
        logger.info(f"Semantic Scholar 补充 {len(dois_still_missing)} 篇...")
        for i in range(0, len(dois_still_missing), 25):
            batch = dois_still_missing[i:i + 25]
            url = "https://api.semanticscholar.org/graph/v1/paper/batch?fields=citationCount,abstract"
            for attempt in range(3):
                try:
                    data_bytes = json.dumps({"ids": [f"DOI:{d}" for d in batch]}).encode()
                    req = urllib.request.Request(
                        url, data=data_bytes,
                        headers={"Content-Type": "application/json"}, method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=20) as resp:
                        results = json.loads(resp.read().decode())
                    for item in results:
                        if not item:
                            continue
                        ext_ids = item.get("externalIds", {}) or {}
                        doi = (ext_ids.get("DOI") or "").lower()
                        if not doi:
                            continue
                        if "citationCount" in item and doi not in doi_to_cite:
                            doi_to_cite[doi] = item["citationCount"]
                        ab = (item.get("abstract") or "").strip()
                        if ab and doi not in doi_to_abstract:
                            doi_to_abstract[doi] = ab
                    break
                except Exception as e:
                    if "429" in str(e):
                        wait = 5 * (2 ** attempt)
                        logger.warning(f"Semantic Scholar 限流，等待 {wait}s 后重试...")
                        time.sleep(wait)
                    else:
                        logger.warning(f"Semantic Scholar batch 失败: {e}")
                        break
            time.sleep(1.0)

    # ── 写回 papers ────────────────────────────────────────────────────────────
    for p in papers:
        if not p.doi:
            continue
        key = p.doi.lower()
        if p.citations < 0:
            p.citations = doi_to_cite.get(key, 0)
        if not p.abstract and key in doi_to_abstract:
            p.abstract = doi_to_abstract[key]
        if not p.keywords and key in doi_to_keywords:
            p.keywords = doi_to_keywords[key]


def _fetch_url_follow_redirect(url: str, timeout: int = 15) -> str:
    """跟踪 HTTP 302 和 HTML meta refresh 重定向，返回最终页面 HTML"""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    }
    current_url = url
    for _ in range(8):  # 最多跟踪 8 次跳转
        req = urllib.request.Request(current_url, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                html = resp.read().decode("utf-8", errors="replace")
                base_url = resp.url
        except urllib.error.HTTPError as e:
            if e.code in (301, 302, 303, 307, 308):
                location = e.headers.get("Location", "")
                if location:
                    current_url = urllib.parse.urljoin(current_url, location)
                    continue
            raise

        # 检查 HTML meta refresh 跳转（处理 Elsevier linkinghub 等）
        m = re.search(r"HTTP-EQUIV=['\"]REFRESH['\"][^>]+content=['\"][^'\"]*url='([^']+)'", html, re.I)
        if not m:
            m = re.search(r"HTTP-EQUIV=['\"]REFRESH['\"][^>]+content=['\"][^'\"]*url=([^\s'\"&>]+)", html, re.I)
        if m:
            raw_next = m.group(1)
            # Elsevier 把真实 URL 编码在 Redirect= 参数里
            redirect_m = re.search(r"Redirect=([^&'\">\s]+)", raw_next)
            if redirect_m:
                raw_next = urllib.parse.unquote(redirect_m.group(1))
            next_url = urllib.parse.urljoin(base_url, raw_next)
            if next_url != current_url:
                current_url = next_url
                continue

        # 没有跳转，返回当前页面
        return html
    return ""


def _extract_abstract_from_html(html: str) -> str:
    """从各出版商 HTML 中提取摘要文本"""
    if not html:
        return ""

    # ── ScienceDirect（Elsevier）——摘要在 class 含 abstract 的 div ──────────
    m = re.search(
        r'<div[^>]+class="[^"]*(?:abstract|Abstract)[^"]*"[^>]*>(.*?)</div>',
        html, re.S | re.I
    )
    if m:
        text = re.sub(r"<[^>]+>", " ", m.group(1))
        text = re.sub(r"\s+", " ", text).strip()
        # 去掉开头的 "Abstract" 词
        text = re.sub(r"^Abstract\s*", "", text, flags=re.I).strip()
        if len(text) > 80:
            return text

    # ── meta 标签（大多数出版商都有）──────────────────────────────────────────
    for pat in [
        r'<meta[^>]+name="citation_abstract"[^>]+content="([^"]{80,})"',
        r'<meta[^>]+content="([^"]{80,})"[^>]+name="citation_abstract"',
        r'<meta[^>]+property="og:description"[^>]+content="([^"]{80,})"',
        r'<meta[^>]+content="([^"]{80,})"[^>]+property="og:description"',
        r'<meta[^>]+name="description"[^>]+content="([^"]{80,})"',
    ]:
        m2 = re.search(pat, html, re.I)
        if m2:
            text = re.sub(r"<[^>]+>", " ", m2.group(1))
            text = re.sub(r"&amp;", "&", text)
            text = re.sub(r"&lt;", "<", text)
            text = re.sub(r"&gt;", ">", text)
            text = re.sub(r"\s+", " ", text).strip()
            if len(text) > 80:
                return text

    # ── JSON-LD structured data ────────────────────────────────────────────────
    for jl in re.findall(r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>', html, re.S | re.I):
        try:
            d = json.loads(jl)
            ab = d.get("description") or d.get("abstract") or ""
            if isinstance(ab, str) and len(ab) > 80:
                return re.sub(r"\s+", " ", ab).strip()
        except Exception:
            pass

    # ── Wiley / Oxford / INFORMS / Springer section/div ───────────────────────
    for pat in [
        r'<section[^>]+class="[^"]*abstract[^"]*"[^>]*>(.*?)</section>',
        r'<div[^>]+class="[^"]*abstractSection[^"]*"[^>]*>(.*?)</div>',
        r'<div[^>]+class="[^"]*c-article-section__content[^"]*"[^>]*>(.*?)</div>',
        r'<div[^>]+id="abstract[^"]*"[^>]*>(.*?)</div>',
    ]:
        m3 = re.search(pat, html, re.S | re.I)
        if m3:
            text = re.sub(r"<h\d[^>]*>.*?</h\d>", " ", m3.group(1), flags=re.S | re.I)
            text = re.sub(r"<[^>]+>", " ", text)
            text = re.sub(r"\s+", " ", text).strip()
            if len(text) > 80:
                return text

    return ""


def _fetch_one_abstract(p: "Paper") -> tuple:
    """单篇摘要补全：Unpaywall → 网页爬取，返回 (paper, abstract)"""
    ab = ""
    # 1) Unpaywall
    if p.doi:
        try:
            up_url = (
                f"https://api.unpaywall.org/v2/"
                f"{urllib.parse.quote(p.doi, safe='')}?email=papertracker@example.com"
            )
            req = urllib.request.Request(up_url, headers={"User-Agent": "PaperTracker/1.0"})
            with urllib.request.urlopen(req, timeout=8) as resp:
                data = json.loads(resp.read().decode())
            ab = (data.get("abstract") or "").strip()
        except Exception:
            pass
    # 2) 网页爬取
    if not ab and p.url:
        try:
            html = _fetch_url_follow_redirect(p.url, timeout=12)
            ab = _extract_abstract_from_html(html)
        except Exception:
            pass
    return (p, ab)


def scrape_missing_abstracts(papers: list[Paper], log_fn=None) -> int:
    """对仍缺摘要的论文并发补全（Unpaywall + 网页爬取），返回成功数"""
    missing = [p for p in papers if not p.abstract and p.url]
    if not missing:
        return 0

    filled = 0
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(_fetch_one_abstract, p): p for p in missing}
        for future in as_completed(futures):
            p, ab = future.result()
            if ab:
                p.abstract = ab
                filled += 1
                if log_fn:
                    log_fn(f"[INFO] 摘要补全: {p.journal_abbr} — {p.title[:55]}")

    if log_fn:
        log_fn(f"[INFO] 网页/Unpaywall 额外补全 {filled}/{len(missing)} 篇")
    return filled


def fetch_english_papers(
    keyword: str, mode: str = "both",
    max_per_journal: int = None, years_back: int = None,
    journals: dict = None,
) -> list[Paper]:
    max_per_journal = max_per_journal or MAX_PAPERS_PER_JOURNAL
    years_back = years_back if years_back is not None else YEARS_BACK
    journals = journals or ENGLISH_JOURNALS

    all_papers: list[Paper] = []
    seen_dois: set[str] = set()

    for issn, jinfo in journals.items():
        logger.info(f"抓取 {jinfo['abbr']} ({issn})...")
        items_latest, items_cited = [], []
        if mode in ("latest", "both"):
            items_latest = _crossref_search_journal(issn, keyword, max_per_journal, years_back, "published")
            time.sleep(REQUEST_DELAY)
        if mode in ("cited", "both"):
            items_cited = _crossref_search_journal(issn, keyword, max_per_journal, years_back, "relevance")
            time.sleep(REQUEST_DELAY)

        combined = {
            item.get("DOI", ""): item
            for item in (items_latest + items_cited)
            if item.get("title")
        }
        for doi, item in combined.items():
            if doi in seen_dois:
                continue
            seen_dois.add(doi)
            paper = _parse_crossref_item(item, jinfo)
            if paper.title:
                all_papers.append(paper)

    logger.info("补全 Semantic Scholar 引用数...")
    _enrich_citations_batch(all_papers)
    logger.info(f"英文期刊共抓取 {len(all_papers)} 篇论文")
    return all_papers


def fetch_top5_only(keyword: str, **kwargs) -> list[Paper]:
    top5 = {k: v for k, v in ENGLISH_JOURNALS.items() if v["tier"] == "Top5"}
    return fetch_english_papers(keyword, journals=top5, **kwargs)


# ── 管理类期刊（交大安泰 A + A- 共 46 本）──────────────────────────────────────
MANAGEMENT_JOURNALS = {
    # A 级（21本）
    "0001-4826": {"name": "Accounting Review",                                          "abbr": "ACCOUNT REV",        "tier": "A"},
    "0165-4101": {"name": "Journal of Accounting & Economics",                          "abbr": "J ACCOUNT ECON",     "tier": "A"},
    "0021-8456": {"name": "Journal of Accounting Research",                             "abbr": "J ACCOUNT RES",      "tier": "A"},
    "1047-7047": {"name": "Information Systems Research",                               "abbr": "INFORM SYST RES",    "tier": "A"},
    "1091-9856": {"name": "INFORMS Journal on Computing",                               "abbr": "INFORMS J COMPUT",   "tier": "A"},
    "0276-7783": {"name": "MIS Quarterly",                                              "abbr": "MIS QUART",          "tier": "A"},
    "0093-5301": {"name": "Journal of Consumer Research",                               "abbr": "J CONSUM RES",       "tier": "A"},
    "0022-2429": {"name": "Journal of Marketing",                                       "abbr": "J MARKETING",        "tier": "A"},
    "0022-2437": {"name": "Journal of Marketing Research",                              "abbr": "J MARKETING RES",    "tier": "A"},
    "0732-2399": {"name": "Marketing Science",                                          "abbr": "MARKET SCI",         "tier": "A"},
    "0025-1909": {"name": "Management Science",                                         "abbr": "MANAGE SCI",         "tier": "A"},
    "0030-364X": {"name": "Operations Research",                                        "abbr": "OPER RES",           "tier": "A"},
    "0272-6963": {"name": "Journal of Operations Management",                           "abbr": "J OPER MANAG",       "tier": "A"},
    "1523-4614": {"name": "M&SOM Manufacturing & Service Operations Management",        "abbr": "M&SOM-MANUF SERV OP","tier": "A"},
    "1059-1478": {"name": "Production and Operations Management",                       "abbr": "PROD OPER MANAG",    "tier": "A"},
    "0001-4273": {"name": "Academy of Management Journal",                              "abbr": "ACAD MANAGE J",      "tier": "A"},
    "0363-7425": {"name": "Academy of Management Review",                               "abbr": "ACAD MANAGE REV",    "tier": "A"},
    "0001-8392": {"name": "Administrative Science Quarterly",                           "abbr": "ADMIN SCI QUART",    "tier": "A"},
    "0047-2506": {"name": "Journal of International Business Studies",                  "abbr": "J INT BUS STUD",     "tier": "A"},
    "1047-7039": {"name": "Organization Science",                                       "abbr": "ORGAN SCI",          "tier": "A"},
    "0143-2095": {"name": "Strategic Management Journal",                               "abbr": "STRATEGIC MANAGE J", "tier": "A"},
    # A- 级（25本）
    "0361-3682": {"name": "Accounting Organizations and Society",                       "abbr": "ACCOUNT ORG SOC",    "tier": "A-"},
    "0823-9150": {"name": "Contemporary Accounting Research",                           "abbr": "CONTEMP ACCOUNT RES","tier": "A-"},
    "1380-6653": {"name": "Review of Accounting Studies",                               "abbr": "REV ACCOUNT STUD",   "tier": "A-"},
    "1057-7408": {"name": "Journal of Consumer Psychology",                             "abbr": "J CONSUM PSYCHOL",   "tier": "A-"},
    "0162-1459": {"name": "Journal of the American Statistical Association",            "abbr": "J AM STAT ASSOC",    "tier": "A-"},
    "1042-2587": {"name": "Entrepreneurship Theory and Practice",                       "abbr": "ENTREP THEORY PRACT","tier": "A-"},
    "0090-4848": {"name": "Human Resource Management",                                  "abbr": "HUM RESOUR MANAGE-US","tier": "A-"},
    "0021-9010": {"name": "Journal of Applied Psychology",                              "abbr": "J APPL PSYCHOL",     "tier": "A-"},
    "0167-4544": {"name": "Journal of Business Ethics",                                 "abbr": "J BUS ETHICS",       "tier": "A-"},
    "0883-9026": {"name": "Journal of Business Venturing",                              "abbr": "J BUS VENTURING",    "tier": "A-"},
    "0022-2380": {"name": "Journal of Management Studies",                              "abbr": "J MANAGE STUD",      "tier": "A-"},
    "0170-8406": {"name": "Organization Studies",                                       "abbr": "ORGAN STUD",         "tier": "A-"},
    "0749-5978": {"name": "Organizational Behavior and Human Decision Processes",       "abbr": "ORGAN BEHAV HUM DEC","tier": "A-"},
    "1558-9080": {"name": "Academy of Management Perspectives",                         "abbr": "ACAD MANAGE PERSPECT","tier": "A-"},
    "0025-5610": {"name": "Mathematical Programming",                                   "abbr": "MATH PROGRAM",       "tier": "A-"},
    "0041-1655": {"name": "Transportation Science",                                     "abbr": "TRANSPORT SCI",      "tier": "A-"},
    "0149-2063": {"name": "Journal of Management",                                      "abbr": "J MANAGE",           "tier": "A-"},
    "0048-7333": {"name": "Research Policy",                                            "abbr": "RES POLICY",         "tier": "A-"},
    "0894-3796": {"name": "Journal of Organizational Behavior",                         "abbr": "J ORGAN BEHAV",      "tier": "A-"},
    "0031-5826": {"name": "Personnel Psychology",                                       "abbr": "PERS PSYCHOL",       "tier": "A-"},
    "0742-1222": {"name": "Journal of Management Information Systems",                  "abbr": "J MANAGE INFORM SYST","tier": "A-"},
    "0092-0703": {"name": "Journal of the Academy of Marketing Science",                "abbr": "J ACAD MARKET SCI",  "tier": "A-"},
    "0022-4359": {"name": "Journal of Retailing",                                       "abbr": "J RETAILING",        "tier": "A-"},
    "0740-817X": {"name": "IIE Transactions",                                           "abbr": "IIE TRANS",          "tier": "A-"},
    "0894-069X": {"name": "Naval Research Logistics",                                   "abbr": "NAV RES LOG",        "tier": "A-"},
}

# ── 经济类期刊（交大安泰 A+A- 46本 ∪ 上财经济学院 Top+First Tier）──────────────
ECONOMICS_JOURNALS = {
    # 交大 A 级（20本）
    "1945-7782": {"name": "American Economic Journal: Applied Economics",               "abbr": "AEJ-APPL ECON",      "tier": "A"},
    "0002-8282": {"name": "American Economic Review",                                   "abbr": "AER",                "tier": "A"},
    "0012-9682": {"name": "Econometrica",                                               "abbr": "ECMA",               "tier": "A"},
    "0020-6598": {"name": "International Economic Review",                              "abbr": "INT ECON REV",        "tier": "A"},
    "0304-3878": {"name": "Journal of Development Economics",                           "abbr": "J DEV ECON",          "tier": "A"},
    "0304-4076": {"name": "Journal of Econometrics",                                    "abbr": "J ECONOMETRICS",      "tier": "A"},
    "0022-0531": {"name": "Journal of Economic Theory",                                 "abbr": "J ECON THEORY",       "tier": "A"},
    "0022-1082": {"name": "Journal of Finance",                                         "abbr": "J FINANC",            "tier": "A"},
    "0022-1090": {"name": "Journal of Financial and Quantitative Analysis",             "abbr": "J FINANC QUANT ANAL", "tier": "A"},
    "0304-405X": {"name": "Journal of Financial Economics",                             "abbr": "J FINANC ECON",       "tier": "A"},
    "0022-1996": {"name": "Journal of International Economics",                         "abbr": "J INT ECON",          "tier": "A"},
    "0734-306X": {"name": "Journal of Labor Economics",                                 "abbr": "J LABOR ECON",        "tier": "A"},
    "0304-3932": {"name": "Journal of Monetary Economics",                              "abbr": "J MONETARY ECON",     "tier": "A"},
    "0022-3808": {"name": "Journal of Political Economy",                               "abbr": "JPE",                 "tier": "A"},
    "0047-2727": {"name": "Journal of Public Economics",                                "abbr": "J PUBLIC ECON",       "tier": "A"},
    "0033-5533": {"name": "Quarterly Journal of Economics",                             "abbr": "QJE",                 "tier": "A"},
    "0741-6261": {"name": "RAND Journal of Economics",                                  "abbr": "RAND J ECON",         "tier": "A"},
    "0034-6527": {"name": "Review of Economic Studies",                                 "abbr": "REStud",              "tier": "A"},
    "0034-6535": {"name": "Review of Economics and Statistics",                         "abbr": "REStat",              "tier": "A"},
    "0893-9454": {"name": "Review of Financial Studies",                                "abbr": "REV FINANC STUD",     "tier": "A"},
    # 交大 A- 级（26本）
    "1945-7731": {"name": "American Economic Journal: Economic Policy",                 "abbr": "AEJ-ECON POLIC",      "tier": "A-"},
    "1945-7707": {"name": "American Economic Journal: Macroeconomics",                  "abbr": "AEJ-MACROECON",       "tier": "A-"},
    "1945-7669": {"name": "American Economic Journal: Microeconomics",                  "abbr": "AEJ-MICROECON",       "tier": "A-"},
    "0266-4666": {"name": "Econometric Theory",                                         "abbr": "ECONOMET THEOR",      "tier": "A-"},
    "0013-0133": {"name": "Economic Journal",                                           "abbr": "ECON J",              "tier": "A-"},
    "0938-2259": {"name": "Economic Theory",                                            "abbr": "ECON THEOR",          "tier": "A-"},
    "1386-4157": {"name": "Experimental Economics",                                     "abbr": "EXP ECON",            "tier": "A-"},
    "0899-8256": {"name": "Games and Economic Behavior",                                "abbr": "GAME ECON BEHAV",     "tier": "A-"},
    "0883-7252": {"name": "Journal of Applied Econometrics",                            "abbr": "J APPL ECONOMET",     "tier": "A-"},
    "0378-4266": {"name": "Journal of Banking & Finance",                               "abbr": "J BANK FINANC",       "tier": "A-"},
    "0735-0015": {"name": "Journal of Business & Economic Statistics",                  "abbr": "J BUS ECON STAT",     "tier": "A-"},
    "0929-1199": {"name": "Journal of Corporate Finance",                               "abbr": "J CORP FINANC",       "tier": "A-"},
    "1381-4338": {"name": "Journal of Economic Growth",                                 "abbr": "J ECON GROWTH",       "tier": "A-"},
    "0095-0696": {"name": "Journal of Environmental Economics and Management",          "abbr": "J ENVIRON ECON MANAG","tier": "A-"},
    "1042-9573": {"name": "Journal of Financial Intermediation",                        "abbr": "J FINANC INTERMED",   "tier": "A-"},
    "1386-4181": {"name": "Journal of Financial Markets",                               "abbr": "J FINANC MARK",       "tier": "A-"},
    "0167-6296": {"name": "Journal of Health Economics",                                "abbr": "J HEALTH ECON",       "tier": "A-"},
    "0022-166X": {"name": "Journal of Human Resources",                                 "abbr": "J HUM RESOUR",        "tier": "A-"},
    "0022-1821": {"name": "Journal of Industrial Economics",                            "abbr": "J IND ECON",          "tier": "A-"},
    "0022-2186": {"name": "Journal of Law & Economics",                                 "abbr": "J LAW ECON",          "tier": "A-"},
    "0022-2879": {"name": "Journal of Money, Credit and Banking",                       "abbr": "J MONEY CREDIT BANK", "tier": "A-"},
    "1542-4766": {"name": "Journal of the European Economic Association",               "abbr": "J EUR ECON ASSOC",    "tier": "A-"},
    "0094-1190": {"name": "Journal of Urban Economics",                                 "abbr": "J URBAN ECON",        "tier": "A-"},
    "0960-1627": {"name": "Mathematical Finance",                                       "abbr": "MATH FINANC",         "tier": "A-"},
    "1094-2025": {"name": "Review of Economic Dynamics",                                "abbr": "REV ECON DYNAM",      "tier": "A-"},
    "0305-750X": {"name": "World Development",                                          "abbr": "WORLD DEV",           "tier": "A-"},
    # 上财 First Tier 补充（不含交大已有的）
    "1759-7323": {"name": "American Economic Journal: Applied Economics",               "abbr": "AEJ-APPL ECON",       "tier": "First"},  # 与 1945-7782 同刊不同 ISSN 版本，保留
    "1554-0626": {"name": "Journal of Economic History",                                "abbr": "J ECON HIST",         "tier": "First"},
    "1759-7331": {"name": "Quantitative Economics",                                     "abbr": "QUANT ECON",          "tier": "First"},
    "1933-6837": {"name": "Theoretical Economics",                                      "abbr": "THEOR ECON",          "tier": "First"},
}


def fetch_management_papers(keyword: str, **kwargs) -> list[Paper]:
    """按管理类期刊（交大安泰 A+A-）抓取"""
    return fetch_english_papers(keyword, journals=MANAGEMENT_JOURNALS, **kwargs)


def fetch_economics_papers(keyword: str, **kwargs) -> list[Paper]:
    """按经济类期刊（交大 A+A- ∪ 上财 First Tier）抓取"""
    return fetch_english_papers(keyword, journals=ECONOMICS_JOURNALS, **kwargs)


def fetch_all_papers(keyword: str, **kwargs) -> list[Paper]:
    """按全部期刊（ENGLISH_JOURNALS ∪ MANAGEMENT_JOURNALS ∪ ECONOMICS_JOURNALS 去重）抓取"""
    all_journals = {**ENGLISH_JOURNALS, **MANAGEMENT_JOURNALS, **ECONOMICS_JOURNALS}
    return fetch_english_papers(keyword, journals=all_journals, **kwargs)


# =============================================================================
# 翻译模块（Claude API 批量翻译）
# =============================================================================

def _call_claude(prompt: str, max_tokens: int = 8192) -> Optional[str]:
    return _call_llm(prompt, max_tokens=max_tokens, kind="fast")


def _is_english_text(text: str) -> bool:
    if not text:
        return False
    return sum(1 for c in text if c.isascii() and c.isalpha()) / max(len(text), 1) > 0.3


def translate_papers(papers: list[Paper], batch_size: int = None) -> None:
    if not ANTHROPIC_API_KEY:
        logger.warning("未配置 ANTHROPIC_API_KEY，跳过翻译步骤")
        return

    to_translate = [
        p for p in papers
        if p.source in ("crossref",) and _is_english_text(p.title) and not p.title_zh
    ]
    if not to_translate:
        logger.info("没有需要翻译的英文论文")
        return

    logger.info(f"开始翻译 {len(to_translate)} 篇英文论文...")

    for i, paper in enumerate(to_translate):
        # 构造单篇翻译 prompt，返回格式固定为两行
        abstract_part = f"\n摘要：{paper.abstract[:1500]}" if paper.abstract else ""
        prompt = (
            f"请将以下经济学论文的英文标题和摘要翻译成中文，严格按如下格式输出，不要任何多余文字：\n"
            f"标题：[中文标题]\n"
            f"摘要：[中文摘要]\n\n"
            f"英文标题：{paper.title}"
            f"{abstract_part}"
        )
        response = _call_claude(prompt, max_tokens=2048)
        if not response:
            logger.warning(f"  第 {i+1} 篇翻译失败，跳过")
            continue

        # 解析固定格式输出
        title_zh = ""
        abstract_zh = ""
        for line in response.strip().splitlines():
            line = line.strip()
            if line.startswith("标题："):
                title_zh = line[3:].strip()
            elif line.startswith("摘要："):
                abstract_zh = line[3:].strip()

        if title_zh:
            paper.title_zh = title_zh
            paper.abstract_zh = abstract_zh
        else:
            # 兜底：整个响应第一行作为标题
            lines = [l.strip() for l in response.strip().splitlines() if l.strip()]
            paper.title_zh = lines[0] if lines else ""

        if (i + 1) % 5 == 0 or (i + 1) == len(to_translate):
            logger.info(f"  翻译进度: {i+1}/{len(to_translate)}")
        time.sleep(0.5)

    logger.info("翻译完成")


# =============================================================================
# Excel 导出模块
# =============================================================================

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

THEME = {
    "header_bg":  "1F3864",
    "header_font": "FFFFFF",
    "en_row_bg":  "EBF3FF",
    "cn_row_bg":  "F0FFF0",
    "alt_en_bg":  "DDEEFF",
    "alt_cn_bg":  "E0FFE0",
    "top5_font":  "C00000",
    "border":     "BFBFBF",
}

COLUMNS = [
    ("序号",       "__idx__",      6),
    ("期刊",       "journal_abbr", 12),
    ("年份",       "year",          7),
    ("引用数",     "citations",     9),
    ("英文标题",   "title",        45),
    ("中文标题",   "title_zh",     40),
    ("作者",       "authors_str",  25),
    ("英文摘要",   "abstract",     60),
    ("中文摘要",   "abstract_zh",  60),
    ("匹配位置",   "match_reason", 16),
    ("DOI / 链接", "url",          35),
    ("来源",       "source",       10),
    ("下载 PDF",   "download_url", 14),
]


def _make_border(color: str = "BFBFBF") -> "Border":
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _row_bg(paper: Paper, idx: int) -> str:
    if paper.source == "cnki":
        return THEME["alt_cn_bg"] if idx % 2 == 0 else THEME["cn_row_bg"]
    return THEME["alt_en_bg"] if idx % 2 == 0 else THEME["en_row_bg"]


def export_to_excel(papers: list[Paper], output_path: str, keyword: str = "") -> str:
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if not papers:
        logger.warning("没有论文数据可导出")
        return ""

    if not HAS_OPENPYXL:
        csv_path = output_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["序号", "期刊", "年份", "引用数", "英文标题", "中文标题", "作者", "英文摘要", "中文摘要", "匹配位置", "链接", "下载PDF"])
            for i, p in enumerate(papers, 1):
                writer.writerow([i, p.journal_abbr, p.year, p.citations, p.title, p.title_zh,
                                  "; ".join(p.authors[:3]), p.abstract, p.abstract_zh, p.match_reason, p.url, p.download_url])
        logger.info(f"CSV 已保存: {csv_path}")
        return csv_path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "论文列表"

    if keyword:
        ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
        tc = ws["A1"]
        tc.value = f'经济学文献检索结果  关键词：{keyword}  共 {len(papers)} 篇'
        tc.font = Font(name="宋体", size=14, bold=True, color="1F3864")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        header_row = 2
    else:
        header_row = 1

    header_fill = PatternFill(fill_type="solid", fgColor=THEME["header_bg"])
    for col_idx, (col_name, _, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = Font(name="宋体", size=11, bold=True, color=THEME["header_font"])
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _make_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[header_row].height = 22

    is_top5_abbrs = {v["abbr"] for v in ENGLISH_JOURNALS.values() if v.get("tier") == "Top5"}

    for pidx, paper in enumerate(papers):
        row_num = header_row + 1 + pidx
        row_fill = PatternFill(fill_type="solid", fgColor=_row_bg(paper, pidx))
        row_data = {
            "__idx__":    pidx + 1,
            "journal_abbr": paper.journal_abbr or paper.journal,
            "year":       paper.year or "",
            "citations":  paper.citations if paper.citations >= 0 else "",
            "title":      paper.title,
            "title_zh":   paper.title_zh,
            "authors_str": "; ".join(paper.authors[:3]) + (" 等" if len(paper.authors) > 3 else ""),
            "abstract":   paper.abstract,
            "abstract_zh": paper.abstract_zh,
            "match_reason": paper.match_reason,
            "url":        paper.url,
            "source":     paper.source,
            "download_url": paper.download_url,
        }
        is_top5 = paper.journal_abbr in is_top5_abbrs
        for col_idx, (_, fname, _) in enumerate(COLUMNS, start=1):
            value = row_data.get(fname, "")
            # 下载 PDF 列：显示"下载"文字，URL 作为超链接
            is_dl = fname == "download_url"
            if is_dl:
                display_value = "⬇ 下载" if value else ""
            else:
                display_value = value
            cell = ws.cell(row=row_num, column=col_idx, value=display_value)
            cell.fill = row_fill
            cell.border = _make_border()
            font_color = THEME["top5_font"] if is_top5 and fname == "journal_abbr" else "000000"
            is_url = fname == "url"
            is_link = is_url or is_dl
            cell.font = Font(name="宋体", size=10,
                             color="0563C1" if is_link else font_color,
                             underline="single" if is_link else None)
            if fname in ("__idx__", "year", "citations"):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif fname in ("match_reason", "download_url"):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif fname in ("abstract", "abstract_zh"):
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if is_url and value:
                cell.hyperlink = value
            if is_dl and value:
                cell.hyperlink = value
        # 按最长换行字段估算行高（中文字符按2计，列宽单位≈字符数）
        def _line_count(text, col_w, char_w=1.1):
            if not text:
                return 1
            # 中文字符宽度约为英文的2倍
            w = sum(2 if ord(c) > 127 else 1 for c in str(text))
            chars_per_line = max(1, int(col_w / char_w))
            return max(1, -(-w // chars_per_line))  # ceiling division
        col_map = {fname: col_w for _, fname, col_w in COLUMNS}
        max_lines = max(
            _line_count(paper.abstract,    col_map.get("abstract",    60)),
            _line_count(paper.abstract_zh, col_map.get("abstract_zh", 60)),
            _line_count(paper.title,       col_map.get("title",       45)),
            _line_count(paper.title_zh,    col_map.get("title_zh",    40)),
        )
        ws.row_dimensions[row_num].height = max(18, min(max_lines * 15, 300))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=2)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{header_row + len(papers)}"

    # 汇总 Sheet
    ws2 = wb.create_sheet("汇总统计")
    journal_counts = Counter(p.journal_abbr or p.journal for p in papers)
    year_counts    = Counter(p.year for p in papers if p.year)
    hdr_fill = PatternFill(fill_type="solid", fgColor=THEME["header_bg"])
    bw = Font(name="宋体", size=11, bold=True, color="FFFFFF")
    ws2["A1"] = "期刊分布"
    ws2["A1"].font = Font(name="宋体", size=12, bold=True, color="1F3864")
    ws2.merge_cells("A1:B1")
    ws2["A2"] = "期刊"; ws2["B2"] = "论文数"
    for c in [ws2["A2"], ws2["B2"]]:
        c.fill = hdr_fill; c.font = bw; c.alignment = Alignment(horizontal="center")
    for i, (jn, cnt) in enumerate(journal_counts.most_common(), start=3):
        ws2.cell(row=i, column=1, value=jn); ws2.cell(row=i, column=2, value=cnt)
    base = 3 + len(journal_counts) + 2
    ws2.cell(row=base, column=1, value="年份分布").font = Font(name="宋体", size=12, bold=True, color="1F3864")
    ws2.merge_cells(f"A{base}:B{base}")
    ws2.cell(row=base+1, column=1, value="年份").fill = hdr_fill
    ws2.cell(row=base+1, column=1).font = bw
    ws2.cell(row=base+1, column=2, value="论文数").fill = hdr_fill
    ws2.cell(row=base+1, column=2).font = bw
    for i, (yr, cnt) in enumerate(sorted(year_counts.items(), reverse=True), start=base+2):
        ws2.cell(row=i, column=1, value=yr); ws2.cell(row=i, column=2, value=cnt)
    ws2.column_dimensions["A"].width = 20; ws2.column_dimensions["B"].width = 12

    wb.save(output_path)
    logger.info(f"Excel 已保存: {output_path}")
    return output_path


# =============================================================================
# 文献综述生成模块
# =============================================================================

def _call_claude_strong(prompt: str, max_tokens: int = 8192) -> str:
    if not ANTHROPIC_API_KEY:
        return "[错误] 未配置 API Key"
    result = _call_llm(prompt, max_tokens=max_tokens, kind="strong")
    return result if result else "[错误] 调用失败"


def _format_paper_list(papers: list[Paper], max_papers: int = 40) -> str:
    lines = []
    for i, p in enumerate(papers[:max_papers], 1):
        en_title = p.title
        zh_title = f"（{p.title_zh}）" if p.title_zh and p.source != "cnki" else ""
        lang_tag = "【中文】" if p.source == "cnki" else ""
        abstract = p.abstract_zh if p.abstract_zh else p.abstract
        abstract_short = (abstract[:300] + "...") if len(abstract) > 300 else abstract
        authors_str = ", ".join(p.authors[:3]) + (" et al." if len(p.authors) > 3 else "")
        cite_str = f"，被引 {p.citations} 次" if p.citations > 0 else ""
        lines.append(
            f"[{i}]{lang_tag} {en_title}{zh_title}\n"
            f"    来源：{p.journal}（{p.year}）{cite_str}\n"
            f"    作者：{authors_str}\n"
            f"    摘要：{abstract_short}"
        )
    return "\n\n".join(lines)


def generate_review(
    papers: list[Paper], keyword: str,
    output_path: str = "", style: str = "academic", language: str = "zh",
) -> str:
    """
    生成文献综述。

    当文献数量 <= 40 篇时，直接调用一次 LLM 生成完整综述。
    当文献数量 > 40 篇时，采用两阶段生成：
      阶段一：每批 40 篇独立提炼核心发现（各批约 500-800 字摘要），多批并行。
      阶段二：将所有批次的摘要合并，调用一次 LLM 生成完整综述。
    这样可以避免单次 prompt 过长导致中转站超时断连。
    """
    BATCH_SIZE = 40  # 每批最大文献数，超过此数启用两阶段生成

    if not papers:
        return "没有可用的论文数据，无法生成综述"

    total = len(papers)

    # ── 阶段一：分批提炼（仅在文献超过 BATCH_SIZE 时执行）────────────────────
    if total > BATCH_SIZE:
        batches = [papers[i:i + BATCH_SIZE] for i in range(0, total, BATCH_SIZE)]
        n_batches = len(batches)
        logger.info(f"文献共 {total} 篇，超过 {BATCH_SIZE} 篇上限，启用两阶段生成（共 {n_batches} 批）...")

        batch_summaries = []
        for idx, batch in enumerate(batches, 1):
            batch_list_str = _format_paper_list(batch, max_papers=len(batch))
            start_no = (idx - 1) * BATCH_SIZE + 1
            end_no   = start_no + len(batch) - 1
            logger.info(f"  阶段一：提炼第 {idx}/{n_batches} 批（第 {start_no}-{end_no} 篇）...")

            if language == "zh":
                batch_prompt = f"""你是一位专业的经济学研究者。以下是关于"{keyword}"的 {len(batch)} 篇经济学文献（编号 {start_no}-{end_no}，共 {total} 篇中的一批）。

请对这批文献进行结构化提炼，输出约 600-900 字，包含：
1. **核心研究问题**：这批文献聚焦哪些研究问题？
2. **主要发现**：列出 3-5 个最重要的实证发现或理论贡献，引用对应文献编号 [序号]
3. **研究方法**：这批文献使用了哪些主要方法或数据？
4. **争议或分歧**：这批文献之间是否存在不同观点？

要求：只基于所给文献，不要捏造；引用时使用 [序号] 标注。

文献列表：
{batch_list_str}

请输出提炼结果："""
            else:
                batch_prompt = f"""You are an expert economist. Below are {len(batch)} papers (nos. {start_no}-{end_no} of {total} total) on "{keyword}".

Extract and summarize (~600 words):
1. **Core research questions** covered by this batch
2. **Key findings** (cite 3-5 most important, use [N] format)
3. **Methods and data** used
4. **Disagreements or debates** among these papers

Only use the given papers. Cite as [N].

Papers:
{batch_list_str}

Output:"""

            summary = _call_claude_strong(batch_prompt, max_tokens=2048)
            if not summary or summary.startswith("[错误]"):
                logger.warning(f"  第 {idx} 批提炼失败，将跳过该批次")
                summary = f"（第 {idx} 批文献提炼失败，编号 {start_no}-{end_no}）"
            batch_summaries.append(f"【第 {idx} 批，文献 {start_no}-{end_no}】\n{summary}")

        combined_summaries = "\n\n" + "─" * 40 + "\n\n".join(batch_summaries)
        logger.info(f"  阶段一完成，进入阶段二：整合 {n_batches} 批提炼结果生成完整综述...")

        # ── 阶段二：基于所有批次摘要生成完整综述 ────────────────────────────
        if language == "zh":
            if style == "academic":
                final_prompt = f"""你是一位专业的经济学研究者。以下是关于"{keyword}"的 {total} 篇经济学文献的分批提炼结果（共 {n_batches} 批，每批约 40 篇）。

请综合所有批次的提炼内容，撰写一篇完整的学术文献综述。

## 综述要求

**结构**：
1. **引言**（100-200字）：介绍该领域研究背景和综述目的
2. **研究脉络梳理**（400-600字）：按研究问题/方法/时间线梳理文献发展脉络，说明各阶段代表性研究
3. **主要发现与争议**（400-600字）：总结主要研究发现，指出学界共识和存在分歧的议题
4. **研究方法综述**（200-300字）：梳理该领域常用的研究方法和数据
5. **研究前沿与展望**（200-300字）：指出当前研究局限和未来可能的研究方向
6. **参考文献**：列出引用的论文，英文文献**必须使用英文原版标题**（格式：Author et al. (Year). *Title*. *Journal*.），中文文献用中文原题目

**要求**：
- 综述约 1500-2500 字
- 引用具体文献时使用 [序号] 格式标注（如 [1][3]）
- 跨批次综合归纳，不要按批次分段罗列
- 语言严谨，符合中文学术写作规范
- 不要凭空捏造提炼内容中没有提到的研究

## 分批提炼结果

{combined_summaries}

请撰写完整综述："""
            else:
                final_prompt = f"""请根据以下关于"{keyword}"的 {total} 篇经济学文献的分批提炼结果，写一份简要的研究概述（约 500-800 字）。

概述要求：
1. 说明这个领域的核心研究问题
2. 总结主要研究结论（引用代表性文献编号）
3. 指出研究空白或争议点

分批提炼结果：
{combined_summaries}

请写概述："""
        else:
            final_prompt = f"""You are an expert economist. Below are batch summaries of {total} papers (in {n_batches} batches) on "{keyword}".

Write a complete academic literature review based on all batches.

Structure:
1. Introduction (100-150 words)
2. Research Streams (400-500 words)
3. Key Findings and Debates (400-500 words)
4. Methodology Overview (150-200 words)
5. Research Gaps and Future Directions (150-200 words)
6. References

Requirements: ~1500-2000 words, cite as [N], synthesize across batches (do not list by batch).

Batch summaries:
{combined_summaries}

Write the review:"""

        review_text = _call_claude_strong(final_prompt, max_tokens=8192)
        displayed = total

    # ── 文献 <= 40 篇：直接一次生成 ──────────────────────────────────────────
    else:
        paper_list_str = _format_paper_list(papers, max_papers=total)
        displayed = total

        if language == "zh":
            if style == "academic":
                prompt = f"""你是一位专业的经济学研究者。请根据以下 {displayed} 篇关于"{keyword}"的经济学文献，撰写一篇学术文献综述。

文献列表中标注【中文】的为中文期刊文献，其余为英文文献。请综合中英文研究成果，在综述中平等引用两类文献的核心观点。

## 综述要求

**结构**：
1. **引言**（100-200字）：介绍该领域研究背景和综述目的
2. **研究脉络梳理**（400-600字）：按研究问题/方法/时间线梳理文献发展脉络，说明各阶段代表性研究
3. **主要发现与争议**（400-600字）：总结主要研究发现，指出学界共识和存在分歧的议题
4. **研究方法综述**（200-300字）：梳理该领域常用的研究方法和数据
5. **研究前沿与展望**（200-300字）：指出当前研究局限和未来可能的研究方向
6. **参考文献**：列出引用的论文，英文文献**必须使用英文原版标题**（格式：Author et al. (Year). *Title*. *Journal*.），中文文献用中文原题目

**要求**：
- 综述约 1500-2500 字
- 引用具体文献时使用 [序号] 格式标注（如 [1][3]）
- 语言严谨，符合中文学术写作规范
- 不要凭空捏造文献中没有提到的研究

## 文献列表

{paper_list_str}

请撰写综述："""
            else:
                prompt = f"""请根据以下关于"{keyword}"的 {displayed} 篇经济学文献，写一份简要的研究概述（约 500-800 字）。

文献列表中标注【中文】的为中文期刊文献，请综合引用中英文文献的核心观点。

概述要求：
1. 说明这个领域的核心研究问题
2. 总结主要研究结论（引用 3-5 篇最重要的文献）
3. 指出研究空白或争议点

文献列表：
{paper_list_str}

请写概述："""
        else:
            prompt = f"""You are an expert economist. Based on the following {displayed} papers on "{keyword}", write an academic literature review.

Structure:
1. Introduction (100-150 words)
2. Research Streams (400-500 words)
3. Key Findings and Debates (400-500 words)
4. Methodology Overview (150-200 words)
5. Research Gaps and Future Directions (150-200 words)
6. References

Requirements: ~1500-2000 words total, cite papers as [N] inline.

Papers:
{paper_list_str}

Write the review:"""

        logger.info(f"调用 Claude 生成综述（共 {displayed} 篇文献）...")
        review_text = _call_claude_strong(prompt, max_tokens=8192)

    # ── 保存文件 ──────────────────────────────────────────────────────────────
    if output_path and review_text and not review_text.startswith("[错误]"):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(f"文献综述：{keyword}\n")
            f.write("=" * 60 + "\n")
            if total > BATCH_SIZE:
                f.write(f"共 {total} 篇文献，采用两阶段生成（每批 {BATCH_SIZE} 篇）\n")
            else:
                f.write(f"共 {total} 篇文献\n")
            f.write("=" * 60 + "\n\n")
            f.write(review_text)
        logger.info(f"综述已保存: {output_path}")

    return review_text


_MANUAL_HTML = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PaperTracker 使用说明书</title>
<style>
  :root {
    --blue:   #1565C0;
    --green:  #2E7D32;
    --orange: #E65100;
    --purple: #6A1B9A;
    --red:    #B71C1C;
    --gray:   #424242;
    --bg:     #FAFAFA;
    --card:   #FFFFFF;
    --border: #E0E0E0;
    --code-bg:#F5F5F5;
    --tag-bg: #E3F2FD;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'SimSun', 'Times New Roman', serif;
    font-size: 15px;
    background: var(--bg);
    color: #212121;
    line-height: 1.8;
  }
  /* ─── 导航侧栏 ─── */
  nav {
    position: fixed;
    top: 0; left: 0;
    width: 240px; height: 100vh;
    background: #1A237E;
    color: #fff;
    overflow-y: auto;
    padding: 24px 0 40px;
    z-index: 100;
  }
  nav .nav-title {
    font-size: 14px;
    font-weight: bold;
    letter-spacing: 1px;
    color: #90CAF9;
    padding: 0 20px 12px;
    border-bottom: 1px solid #283593;
    margin-bottom: 8px;
  }
  nav a {
    display: block;
    padding: 7px 20px;
    color: #CFD8DC;
    text-decoration: none;
    font-size: 13px;
    transition: background .15s;
  }
  nav a:hover { background: #283593; color: #fff; }
  nav a.h2-link { font-weight: bold; color: #E3F2FD; margin-top: 6px; }
  nav a.h3-link { padding-left: 32px; }
  /* ─── 主内容 ─── */
  main {
    margin-left: 240px;
    padding: 40px 56px 80px;
    max-width: 1000px;
  }
  h1 {
    font-size: 28px;
    color: var(--blue);
    border-bottom: 3px solid var(--blue);
    padding-bottom: 12px;
    margin-bottom: 8px;
  }
  .subtitle {
    color: #78909C;
    font-size: 13px;
    margin-bottom: 40px;
  }
  h2 {
    font-size: 20px;
    color: var(--blue);
    margin: 48px 0 16px;
    padding: 8px 16px;
    border-left: 5px solid var(--blue);
    background: #E8EAF6;
    border-radius: 0 6px 6px 0;
  }
  h3 {
    font-size: 16px;
    color: var(--gray);
    margin: 28px 0 10px;
    border-bottom: 1px solid var(--border);
    padding-bottom: 4px;
  }
  h4 {
    font-size: 14px;
    color: var(--gray);
    margin: 18px 0 6px;
    font-weight: bold;
  }
  p { margin: 8px 0; }
  ul, ol { padding-left: 22px; margin: 8px 0; }
  li { margin: 4px 0; }
  /* ─── 卡片 ─── */
  .card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 20px 24px;
    margin: 16px 0;
    box-shadow: 0 1px 3px rgba(0,0,0,.06);
  }
  /* ─── 代码块 ─── */
  pre {
    background: #263238;
    color: #CFD8DC;
    border-radius: 6px;
    padding: 16px 20px;
    overflow-x: auto;
    margin: 12px 0;
    font-family: 'Consolas', monospace;
    font-size: 13px;
    line-height: 1.6;
  }
  code {
    background: var(--code-bg);
    padding: 1px 5px;
    border-radius: 3px;
    font-family: 'Consolas', monospace;
    font-size: 13px;
    color: var(--red);
  }
  pre code { background: none; color: inherit; padding: 0; }
  /* ─── 表格 ─── */
  table {
    width: 100%;
    border-collapse: collapse;
    margin: 12px 0;
    font-size: 14px;
  }
  th {
    background: var(--blue);
    color: #fff;
    padding: 9px 12px;
    text-align: left;
    font-weight: normal;
  }
  td { padding: 8px 12px; border-bottom: 1px solid var(--border); }
  tr:nth-child(even) td { background: #F5F5F5; }
  /* ─── 标签 ─── */
  .tag {
    display: inline-block;
    padding: 2px 10px;
    border-radius: 12px;
    font-size: 12px;
    font-weight: bold;
    margin: 2px;
  }
  .tag-blue   { background: #E3F2FD; color: #1565C0; }
  .tag-green  { background: #E8F5E9; color: #2E7D32; }
  .tag-orange { background: #FFF3E0; color: #E65100; }
  .tag-red    { background: #FFEBEE; color: #B71C1C; }
  .tag-purple { background: #F3E5F5; color: #6A1B9A; }
  /* ─── 流程图 ─── */
  .flow {
    display: flex;
    align-items: center;
    flex-wrap: wrap;
    gap: 4px;
    margin: 16px 0;
  }
  .flow-box {
    background: #E3F2FD;
    border: 1.5px solid #1565C0;
    border-radius: 6px;
    padding: 8px 16px;
    font-size: 13px;
    color: #1565C0;
    font-weight: bold;
  }
  .flow-arrow {
    color: #90A4AE;
    font-size: 18px;
    font-weight: bold;
  }
  /* ─── 警告框 ─── */
  .notice {
    border-radius: 6px;
    padding: 12px 16px;
    margin: 12px 0;
    font-size: 14px;
  }
  .notice-info   { background:#E3F2FD; border-left: 4px solid #1565C0; }
  .notice-warn   { background:#FFF8E1; border-left: 4px solid #F9A825; }
  .notice-danger { background:#FFEBEE; border-left: 4px solid #C62828; }
  .notice-tip    { background:#E8F5E9; border-left: 4px solid #2E7D32; }
  /* ─── 章节锚点偏移（避免被固定导航遮挡）─── */
  [id] { scroll-margin-top: 24px; }
  @media print {
    nav { display: none; }
    main { margin-left: 0; }
  }
</style>
</head>
<body>

<!-- ════════════════════════════════════════════
     导航侧栏
════════════════════════════════════════════ -->
<nav>
  <div class="nav-title">目录</div>
  <a class="h2-link" href="#overview">一、软件概述</a>
  <a class="h3-link" href="#what">是什么</a>
  <a class="h3-link" href="#arch">技术架构</a>

  <a class="h2-link" href="#start">二、快速上手</a>
  <a class="h3-link" href="#install">安装与启动</a>
  <a class="h3-link" href="#api-setup">配置 API</a>
  <a class="h3-link" href="#first-search">第一次搜索</a>

  <a class="h2-link" href="#ui">三、界面说明</a>
  <a class="h3-link" href="#ui-keyword">关键词输入区</a>
  <a class="h3-link" href="#ui-en">英文期刊设置</a>
  <a class="h3-link" href="#ui-cn">中文期刊设置</a>
  <a class="h3-link" href="#ui-model">模型设置</a>
  <a class="h3-link" href="#ui-filter">篇关摘过滤</a>
  <a class="h3-link" href="#ui-preview">结果预览</a>

  <a class="h2-link" href="#workflow">四、工作原理</a>
  <a class="h3-link" href="#wf-kw">关键词处理</a>
  <a class="h3-link" href="#wf-en">英文文献抓取</a>
  <a class="h3-link" href="#wf-abstract">摘要补全</a>
  <a class="h3-link" href="#wf-cn">中文文献抓取</a>
  <a class="h3-link" href="#wf-filter">篇关摘过滤</a>
  <a class="h3-link" href="#wf-translate">翻译</a>
  <a class="h3-link" href="#wf-review">综述生成</a>
  <a class="h3-link" href="#wf-export">导出</a>

  <a class="h2-link" href="#journals">五、收录期刊</a>
  <a class="h3-link" href="#j-en">英文期刊</a>
  <a class="h3-link" href="#j-cn">中文期刊</a>

  <a class="h2-link" href="#bool">六、布尔检索语法</a>
  <a class="h3-link" href="#bool-ops">运算符</a>
  <a class="h3-link" href="#bool-example">示例</a>

  <a class="h2-link" href="#output">七、输出文件</a>
  <a class="h3-link" href="#out-excel">Excel 说明</a>
  <a class="h3-link" href="#out-review">综述说明</a>

  <a class="h2-link" href="#faq">八、常见问题</a>
  <a class="h2-link" href="#devlog">九、版本历史</a>
</nav>

<!-- ════════════════════════════════════════════
     主内容
════════════════════════════════════════════ -->
<main>

<h1>PaperTracker 使用说明书</h1>
<div class="subtitle">经济学论文追踪系统 · 版本 2026-04-08</div>

<!-- ════ 一、软件概述 ════ -->
<h2 id="overview">一、软件概述</h2>

<h3 id="what">是什么</h3>
<p>PaperTracker 是一款面向经济学研究者的论文检索与管理桌面工具。只需输入研究关键词，软件可自动完成以下工作流：</p>
<div class="flow">
  <div class="flow-box">输入关键词</div>
  <div class="flow-arrow">→</div>
  <div class="flow-box">搜索顶刊</div>
  <div class="flow-arrow">→</div>
  <div class="flow-box">补全摘要</div>
  <div class="flow-arrow">→</div>
  <div class="flow-box">AI翻译</div>
  <div class="flow-arrow">→</div>
  <div class="flow-box">导出Excel</div>
  <div class="flow-arrow">→</div>
  <div class="flow-box">生成综述</div>
</div>
<p>软件覆盖 <strong>英文顶刊 100+ 本</strong>（含 Top5、交大安泰经济类/管理类 A 级以上）和 <strong>中文顶刊 24 本</strong>（通过知网高级检索）。</p>
<p>软件打包为单一 <code>PaperTracker.exe</code>，无需安装 Python 环境，双击即用。</p>

<h3 id="arch">技术架构</h3>
<div class="card">
  <table>
    <tr><th>功能层</th><th>技术方案</th><th>说明</th></tr>
    <tr><td>GUI 界面</td><td>PyQt6</td><td>窗口、控件、布局；深色日志区、结果表格</td></tr>
    <tr><td>英文文献抓取</td><td>CrossRef 公开 API</td><td>按 ISSN + 关键词全文检索，免费无需注册</td></tr>
    <tr><td>引用数 + 摘要补全</td><td>OpenAlex → Semantic Scholar → 网页爬取</td><td>三级瀑布式补全，多线程并发</td></tr>
    <tr><td>中文文献抓取</td><td>QWebEngineView（内嵌 Chromium）</td><td>真实浏览器登录知网，状态机驱动翻页</td></tr>
    <tr><td>AI 翻译 / 综述</td><td>Claude API / OpenAI 兼容接口</td><td>翻译用轻量模型（haiku），综述用强力模型（sonnet）</td></tr>
    <tr><td>Excel 导出</td><td>openpyxl</td><td>格式化、超链接、冻结窗格、行高自适应</td></tr>
    <tr><td>打包</td><td>PyInstaller</td><td>单文件 exe，内含所有依赖</td></tr>
  </table>
</div>

<!-- ════ 二、快速上手 ════ -->
<h2 id="start">二、快速上手</h2>

<h3 id="install">安装与启动</h3>
<p><strong>方式一（推荐）：直接运行 exe</strong></p>
<pre><code>dist\PaperTracker.exe</code></pre>
<p>双击即可启动，无需安装任何环境。</p>
<p><strong>方式二：从源代码运行</strong></p>
<pre><code>pip install PyQt6 PyQt6-WebEngine openpyxl anthropic
python paper_tracker.py</code></pre>

<div class="notice notice-warn">
  <strong>注意</strong>：使用知网（中文文献）功能时，需要额外安装 <code>PyQt6-WebEngine</code>，且版本须与 <code>PyQt6</code> 一致（如均为 6.11.0）。
</div>

<h3 id="api-setup">配置 API</h3>
<p>软件翻译和综述功能依赖大语言模型 API。在主界面「模型设置」区填写：</p>

<div class="card">
  <table>
    <tr><th>字段</th><th>说明</th><th>示例</th></tr>
    <tr><td>接口类型</td><td>选择 Anthropic（Claude）或 OpenAI 兼容</td><td>Anthropic（Claude）</td></tr>
    <tr><td>API Key</td><td>模型服务商提供的密钥</td><td><code>sk-ant-xxxxxxxx</code></td></tr>
    <tr><td>API 地址</td><td>原生留空；中转站填入完整 URL</td><td><code>https://xxx.com</code></td></tr>
    <tr><td>翻译模型</td><td>标题/摘要翻译用的轻量模型</td><td><code>claude-haiku-4-5-20251001</code></td></tr>
    <tr><td>综述模型</td><td>文献综述生成用的强力模型</td><td><code>claude-sonnet-4-6</code></td></tr>
  </table>
</div>

<p>所有设置会自动保存到 <code>paper_tracker_settings.json</code>，下次启动自动加载。</p>
<div class="notice notice-tip">
  <strong>使用 DeepSeek / 豆包 / Gemini</strong>：将接口类型设为「OpenAI 兼容」，填入对应的 API Key、Base URL（如 <code>https://api.deepseek.com</code>）和模型名（如 <code>deepseek-chat</code>）即可。
</div>

<h3 id="first-search">第一次搜索</h3>
<ol>
  <li>在顶部关键词框中输入研究词（如 <code>minimum wage</code>），按 <kbd>Enter</kbd> 添加</li>
  <li>在「英文期刊搜索设置」中选择期刊范围（如「Top5」）、年限（如「最近3年」）</li>
  <li>确认「模型设置」中填写了有效的 API Key</li>
  <li>点击蓝色「开始抓取」按钮</li>
  <li>观察右侧日志区进度，完成后在「结果预览」标签页查看论文列表</li>
  <li>结果自动导出到 <code>D:\claude\research\results\tables\YYYYMMDD\</code> 目录</li>
</ol>

<!-- ════ 三、界面说明 ════ -->
<h2 id="ui">三、界面说明</h2>

<h3 id="ui-keyword">关键词输入区</h3>
<div class="card">
  <p>位于窗口顶部，分为两个部分：</p>
  <ul>
    <li><strong>输入框</strong>：输入关键词后按 <kbd>Enter</kbd> 或逗号添加为「词条」。支持中文（自动翻译为英文学术词后检索）和英文。</li>
    <li><strong>词条（Chip）区</strong>：已添加的关键词以彩色标签展示，点击「×」删除。</li>
    <li><strong>运算符按钮</strong>：输入框下方提供 5 个彩色运算符按钮（<span class="tag tag-blue">*</span><span class="tag tag-green">+</span><span class="tag tag-orange">-</span><span class="tag tag-purple">(</span><span class="tag tag-purple">)</span>），点击将符号插入到光标处。</li>
  </ul>
  <p>输入框实时高亮运算符，与按钮配色一致。</p>
</div>

<div class="notice notice-info">
  <strong>多关键词行为</strong>：添加多个词条时，软件会<strong>逐个词条</strong>独立执行一次完整的搜索流程，各自输出独立的 Excel 和综述文件。
</div>

<h3 id="ui-en">英文期刊搜索设置</h3>
<div class="card">
  <table>
    <tr><th>控件</th><th>说明</th></tr>
    <tr><td>期刊范围</td><td>全部期刊 / 仅 Top5 / 经济类 / 管理类 / 仅英文顶刊（22本）</td></tr>
    <tr><td>年限</td><td>最近 1 / 3 / 5 / 10 年</td></tr>
    <tr><td>搜索模式</td><td>最新+高引 / 仅最新 / 仅高引</td></tr>
    <tr><td>最大数量</td><td>每个期刊最多抓取的论文篇数上限</td></tr>
    <tr><td>翻译标题/摘要</td><td>勾选后调用 AI 翻译中文，不勾选则直接导出英文</td></tr>
  </table>
</div>

<h3 id="ui-cn">中文期刊搜索设置</h3>
<div class="card">
  <table>
    <tr><th>控件</th><th>说明</th></tr>
    <tr><td>搜索中文文献（知网）</td><td>启用知网抓取功能（需已安装 PyQt6-WebEngine）</td></tr>
    <tr><td>抓取数量</td><td>每本期刊最多抓取的论文篇数（可自由填写）</td></tr>
    <tr><td>排序方式</td><td>发表时间 / 被引次数 / 相关度 / 下载次数 / 综合</td></tr>
    <tr><td>选择中文期刊</td><td>点击弹出多选对话框，可逐本勾选；勾选状态跨次启动保留</td></tr>
    <tr><td>知网登录</td><td>点击弹出内嵌浏览器，在校园网/VPN 环境中手动登录，Cookie 自动持久化</td></tr>
  </table>
  <p>「选择期刊」按钮显示当前勾选数量（如「选择期刊 (20/24)」）。</p>
</div>

<h3 id="ui-model">模型设置</h3>
<p>见<a href="#api-setup">配置 API</a>节。翻译模型和综述模型均支持下拉选择预置选项，也可手动输入自定义模型名。</p>

<h3 id="ui-filter">篇关摘过滤</h3>
<p>位于「开始抓取」按钮左侧：</p>
<div class="card">
  <ul>
    <li><strong>篇关摘过滤</strong>（蓝色开关）：启用后，只保留检索词出现在指定字段的论文。</li>
    <li><strong>篇名 / 关键词 / 摘要</strong>（绿色字段选项）：选择要匹配哪些字段。字段选项仅在开关启用时可点。</li>
  </ul>
  <p><strong>对中文文献的处理</strong>：知网检索本身已在服务端完成主题词匹配，程序不对知网结果进行二次过滤，但仍会标注「匹配位置」列。</p>
</div>

<h3 id="ui-preview">结果预览</h3>
<p>搜索完成后切换到「结果预览」标签页，展示全部返回论文（无篇数上限）。可在此选中知网来源的行后点击「⬇ 下载选中论文 PDF（知网）」批量下载全文。</p>

<!-- ════ 四、工作原理 ════ -->
<h2 id="workflow">四、工作原理</h2>

<h3 id="wf-kw">1. 关键词处理</h3>
<p>用户输入的关键词首先经过语言检测：若含有中文字符（CJK Unicode），则调用 LLM（轻量模型）将其翻译为对应的英文学术词，再用英文词去检索英文期刊；原始中文词直接用于知网检索。</p>
<div class="card">
  <strong>示例</strong>：输入「创新」→ LLM 翻译为 <code>innovation</code> → 用 <code>innovation</code> 检索 CrossRef；知网检索仍用「创新」作为主题词。
</div>

<h3 id="wf-en">2. 英文文献抓取（CrossRef）</h3>
<p>核心函数：<code>_crossref_search_journal()</code></p>
<p>对每本收录期刊，构造 CrossRef API 请求：</p>
<pre><code>GET https://api.crossref.org/works
  ?query={关键词}
  &filter=has-issn:true,issn:{ISSN},from-pub-date:{年份起}
  &select=DOI,title,abstract,author,published,container-title,subject
  &sort=relevance|published
  &rows={最大数量}
  &mailto={邮箱}</code></pre>

<div class="notice notice-info">
  <strong>NBER Working Paper 特殊处理</strong>：NBER 没有标准 ISSN，改用 CrossRef 机构 ID：<code>filter=member:1960,type:report,from-pub-date:{年份}</code>。
</div>

<p>所有期刊的请求通过 <code>ThreadPoolExecutor</code> 多线程并发发出，加快抓取速度。</p>

<p><strong>搜索模式</strong>对抓取行为的影响：</p>
<ul>
  <li><span class="tag tag-blue">最新+高引</span>：先抓 <code>sort=published</code> 最近若干篇，再抓 <code>sort=relevance</code>（相关度高→引用数高）若干篇，合并去重</li>
  <li><span class="tag tag-green">仅最新</span>：仅按发表时间倒序</li>
  <li><span class="tag tag-orange">仅高引</span>：仅按相关度（在 CrossRef 侧代理高引用）</li>
</ul>

<h3 id="wf-abstract">3. 摘要补全（三级瀑布）</h3>
<p>CrossRef 原始数据中约 60–80% 的论文缺少摘要。程序采用三级依次补全策略：</p>

<div class="flow" style="flex-direction:column; align-items:flex-start; gap:8px;">
  <div style="display:flex; align-items:center; gap:8px;">
    <div class="flow-box" style="width:180px; text-align:center;">① OpenAlex API</div>
    <span>批量 DOI 查询，倒排索引还原摘要。覆盖率约 75–85%。<strong>8 线程并发</strong>，处理千篇约 10 秒。</span>
  </div>
  <div style="display:flex; align-items:center; gap:8px;">
    <div class="flow-box" style="width:180px; text-align:center;">② Semantic Scholar</div>
    <span>小批次 25 条 + 指数退避重试（应对 HTTP 429 限流）。补充 OpenAlex 未能覆盖的部分。</span>
  </div>
  <div style="display:flex; align-items:center; gap:8px;">
    <div class="flow-box" style="width:180px; text-align:center;">③ 网页爬取</div>
    <span>直接爬取 DOI 对应论文页面，提取 <code>&lt;meta name="citation_abstract"&gt;</code> 等标签。<strong>10 线程并发</strong>。特别处理 Elsevier 两级跳转（doi.org → linkinghub → ScienceDirect）。</span>
  </div>
</div>

<p>三个阶段完成后，日志仅显示最终仍缺失摘要的篇数（「无公开摘要」），不打印逐篇 warning。</p>

<h3 id="wf-cn">4. 中文文献抓取（知网）</h3>
<p>核心组件：<code>CnkiManager(QObject)</code></p>

<p>由于知网是 SPA（单页应用），需要真实浏览器环境。程序内嵌 <strong>QWebEngineView（Chromium 内核）</strong>，通过状态机驱动整个搜索流程：</p>

<div class="card">
  <p><strong>状态机流转：</strong></p>
  <div class="flow" style="flex-wrap:wrap;">
    <div class="flow-box">idle</div>
    <div class="flow-arrow">→</div>
    <div class="flow-box">navigating</div>
    <div class="flow-arrow">→</div>
    <div class="flow-box">filling</div>
    <div class="flow-arrow">→</div>
    <div class="flow-box">filling_submitted</div>
    <div class="flow-arrow">→</div>
    <div class="flow-box">polling</div>
    <div class="flow-arrow">→</div>
    <div class="flow-box">abs_loading</div>
    <div class="flow-arrow">→</div>
    <div class="flow-box">idle</div>
  </div>
  <ul style="margin-top:12px;">
    <li><strong>navigating</strong>：导航至知网高级检索页 (<code>kns8s/AdvSearch</code>)，等待 <code>loadFinished</code></li>
    <li><strong>filling</strong>：注入 JS 填写「主题词」（<code>gradetxt-1</code>）和「文献来源」（<code>gradetxt-3</code>，填入 <code>期刊A OR 期刊B OR ...</code>）并提交</li>
    <li><strong>filling_submitted</strong>：等待搜索结果页加载完成</li>
    <li><strong>polling</strong>：每 600ms 注入 JS 轮询结果列表，提取论文信息；同时应用年份过滤（自动勾选结果页左侧的年度复选框）和排序切换</li>
    <li><strong>abs_loading</strong>：对每篇论文逐一注入 JS 点击详情链接，等待详情页加载后提取摘要和下载链接</li>
  </ul>
</div>

<p>程序使用 <code>QWebEngineProfile("cnki_session")</code> 持久化 Cookie，<strong>一次登录，跨启动复用</strong>。</p>
<p><strong>篇关摘过滤对知网的处理</strong>：知网的「主题词」检索本身已覆盖篇名+关键词+摘要三字段，程序信任知网服务端的匹配结果，不对知网返回论文进行本地二次过滤，但仍标注「匹配位置」列（若摘要缺失显示「摘要缺失（知网已匹配）」）。</p>

<h3 id="wf-filter">5. 篇关摘过滤</h3>
<p>仅作用于英文文献，在摘要补全完成后、翻译开始前执行：</p>
<ol>
  <li>将用户勾选字段（篇名 / 关键词 / 摘要）的文本拼合为整体</li>
  <li>对整体文本求值布尔表达式（详见<a href="#bool">布尔检索语法</a>节）</li>
  <li>通过过滤的论文保留；未通过的丢弃（日志显示丢弃数量）</li>
  <li>无论是否开启过滤，<strong>始终</strong>计算并填写 Excel「匹配位置」列</li>
</ol>

<table>
  <tr><th>匹配位置值</th><th>含义</th></tr>
  <tr><td>篇名</td><td>检索词出现在标题中</td></tr>
  <tr><td>关键词</td><td>检索词出现在论文关键词中</td></tr>
  <tr><td>摘要</td><td>检索词出现在摘要中</td></tr>
  <tr><td>篇名、摘要</td><td>两处均出现</td></tr>
  <tr><td>摘要缺失</td><td>该论文无公开摘要，无法判断</td></tr>
  <tr><td>词项未出现</td><td>有内容但不含检索词（未启用过滤时可能保留）</td></tr>
  <tr><td>知网主题词已匹配</td><td>知网文献，服务端已确认相关</td></tr>
</table>

<h3 id="wf-translate">6. 翻译</h3>
<p>核心函数：<code>_call_claude()</code>（内部调用 <code>_call_llm()</code>）</p>
<p>采用<strong>逐篇翻译</strong>方式，对每篇论文发送独立请求，prompt 格式固定：</p>
<pre><code>请将以下经济学论文的标题和摘要翻译成中文，保持学术语言风格。
英文标题：{title}
英文摘要：{abstract}
请直接输出：
标题：[中文标题]
摘要：[中文摘要]</code></pre>
<p>不依赖 JSON 解析，任意一篇翻译失败不影响其他篇。使用<strong>轻量模型</strong>（默认 claude-haiku-4-5-20251001），速度快、成本低。</p>

<h3 id="wf-review">7. 综述生成</h3>
<p>核心函数：<code>generate_review()</code>，使用<strong>强力模型</strong>（默认 claude-sonnet-4-6）。</p>
<p>以筛选后的全部论文作为输入，每篇送入：英文标题（+中文译名）、期刊+年份+被引次数、作者（≤3人+et al.）、摘要前 300 字。</p>

<div class="card">
  <p><strong>两阶段生成机制（论文数 &gt; 40 篇时自动启用）</strong></p>
  <ul>
    <li><strong>阶段一</strong>：将文献按每批 40 篇分组，每批独立调用 LLM 提炼核心发现（约 600–900 字/批，<code>max_tokens=2048</code>），某批失败时跳过</li>
    <li><strong>阶段二</strong>：将全部批次的提炼结果拼合后，一次调用 LLM 生成完整综述（<code>max_tokens=8192</code>），要求跨批次综合归纳</li>
  </ul>
  <p>此设计可规避中转站代理约 120 秒的超时限制。</p>
</div>

<p><strong>综述风格</strong>（可在 UI 中切换）：</p>
<ul>
  <li><span class="tag tag-blue">学术综述</span>：详细 1500–2500 字，含研究问题、主要发现、方法论和争议</li>
  <li><span class="tag tag-green">简洁综述</span>：500–800 字，精炼摘要</li>
  <li><span class="tag tag-orange">英文综述</span>：与学术综述等长，全英文输出</li>
</ul>
<p>参考文献格式：<em>Author et al. (Year). *Title*. *Journal*.</em>（英文原版，中文译名作括号补充）。</p>

<h3 id="wf-export">8. 导出</h3>
<p>完成后自动调用 openpyxl 生成 Excel 文件。输出路径：</p>
<pre><code>D:\claude\research\results\tables\YYYYMMDD\papers_{关键词}_{YYYYMMDD}.xlsx
D:\claude\research\paper\review_{关键词}_{YYYYMMDD}.txt</code></pre>
<p>若默认目录不可写，则退回到桌面 <code>PaperTracker\</code> 文件夹。文件名中的 Windows 非法字符（<code>\ / : * ? " &lt; &gt; |</code>）自动替换为下划线。</p>

<!-- ════ 五、收录期刊 ════ -->
<h2 id="journals">五、收录期刊</h2>

<h3 id="j-en">英文期刊</h3>

<h4>Top5 综合经济学（5本）</h4>
<table>
  <tr><th>缩写</th><th>期刊名</th></tr>
  <tr><td>AER</td><td>American Economic Review</td></tr>
  <tr><td>QJE</td><td>Quarterly Journal of Economics</td></tr>
  <tr><td>JPE</td><td>Journal of Political Economy</td></tr>
  <tr><td>REStud</td><td>Review of Economic Studies</td></tr>
  <tr><td>ECMA</td><td>Econometrica</td></tr>
</table>

<h4>英文其他顶刊（18本，含 NBER WP）</h4>
<table>
  <tr><th>缩写</th><th>期刊名</th><th>领域</th></tr>
  <tr><td>REStat</td><td>Review of Economics and Statistics</td><td>综合</td></tr>
  <tr><td>JEL</td><td>Journal of Economic Literature</td><td>综合</td></tr>
  <tr><td>JEP</td><td>Journal of Economic Perspectives</td><td>综合</td></tr>
  <tr><td>AEJ:AE / EP / Mac / Mic</td><td>American Economic Journal 系列</td><td>综合</td></tr>
  <tr><td>JF</td><td>Journal of Finance</td><td>金融</td></tr>
  <tr><td>JFE</td><td>Journal of Financial Economics</td><td>金融</td></tr>
  <tr><td>RFS</td><td>Review of Financial Studies</td><td>金融</td></tr>
  <tr><td>JFQA</td><td>Journal of Financial and Quantitative Analysis</td><td>金融</td></tr>
  <tr><td>JOLE</td><td>Journal of Labor Economics</td><td>劳动</td></tr>
  <tr><td>JHR</td><td>Journal of Human Resources</td><td>劳动</td></tr>
  <tr><td>JDE</td><td>Journal of Development Economics</td><td>发展</td></tr>
  <tr><td>JIE</td><td>Journal of International Economics</td><td>国际</td></tr>
  <tr><td>JoE</td><td>Journal of Econometrics</td><td>计量</td></tr>
  <tr><td>NBER WP</td><td>NBER Working Paper</td><td>工作论文</td></tr>
</table>

<h4>经济类期刊（交大安泰 A/A-，共 50 本）</h4>
<p>涵盖 RAND J ECON、J PUBLIC ECON、J MONETARY ECON、J ECON THEORY 等国际主流经济学期刊。完整列表见程序内「收录期刊」弹窗。</p>

<h4>管理类期刊（交大安泰 A/A-，共 46 本）</h4>
<p>涵盖 MANAGE SCI、ORGAN SCI、J MARKETING、J FINANC ECON 等国际主流管理学期刊。完整列表见程序内「收录期刊」弹窗。</p>

<h3 id="j-cn">中文期刊（24本）</h3>
<table>
  <tr><th>级别</th><th>期刊</th></tr>
  <tr><td>特A / A+</td><td>经济研究、管理世界、经济学(季刊)、世界经济</td></tr>
  <tr><td>A</td><td>中国工业经济、数量经济技术经济研究、金融研究、统计研究、经济学动态、中国农村经济</td></tr>
  <tr><td>A-</td><td>中国人口·资源与环境、财贸经济、国际贸易问题、南开经济研究、经济科学、财经研究、经济评论、经济学家、国际金融研究、产业经济研究</td></tr>
  <tr><td>管理类补充</td><td>会计研究、管理科学学报、系统工程理论与实践、中国行政管理</td></tr>
</table>

<!-- ════ 六、布尔检索语法 ════ -->
<h2 id="bool">六、布尔检索语法</h2>
<p>在关键词输入框中可使用布尔运算符组合多个检索词，实现精准筛选。</p>

<h3 id="bool-ops">运算符说明</h3>
<table>
  <tr><th>运算符</th><th>含义</th><th>优先级</th><th>示例</th></tr>
  <tr><td><code>A * B</code></td><td>AND（且）：A 和 B 同时出现</td><td>高</td><td><code>innovation * technology</code></td></tr>
  <tr><td><code>A + B</code></td><td>OR（或）：A 或 B 出现其一</td><td>低</td><td><code>innovation + technology</code></td></tr>
  <tr><td><code>- A</code></td><td>NOT（非，一元前缀）：A 不出现</td><td>高于 AND</td><td><code>innovation * - policy</code></td></tr>
  <tr><td><code>( )</code></td><td>括号分组，改变优先级</td><td>最高</td><td><code>(innovation + robot) * wage</code></td></tr>
  <tr><td><code>'...'</code></td><td>精确短语匹配</td><td>原子</td><td><code>'minimum wage' * china</code></td></tr>
</table>

<p><strong>优先级顺序（高→低）</strong>：括号 &gt; NOT &gt; AND(<code>*</code>) &gt; OR(<code>+</code>)</p>

<h3 id="bool-example">完整示例</h3>
<div class="card">
  <p><strong>目标</strong>：检索「与机器人或人工智能相关的劳动市场研究，但不包括纯理论模型论文」</p>
  <pre><code>(robot + 'artificial intelligence') * labor * - theory</code></pre>
  <p><strong>解释</strong>：</p>
  <ul>
    <li><code>(robot + 'artificial intelligence')</code>：标题/摘要/关键词中需出现「robot」或「artificial intelligence」</li>
    <li><code>* labor</code>：同时需要出现「labor」</li>
    <li><code>* - theory</code>：同时不能出现「theory」</li>
  </ul>
  <p><strong>CrossRef 侧行为</strong>：程序自动提取所有词项（robot、artificial intelligence、labor、theory），拼接为 CrossRef 相关性搜索词，保证召回率。过滤由程序在本地执行。</p>
</div>

<div class="notice notice-warn">
  <strong>注意</strong>：布尔过滤是本地文本匹配，大小写不敏感，但需要检索词精确出现在字段文本中（子串匹配）。对知网文献<strong>不生效</strong>（知网服务端已处理）。
</div>

<!-- ════ 七、输出文件 ════ -->
<h2 id="output">七、输出文件</h2>

<h3 id="out-excel">Excel 文件说明</h3>
<p>文件包含「论文列表」和「汇总统计」两个 Sheet。</p>

<h4>论文列表 Sheet 列说明</h4>
<table>
  <tr><th>列名</th><th>说明</th></tr>
  <tr><td>序号</td><td>从 1 开始的序号</td></tr>
  <tr><td>英文标题</td><td>原始英文标题</td></tr>
  <tr><td>中文标题</td><td>AI 翻译的中文标题</td></tr>
  <tr><td>作者</td><td>最多 3 位作者 + et al.（知网文献列全部作者）</td></tr>
  <tr><td>期刊</td><td>期刊全名</td></tr>
  <tr><td>年份</td><td>发表年份</td></tr>
  <tr><td>被引次数</td><td>来自 CrossRef / Semantic Scholar / OpenAlex</td></tr>
  <tr><td>英文摘要</td><td>原始英文摘要（经三级补全流程）</td></tr>
  <tr><td>中文摘要</td><td>AI 翻译的中文摘要</td></tr>
  <tr><td>匹配位置</td><td>检索词出现的字段位置（篇名/关键词/摘要）</td></tr>
  <tr><td>DOI / 链接</td><td>超链接，点击跳转论文原始页面</td></tr>
  <tr><td>来源</td><td>crossref（英文）或 cnki（知网）</td></tr>
  <tr><td>下载 PDF</td><td>知网文献：「⬇ 下载」超链接（需登录知网）；英文文献为空</td></tr>
</table>

<h4>格式特性</h4>
<ul>
  <li>首行冻结（滚动时标题行固定）</li>
  <li>摘要列启用自动换行，行高根据摘要长度自适应（18–300 pt）</li>
  <li>知网论文行背景色区分（浅黄色）</li>
  <li>DOI / 下载链接为超链接格式，蓝色下划线</li>
</ul>

<h3 id="out-review">综述文件说明</h3>
<p>纯文本格式（.txt），UTF-8 编码。内容包含：</p>
<ul>
  <li>标题（关键词 + 日期）</li>
  <li>正文（1500–2500 字，含研究问题、发现、方法、争议）</li>
  <li>参考文献列表（英文原版格式：<em>Author et al. (Year). *Title*. *Journal*.</em>）</li>
</ul>

<!-- ════ 八、常见问题 ════ -->
<h2 id="faq">八、常见问题</h2>

<div class="card">
  <h4>Q：知网功能显示「需安装 PyQt6-WebEngine」</h4>
  <p>A：请确认已安装，且版本与 PyQt6 完全一致：</p>
  <pre><code>pip install --force-reinstall PyQt6-WebEngine</code></pre>
  <p>安装后重启程序。如果运行 exe 版本，则不支持知网功能（需从源代码运行）。</p>
</div>

<div class="card">
  <h4>Q：知网返回 0 篇</h4>
  <p>A：常见原因：</p>
  <ul>
    <li>未登录知网 → 点击「知网登录」按钮，在弹出浏览器中完成登录</li>
    <li>当前 IP 无知网访问权限 → 确保在校园网或机构 VPN 环境下使用</li>
    <li>知网页面结构变动 → 可尝试重新登录</li>
  </ul>
</div>

<div class="card">
  <h4>Q：摘要列大量为空</h4>
  <p>A：程序已尝试三级补全（OpenAlex → Semantic Scholar → 网页爬取）。仍为空的论文属于出版商未公开摘要（常见于 Elsevier 旗下期刊的早期论文），属正常现象，日志会汇总显示「仍缺失 X 篇（无公开摘要）」。</p>
</div>

<div class="card">
  <h4>Q：篇关摘过滤开启后中文文献全被过滤</h4>
  <p>A：这是已修复的问题。知网文献不参与本地过滤，如仍出现此情况请确认使用的是 2026-04-08 及以后的版本。</p>
</div>

<div class="card">
  <h4>Q：生成综述时报连接错误</h4>
  <p>A：通常是中转站代理超时（论文数量多时 prompt 大，响应慢）。程序已内置两阶段生成（>40篇自动拆批），若仍超时可先用篇关摘过滤缩小论文数量再生成。</p>
</div>

<div class="card">
  <h4>Q：窗口启动后占满全屏无法调节</h4>
  <p>A：删除 <code>paper_tracker_settings.json</code> 中的 <code>window_geometry</code> 字段，或直接删除该文件后重启程序。</p>
</div>

<div class="card">
  <h4>Q：中文关键词搜出的英文论文不相关</h4>
  <p>A：程序会用 AI 将中文词翻译为英文学术词再检索，翻译结果可在日志中查看。如果翻译不准确，可直接在输入框输入英文关键词。</p>
</div>

<!-- ════ 九、版本历史 ════ -->
<h2 id="devlog">九、版本历史</h2>

<table>
  <tr><th>日期</th><th>主要更新</th></tr>
  <tr><td>2026-04-08</td><td>修复篇关摘误删知网文献；综述两阶段生成（>40篇）；Excel 新增下载PDF列；预览改为显示全部；行高自适应；模型下拉选择</td></tr>
  <tr><td>2026-04-07</td><td>集成知网内嵌浏览器抓取；知网期刊多选对话框；参数区三分组重构；复选框打勾样式；NBER WP 支持；字体改宋体；全局字号调大；知网排序/年份过滤；PDF下载功能；布尔检索运算符按钮；接口兼容 OpenAI 格式</td></tr>
  <tr><td>2026-04-05</td><td>添加管理类/经济类期刊（交大安泰）；修复 Excel 摘要为空（OpenAlex + 网页三级补全）；多线程并发提速；篇关摘三字段独立勾选；布尔检索表达式支持；综述参考文献改英文原版</td></tr>
  <tr><td>2026-04-01</td><td>合并为单文件 paper_tracker.py；打包为 PaperTracker.exe；中文关键词自动翻译；中转站 API 支持；设置持久化</td></tr>
  <tr><td>2026-03-xx</td><td>初始版本（多文件版）：CrossRef 抓取、Claude 翻译/综述、PyQt6 GUI、Excel 导出</td></tr>
</table>

<div style="margin-top:60px; color:#90A4AE; font-size:12px; text-align:center; border-top:1px solid var(--border); padding-top:20px;">
  PaperTracker · 经济学论文追踪系统 · 说明书生成于 2026-04-09
</div>

</main>

</body>
</html>
"""

# =============================================================================
# PyQt6 GUI
# =============================================================================

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGridLayout, QLabel, QLineEdit, QPushButton, QComboBox, QCheckBox,
    QTextEdit, QProgressBar, QFrame, QSplitter, QTabWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QScrollArea,
    QMessageBox, QSizePolicy, QGroupBox, QLayout, QDialog, QDialogButtonBox,
    QSpinBox, QFileDialog, QStackedWidget,
)
from PyQt6.QtCore import Qt, QThread, QObject, pyqtSignal, QSize, QRect, QPoint
from PyQt6.QtGui import QFont, QTextCursor, QColor, QPainter, QTextCharFormat, QPalette
from PyQt6.QtWidgets import QSizePolicy as SP

try:
    from PyQt6.QtWebEngineWidgets import QWebEngineView
    from PyQt6.QtWebEngineCore import QWebEngineProfile, QWebEnginePage
    WEBENGINE_AVAILABLE = True
except ImportError:
    WEBENGINE_AVAILABLE = False
    QWebEngineView = None
    QWebEngineProfile = None
    QWebEnginePage = None

import threading
from PyQt6.QtCore import QMetaObject, Q_ARG, pyqtSlot

# 生成临时 checkmark SVG 文件，供 QSS image: url() 使用
import tempfile as _tempfile, os as _os
_CHECK_SVG_PATH = _os.path.join(_tempfile.gettempdir(), "pt_check.svg").replace("\\", "/")
with open(_CHECK_SVG_PATH, "w") as _f:
    _f.write('<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 16 16">'
             '<path d="M2 8L6 12L14 4" stroke="white" stroke-width="2.5" '
             'stroke-linecap="round" stroke-linejoin="round" fill="none"/></svg>')

# ─────────────────────────────────────────────────────────────────────────────
# FlowLayout（Qt 官方示例 Python 版，用于关键词 chip 换行排列）
# ─────────────────────────────────────────────────────────────────────────────

class FlowLayout(QLayout):
    def __init__(self, parent=None, h_spacing=6, v_spacing=6):
        super().__init__(parent)
        self._items = []
        self._h_spacing = h_spacing
        self._v_spacing = v_spacing

    def addItem(self, item):
        self._items.append(item)

    def horizontalSpacing(self):
        return self._h_spacing

    def verticalSpacing(self):
        return self._v_spacing

    def count(self):
        return len(self._items)

    def itemAt(self, index):
        if 0 <= index < len(self._items):
            return self._items[index]
        return None

    def takeAt(self, index):
        if 0 <= index < len(self._items):
            return self._items.pop(index)
        return None

    def expandingDirections(self):
        return Qt.Orientation(0)

    def hasHeightForWidth(self):
        return True

    def heightForWidth(self, width):
        return self._do_layout(QRect(0, 0, width, 0), test_only=True)

    def setGeometry(self, rect):
        super().setGeometry(rect)
        self._do_layout(rect, test_only=False)

    def sizeHint(self):
        return self.minimumSize()

    def minimumSize(self):
        size = QSize()
        for item in self._items:
            size = size.expandedTo(item.minimumSize())
        margins = self.contentsMargins()
        size += QSize(margins.left() + margins.right(), margins.top() + margins.bottom())
        return size

    def _do_layout(self, rect, test_only):
        margins = self.contentsMargins()
        effective = rect.adjusted(margins.left(), margins.top(), -margins.right(), -margins.bottom())
        x, y = effective.x(), effective.y()
        line_height = 0
        for item in self._items:
            space_x = self._h_spacing
            space_y = self._v_spacing
            next_x = x + item.sizeHint().width() + space_x
            if next_x - space_x > effective.right() and line_height > 0:
                x = effective.x()
                y += line_height + space_y
                next_x = x + item.sizeHint().width() + space_x
                line_height = 0
            if not test_only:
                item.setGeometry(QRect(QPoint(x, y), item.sizeHint()))
            x = next_x
            line_height = max(line_height, item.sizeHint().height())
        return y + line_height - rect.y() + margins.bottom()


# ─────────────────────────────────────────────────────────────────────────────
# 常量
# ─────────────────────────────────────────────────────────────────────────────

SCOPE_LABELS  = [
    "无（不搜索英文期刊）",
    "全部期刊（经济+管理合并去重）",
    "全部英文顶刊（23本）",
    "仅 Top5（AER/QJE/JPE/REStud/ECMA）",
    "管理类（交大安泰 A+A- 共46本）",
    "经济类（交大A+A-∪上财 共50本）",
]
YEARS_OPTIONS = [1, 3, 5, 10]
YEARS_LABELS  = ["最近 1 年", "最近 3 年", "最近 5 年", "最近 10 年"]
MODE_OPTIONS  = ["both", "latest", "cited"]
MODE_LABELS   = ["最新 + 高引（推荐）", "仅最新发表", "仅高引用"]
CHIP_COLORS   = [
    "#E3F2FD", "#F3E5F5", "#E8F5E9", "#FFF3E0",
    "#FCE4EC", "#E0F7FA", "#F9FBE7", "#EDE7F6",
]


# ─────────────────────────────────────────────────────────────────────────────
# Logging → Qt Signal 桥接
# ─────────────────────────────────────────────────────────────────────────────

class LogSignalHandler(logging.Handler):
    def __init__(self, signal):
        super().__init__()
        self.signal = signal
        self.setFormatter(logging.Formatter(
            "%(asctime)s  [%(levelname)s]  %(message)s", datefmt="%H:%M:%S"
        ))

    def emit(self, record):
        try:
            self.signal.emit(self.format(record))
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────────────
# CnkiManager — 知网嵌入式浏览器控制器（必须在主线程操作）
# ─────────────────────────────────────────────────────────────────────────────

class CnkiManager(QObject):
    status_update    = pyqtSignal(str)
    captcha_detected = pyqtSignal()
    search_complete  = pyqtSignal(list)
    download_started = pyqtSignal(str)   # 文件名
    download_done    = pyqtSignal(str)   # 保存路径
    download_failed  = pyqtSignal(str)   # 错误信息

    def __init__(self, parent=None):
        super().__init__(parent)
        self.login_confirmed = False
        self._state = "idle"
        self._pending = {}
        self._raw_papers = []
        self._abs_idx = 0
        self._worker_event = None
        self._worker_result = []
        self._current_page = 1
        self._max_pages = 1
        self._download_save_dir = ""   # 下载保存目录

        if WEBENGINE_AVAILABLE:
            from PyQt6.QtWebEngineCore import QWebEngineProfile, QWebEnginePage
            from PyQt6.QtWebEngineWidgets import QWebEngineView
            self._profile = QWebEngineProfile("cnki_session", self)
            self._page = QWebEnginePage(self._profile, self)
            self._view = QWebEngineView()
            self._view.setPage(self._page)
            self._view.hide()
            self._page.loadFinished.connect(self._on_load_finished)
            # 拦截下载请求
            self._profile.downloadRequested.connect(self._on_download_requested)
            from PyQt6.QtCore import QTimer
            self._timer = QTimer(self)
            self._timer.setInterval(600)
            self._timer.timeout.connect(self._poll_results)
        else:
            self._profile = None
            self._page = None
            self._view = None
            self._timer = None

    def _run_js(self, js, callback):
        if self._page:
            self._page.runJavaScript(js, callback)

    @pyqtSlot(str, int, int, str, str)
    def start_search(self, keyword: str, max_results: int, years_back: int, journal_filter: str, sort_id: str = "PT"):
        if not WEBENGINE_AVAILABLE or self._page is None:
            if self._worker_event:
                self._worker_event.set()
            return
        # journal_filter 是 JSON 字符串（期刊名列表）
        import json as _json
        try:
            journals_names = _json.loads(journal_filter) if journal_filter else []
        except Exception:
            journals_names = []

        self._state = "navigating"
        self._pending = {
            "keyword": keyword,
            "max_results": max_results,   # 每本期刊最多抓多少篇
            "years_back": years_back,
            "journal_queue": journals_names if journals_names else [None],  # None = 不限期刊
            "sort_id": sort_id if sort_id else "PT",
        }
        self._raw_papers = []
        self._abs_idx = 0
        self._current_page = 1
        self._year_filter_applied = False
        self._sort_applied = False
        self._journal_idx = 0
        self._journal_collected = 0
        kw = keyword
        total = len(self._pending["journal_queue"])
        self.status_update.emit(f"知网：开始逐刊搜索（共{total}本期刊）")
        self.status_update.emit("知网：导航到高级检索页…")
        from PyQt6.QtCore import QUrl
        self._page.load(QUrl("https://kns.cnki.net/kns8s/AdvSearch"))

    def _on_load_finished(self, ok):
        current_url = self._page.url().toString() if self._page else ""
        if self._state == "navigating":
            # 知网可能先跳验证码页再跳回 AdvSearch，必须确认是目标页才填写
            if "AdvSearch" not in current_url:
                return
            self._state = "filling"
            kw = self._pending.get("keyword", "")
            queue = self._pending.get("journal_queue", [None])
            jname = queue[self._journal_idx] if self._journal_idx < len(queue) else None
            self.status_update.emit(f"知网：搜索期刊 [{self._journal_idx+1}/{len(queue)}] {jname or '不限期刊'}，关键词={kw!r}")
            import json as _json
            js = (_CNKI_JS_FILL_ADVSEARCH
                  .replace('"{KEYWORD}"', repr(kw))
                  .replace('{JOURNAL_NAME_JSON}', _json.dumps(jname, ensure_ascii=False)))
            # 同步 JS，直接用 callback
            self._page.runJavaScript(js, self._on_fill_done_direct)
        elif self._state in ("filling_submitted", "next_page_loading"):
            self._state = "polling"
            self._timer.setInterval(600)
            self._timer.timeout.disconnect()
            self._timer.timeout.connect(self._poll_results)
            self._timer.start()
        elif self._state == "abs_loading":
            self._run_js(_CNKI_JS_EXTRACT_ABSTRACT, self._on_abs_extracted)

    def _on_fill_done_direct(self, result):
        """同步 JS 填写完成的直接 callback"""
        if result == "captcha":
            self._state = "idle"
            self.captcha_detected.emit()
            if self._worker_event:
                self._worker_event.set()
            return
        if result in ("no_topic", "no_button", None):
            # 页面元素未就绪，等 500ms 重试
            self.status_update.emit("知网：页面初始化中，等待就绪…")
            self._fill_poll_count = 0
            self._timer.setInterval(500)
            self._timer.timeout.disconnect()
            self._timer.timeout.connect(self._poll_fill_retry)
            self._timer.start()
            return
        # submitted — 启动结果轮询
        self._state = "polling"
        self._poll_count = 0
        self._year_filter_applied = False  # 每本期刊独立应用年份筛选
        self._timer.setInterval(600)
        self._timer.timeout.disconnect()
        self._timer.timeout.connect(self._poll_results)
        self._timer.start()

    def _poll_fill_retry(self):
        """SearchQueryState 未就绪时重试填写，最多等 20 秒"""
        self._fill_poll_count = getattr(self, '_fill_poll_count', 0) + 1
        if self._fill_poll_count > 40:  # 20秒超时
            self._timer.stop()
            self.status_update.emit("知网：页面初始化超时，跳过知网搜索")
            self._state = "idle"
            if self._worker_event:
                self._worker_event.set()
            return
        # 重新执行填写 JS
        import json as _json
        kw = self._pending.get("keyword", "")
        queue = self._pending.get("journal_queue", [None])
        jname = queue[self._journal_idx] if self._journal_idx < len(queue) else None
        js = (_CNKI_JS_FILL_ADVSEARCH
              .replace('"{KEYWORD}"', repr(kw))
              .replace('{JOURNAL_NAME_JSON}', _json.dumps(jname, ensure_ascii=False)))
        self._timer.stop()
        self._page.runJavaScript(js, self._on_fill_done_direct)

    def _poll_fill_done(self):
        self._fill_poll_count = getattr(self, '_fill_poll_count', 0) + 1
        # 最多等 90 秒（20本期刊 × 约 2.5秒/本 + 余量）
        if self._fill_poll_count > 300:
            self._timer.stop()
            self.status_update.emit("知网：填写期刊超时，直接提交")
            self._page.runJavaScript("document.querySelector('.btn-search')?.click()")
            self._state = "polling"
            self._poll_count = 0
            self._timer.setInterval(600)
            self._timer.timeout.disconnect()
            self._timer.timeout.connect(self._poll_results)
            self._timer.start()
            return
        self._run_js(_CNKI_JS_CHECK_FILL_DONE, self._on_check_fill_done)

    def _on_check_fill_done(self, result):
        if result == "pending":
            return  # 继续等
        self._timer.stop()
        # 填写完成，处理结果
        if result == "captcha":
            self._state = "idle"
            self.captcha_detected.emit()
            if self._worker_event:
                self._worker_event.set()
            return
        if result in ("no_topic", "no_button"):
            self.status_update.emit(f"知网：页面元素未找到({result})，跳过知网搜索")
            self._state = "idle"
            if self._worker_event:
                self._worker_event.set()
            return
        # submitted 或 done — 启动结果轮询
        self._state = "polling"
        self._poll_count = 0
        self._timer.setInterval(600)
        self._timer.timeout.disconnect()
        self._timer.timeout.connect(self._poll_results)
        self._timer.start()

    def _poll_results(self):
        self._poll_count = getattr(self, '_poll_count', 0) + 1
        if self._poll_count > 100:  # 60秒超时（600ms × 100）
            self._timer.stop()
            self.status_update.emit("知网：等待结果超时，结束搜索")
            # 若还有未处理的期刊，跳到下一本；否则抓摘要
            queue = self._pending.get("journal_queue", [None])
            if len(queue) > 1 and self._journal_idx < len(queue) - 1:
                self._next_journal()
            else:
                self._fetch_abstracts()
            return
        self._run_js(_CNKI_JS_CHECK_RESULTS, self._on_check_results)

    def _on_check_results(self, result_str):
        try:
            data = json.loads(result_str)
        except Exception:
            return
        if data.get("error") == "captcha":
            self._timer.stop()
            self._state = "idle"
            self.captcha_detected.emit()
            if self._worker_event:
                self._worker_event.set()
            return
        if data.get("ready"):
            self._timer.stop()
            # 首次出现结果时，先应用排序（只做一次）
            sort_id = self._pending.get("sort_id", "PT")
            if sort_id and not getattr(self, '_sort_applied', False):
                self._sort_applied = True
                js = _CNKI_JS_APPLY_SORT.replace('"{SORT_ID}"', f'"{sort_id}"')
                self._run_js(js, self._on_sort_applied)
                return
            years_back = self._pending.get("years_back", 3)
            # 应用年份筛选（只做一次）
            if years_back > 0 and not getattr(self, '_year_filter_applied', False):
                self._year_filter_applied = True
                import datetime as _dt
                cutoff = _dt.datetime.now().year - years_back
                js = _CNKI_JS_APPLY_YEAR_FILTER.replace('{CUTOFF_YEAR}', str(cutoff))
                self._run_js(js, self._on_year_filter_applied)
            else:
                self._run_js(_CNKI_JS_EXTRACT_LIST, self._on_list_extracted)

    def _on_sort_applied(self, result_str):
        try:
            data = json.loads(result_str)
        except Exception:
            data = {}
        sort_names = {"FFD": "相关度", "PT": "发表时间", "CF": "被引", "DFR": "下载", "ZH": "综合"}
        sort_id = self._pending.get("sort_id", "PT")
        sort_name = sort_names.get(sort_id, sort_id)
        if data.get("already"):
            self.status_update.emit(f"知网：排序方式已是「{sort_name}」，无需切换")
        elif data.get("ok"):
            self.status_update.emit(f"知网：已切换排序方式为「{sort_name}」，等待结果刷新…")
            # 排序切换触发页面刷新，重新等待结果
            self._state = "polling"
            self._poll_count = 0
            self._timer.setInterval(800)
            self._timer.timeout.disconnect()
            self._timer.timeout.connect(self._poll_results)
            self._timer.start()
            return
        else:
            self.status_update.emit(f"知网：排序按钮未找到（{data.get('reason','')}），跳过排序")
        # 排序已就绪或失败，继续年份过滤
        years_back = self._pending.get("years_back", 3)
        if years_back > 0 and not getattr(self, '_year_filter_applied', False):
            self._year_filter_applied = True
            import datetime as _dt
            cutoff = _dt.datetime.now().year - years_back
            js = _CNKI_JS_APPLY_YEAR_FILTER.replace('{CUTOFF_YEAR}', str(cutoff))
            self._run_js(js, self._on_year_filter_applied)
        else:
            self._run_js(_CNKI_JS_EXTRACT_LIST, self._on_list_extracted)

    def _on_year_filter_applied(self, result_str):
        try:
            data = json.loads(result_str)
        except Exception:
            data = {}
        if data.get("applied") and data.get("clicked", 0) > 0:
            self.status_update.emit(f"知网：已应用年份筛选（截止 {self._pending.get('years_back',3)} 年内），等待结果刷新…")
            # 年份筛选触发了页面刷新，需等待新结果加载
            self._state = "polling"
            self._poll_count = 0
            self._timer.setInterval(800)
            self._timer.timeout.disconnect()
            self._timer.timeout.connect(self._poll_results)
            self._timer.start()
        else:
            # 没有可点击的年份复选框（页面结构不同），直接提取
            if data.get("reason"):
                self.status_update.emit(f"知网：年份筛选跳过（{data.get('reason')}），直接提取结果")
            self._run_js(_CNKI_JS_EXTRACT_LIST, self._on_list_extracted)

    def _on_list_extracted(self, result_str):
        try:
            data = json.loads(result_str)
        except Exception:
            self._next_journal()
            return
        results = data.get("results", [])
        years_back = self._pending.get("years_back", 3)
        max_results = self._pending.get("max_results", 20)  # 每本期刊上限
        import datetime
        cutoff_year = datetime.datetime.now().year - years_back if years_back > 0 else 0

        # 记录当前期刊已收集的数量（翻页时累加）
        journal_collected_before = getattr(self, '_journal_collected', 0)
        journal_added = 0

        for item in results:
            if journal_added >= max_results - journal_collected_before:
                break
            date_str = item.get("date", "")
            year = 0
            try:
                year = int(date_str[:4])
            except Exception:
                pass
            if cutoff_year and year and year < cutoff_year:
                continue
            self._raw_papers.append(item)
            journal_added += 1

        self._journal_collected = journal_collected_before + journal_added
        queue = self._pending.get("journal_queue", [None])
        jname = queue[self._journal_idx] if self._journal_idx < len(queue) else "?"
        self.status_update.emit(f"知网：[{self._journal_idx+1}/{len(queue)}] {jname or '不限'} 已收集 {self._journal_collected}/{max_results} 篇，总计 {len(self._raw_papers)} 篇")

        # 当前期刊够了，或者无结果 → 下一本期刊
        if self._journal_collected >= max_results or journal_added == 0:
            self._next_journal()
            return

        # 翻页继续收集
        self._run_js(_CNKI_JS_NEXT_PAGE, self._on_next_page)

    def _on_next_page(self, result_str):
        try:
            data = json.loads(result_str)
        except Exception:
            self._next_journal()
            return
        if data.get("clicked"):
            # SPA pushState，直接启动轮询
            self._state = "polling"
            self._poll_count = 0
            self._timer.start()
        else:
            # 没有下一页 → 当前期刊搜索完毕，跳下一本
            self._next_journal()

    def _next_journal(self):
        """当前期刊搜索完毕，切换到下一本期刊，或全部完成进入摘要抓取"""
        self._journal_idx += 1
        self._journal_collected = 0
        self._year_filter_applied = False
        queue = self._pending.get("journal_queue", [None])
        if self._journal_idx >= len(queue):
            # 所有期刊搜索完毕，开始抓摘要
            self.status_update.emit(f"知网：所有期刊搜索完毕，共 {len(self._raw_papers)} 篇，开始抓取摘要…")
            self._fetch_abstracts()
        else:
            # 导航到下一本期刊的搜索页
            jname = queue[self._journal_idx]
            self.status_update.emit(f"知网：切换到下一本期刊 [{self._journal_idx+1}/{len(queue)}] {jname}")
            self._state = "navigating"
            self._current_page = 1
            from PyQt6.QtCore import QUrl
            self._page.load(QUrl("https://kns.cnki.net/kns8s/AdvSearch"))

    def _fetch_abstracts(self):
        if not self._raw_papers:
            self._compile()
            return
        # 文章数量超过阈值时跳过摘要抓取（每篇需独立加载页面，太慢）
        if len(self._raw_papers) > 1000:
            self.status_update.emit(f"知网：共 {len(self._raw_papers)} 篇，文章较多，跳过摘要逐篇抓取，直接汇总")
            self._compile()
            return
        self._abs_idx = 0
        self._fetch_next_abstract()

    def _fetch_next_abstract(self):
        if self._abs_idx >= len(self._raw_papers):
            self._compile()
            return
        item = self._raw_papers[self._abs_idx]
        href = item.get("href", "")
        if not href:
            self._abs_idx += 1
            self._fetch_next_abstract()
            return
        self._state = "abs_loading"
        from PyQt6.QtCore import QUrl
        self._page.load(QUrl(href))

    def _on_abs_extracted(self, result_str):
        try:
            data = json.loads(result_str)
            self._raw_papers[self._abs_idx]["abstract"] = data.get("abstract", "")
            self._raw_papers[self._abs_idx]["keywords"] = data.get("keywords", [])
        except Exception:
            pass
        self._abs_idx += 1
        self.status_update.emit(f"知网：摘要 {self._abs_idx}/{len(self._raw_papers)}")
        self._fetch_next_abstract()

    def _compile(self):
        papers = []
        for item in self._raw_papers:
            date_str = item.get("date", "")
            year = 0
            try:
                year = int(date_str[:4])
            except Exception:
                pass
            cite_str = item.get("citations", "").replace(",", "").strip()
            try:
                cites = int(cite_str)
            except Exception:
                cites = -1
            p = Paper(
                title=item.get("title", ""),
                title_zh=item.get("title", ""),  # 知网中文期刊，标题即中文
                abstract=item.get("abstract", ""),
                abstract_zh="",
                authors=item.get("authors", "").split("; ") if item.get("authors") else [],
                year=year,
                journal=item.get("journal", ""),
                journal_abbr=item.get("journal", ""),
                doi="",
                url=item.get("href", ""),
                citations=cites,
                source="cnki",
                match_reason="",
                keywords=item.get("keywords", []),
                download_url=item.get("download_url", ""),
            )
            papers.append(p)
        self._state = "idle"
        self._worker_result = papers
        if self._worker_event:
            self._worker_event.set()
        self.search_complete.emit(papers)

    def _on_download_requested(self, download):
        """拦截 WebEngine 下载请求，保存到指定目录"""
        import os
        # 停止超时计时器
        self._dl_triggered = True
        if hasattr(self, '_dl_timeout_timer'):
            self._dl_timeout_timer.stop()
        save_dir = self._download_save_dir or os.path.expanduser("~")
        os.makedirs(save_dir, exist_ok=True)
        filename = download.suggestedFileName() or "paper.pdf"
        # 避免文件名冲突
        base, ext = os.path.splitext(filename)
        candidate = os.path.join(save_dir, filename)
        counter = 1
        while os.path.exists(candidate):
            candidate = os.path.join(save_dir, f"{base}_{counter}{ext}")
            counter += 1
        download.setDownloadDirectory(save_dir)
        download.setDownloadFileName(os.path.basename(candidate))
        download.accept()
        self.download_started.emit(os.path.basename(candidate))
        download.isFinishedChanged.connect(
            lambda: self._on_download_item_finished(download, candidate)
        )

    def _on_download_item_finished(self, download, save_path):
        try:
            from PyQt6.QtWebEngineCore import QWebEngineDownloadRequest
            if download.state() == QWebEngineDownloadRequest.DownloadState.DownloadCompleted:
                self.download_done.emit(save_path)
            else:
                self.download_failed.emit(f"下载失败（{download.state().name}）")
        except Exception as e:
            self.download_failed.emit(str(e))

    @pyqtSlot(str, str)
    def download_pdf(self, url: str, save_dir: str):
        """让 WebEngine 用已有 session 导航下载链接，由 downloadRequested 信号处理保存"""
        if not WEBENGINE_AVAILABLE or self._page is None:
            self.download_failed.emit("WebEngine 不可用")
            return
        import os
        os.makedirs(save_dir, exist_ok=True)
        self._download_save_dir = save_dir
        self._dl_triggered = False  # 标记本次下载是否触发了 downloadRequested

        # 超时保护：15秒内若 downloadRequested 未触发，报失败
        from PyQt6.QtCore import QTimer
        dl_timer = QTimer(self)
        dl_timer.setSingleShot(True)
        def _on_dl_timeout():
            if not self._dl_triggered:
                self.download_failed.emit(f"下载超时（未触发下载请求，可能需要重新登录知网）")
        dl_timer.timeout.connect(_on_dl_timeout)
        dl_timer.start(15000)
        self._dl_timeout_timer = dl_timer

        # 在当前页面注入 JS，创建隐藏 <a> 并点击，让浏览器用已有 session 触发下载
        js = f"""
(function() {{
    var a = document.createElement('a');
    a.href = {repr(url)};
    a.style.display = 'none';
    document.body.appendChild(a);
    a.click();
    setTimeout(function(){{ document.body.removeChild(a); }}, 1000);
}})()
"""
        self._page.runJavaScript(js)


# ─────────────────────────────────────────────────────────────────────────────
# EnJournalDialog — 多选英文期刊对话框（二级菜单：来源分组 → tier 小节）
# ─────────────────────────────────────────────────────────────────────────────

# 构建二级分组数据：{来源标题: {tier标题: [(issn, abbr, name), ...]}}
def _build_en_journal_tree():
    tree = {}

    def _add(src_label, tier_label, journals_dict):
        if src_label not in tree:
            tree[src_label] = {}
        if tier_label not in tree[src_label]:
            tree[src_label][tier_label] = []
        for issn, info in journals_dict.items():
            tree[src_label][tier_label].append((issn, info["abbr"], info["name"]))

    # 英文顶刊分 tier
    top5   = {k: v for k, v in ENGLISH_JOURNALS.items() if v["tier"] == "Top5" and k != "nber-wp"}
    top    = {k: v for k, v in ENGLISH_JOURNALS.items() if v["tier"] == "Top"  and k != "nber-wp"}
    nber   = {k: v for k, v in ENGLISH_JOURNALS.items() if k == "nber-wp"}
    _add("英文顶刊（综合经济学）", "Top 5", top5)
    _add("英文顶刊（综合经济学）", "Top",   top)
    if nber:
        _add("英文顶刊（综合经济学）", "工作论文", nber)

    # 管理类
    mgmt_a  = {k: v for k, v in MANAGEMENT_JOURNALS.items() if v["tier"] == "A"}
    mgmt_am = {k: v for k, v in MANAGEMENT_JOURNALS.items() if v["tier"] == "A-"}
    _add("管理类（交大安泰）", "A 级",  mgmt_a)
    _add("管理类（交大安泰）", "A- 级", mgmt_am)

    # 经济类
    econ_a  = {k: v for k, v in ECONOMICS_JOURNALS.items() if v["tier"] == "A"}
    econ_am = {k: v for k, v in ECONOMICS_JOURNALS.items() if v["tier"] == "A-"}
    econ_f  = {k: v for k, v in ECONOMICS_JOURNALS.items() if v["tier"] == "First"}
    _add("经济类（交大安泰 / 上财）", "A 级",         econ_a)
    _add("经济类（交大安泰 / 上财）", "A- 级",        econ_am)
    _add("经济类（交大安泰 / 上财）", "First Tier",   econ_f)

    return tree

_EN_JOURNAL_TREE = _build_en_journal_tree()


class EnJournalDialog(QDialog):
    """多选英文期刊对话框，按来源和 tier 两级分组展示。"""

    def __init__(self, selected_issns: set, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择英文期刊")
        self.setMinimumWidth(560)
        self.resize(600, 680)
        self._checkboxes: dict = {}   # issn → QCheckBox

        _check_svg = _CHECK_SVG_PATH
        self.setStyleSheet(f"""
            QDialog   {{ background: #FFFFFF; color: #212121; }}
            QGroupBox {{
                background: #F5F5F5; color: #212121;
                border: 1px solid #BDBDBD; border-radius: 4px;
                margin-top: 8px; font-weight: bold; font-size: 13pt;
            }}
            QGroupBox::title {{ subcontrol-origin: margin; left: 8px; padding: 0 4px; color: #1565C0; }}
            QLabel {{ color: #37474F; font-weight: bold; font-size: 12pt; margin-top: 4px; }}
            QCheckBox {{
                color: #212121; font-size: 12pt; spacing: 8px; padding: 1px 0;
            }}
            QCheckBox::indicator {{
                width: 16px; height: 16px;
                border: 2px solid #9E9E9E; border-radius: 3px; background: #FFFFFF;
            }}
            QCheckBox::indicator:checked {{
                background: #2E7D32; border: 2px solid #2E7D32;
                image: url("{_check_svg}");
            }}
            QCheckBox::indicator:hover {{ border-color: #1565C0; }}
            QPushButton {{
                color: #212121; background: #EEEEEE;
                border: 1px solid #BDBDBD; border-radius: 3px; padding: 2px 8px;
            }}
            QPushButton:hover {{ background: #E0E0E0; }}
        """)

        outer = QVBoxLayout(self)
        outer.setSpacing(6)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        inner = QWidget()
        inner.setStyleSheet("background: #FFFFFF;")
        layout = QVBoxLayout(inner)
        layout.setSpacing(8)
        scroll.setWidget(inner)
        outer.addWidget(scroll, stretch=1)

        for src_label, tier_dict in _EN_JOURNAL_TREE.items():
            src_grp = QGroupBox(src_label)
            src_grp_layout = QVBoxLayout(src_grp)
            src_grp_layout.setSpacing(6)

            # 来源级全选/全不选
            src_issns = [issn for tiers in tier_dict.values() for issn, _, _ in tiers]
            src_row = QHBoxLayout()
            s_all  = QPushButton("全选");  s_all.setFixedHeight(24);  s_all.setFixedWidth(72)
            s_none = QPushButton("全不选"); s_none.setFixedHeight(24); s_none.setFixedWidth(72)
            def _make_src_setter(issns, v):
                def _f():
                    for i in issns:
                        if i in self._checkboxes:
                            self._checkboxes[i].setChecked(v)
                return _f
            s_all.clicked.connect(_make_src_setter(src_issns, True))
            s_none.clicked.connect(_make_src_setter(src_issns, False))
            src_row.addWidget(s_all); src_row.addWidget(s_none); src_row.addStretch()
            src_grp_layout.addLayout(src_row)

            for tier_label, journals in tier_dict.items():
                tier_issns = [issn for issn, _, _ in journals]
                # tier 标题行 + 全选/全不选
                tier_row = QHBoxLayout()
                tier_lbl = QLabel(f"  {tier_label}")
                tier_lbl.setStyleSheet("color: #1565C0; font-weight: bold; font-size: 12pt;")
                t_all  = QPushButton("全选");  t_all.setFixedHeight(22);  t_all.setFixedWidth(60)
                t_none = QPushButton("全不选"); t_none.setFixedHeight(22); t_none.setFixedWidth(60)
                def _make_tier_setter(issns, v):
                    def _f():
                        for i in issns:
                            if i in self._checkboxes:
                                self._checkboxes[i].setChecked(v)
                    return _f
                t_all.clicked.connect(_make_tier_setter(tier_issns, True))
                t_none.clicked.connect(_make_tier_setter(tier_issns, False))
                tier_row.addWidget(tier_lbl); tier_row.addWidget(t_all)
                tier_row.addWidget(t_none);   tier_row.addStretch()
                src_grp_layout.addLayout(tier_row)

                for issn, abbr, name in journals:
                    cb = QCheckBox(f"{abbr}  —  {name}")
                    cb.setChecked(issn in selected_issns)
                    self._checkboxes[issn] = cb
                    src_grp_layout.addWidget(cb)

            layout.addWidget(src_grp)

        # 全局按钮行
        global_row = QHBoxLayout()
        g_all  = QPushButton("全部选中"); g_all.setFixedHeight(28);  g_all.setFixedWidth(96)
        g_none = QPushButton("全部取消"); g_none.setFixedHeight(28); g_none.setFixedWidth(96)
        g_all.clicked.connect(lambda: [c.setChecked(True)  for c in self._checkboxes.values()])
        g_none.clicked.connect(lambda: [c.setChecked(False) for c in self._checkboxes.values()])
        global_row.addWidget(g_all); global_row.addWidget(g_none); global_row.addStretch()
        outer.addLayout(global_row)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        outer.addWidget(btns)

    def get_selected_issns(self) -> set:
        return {issn for issn, cb in self._checkboxes.items() if cb.isChecked()}


# ─────────────────────────────────────────────────────────────────────────────
# FeedbackDialog — 用户反馈 Bug，通过 GitHub Issues API 提交
# ─────────────────────────────────────────────────────────────────────────────

_FEEDBACK_REPO  = "liupinliang666-tech/paper-tracker-feedback"
_FEEDBACK_TOKEN = ""  # filled at build time


class _FeedbackWorker(QObject):
    success = pyqtSignal()
    failure = pyqtSignal(str)

    def __init__(self, payload: bytes):
        super().__init__()
        self._payload = payload

    def run(self):
        try:
            req = urllib.request.Request(
                f"https://api.github.com/repos/{_FEEDBACK_REPO}/issues",
                data=self._payload,
                headers={
                    "Authorization": f"Bearer {_FEEDBACK_TOKEN}",
                    "Accept": "application/vnd.github+json",
                    "Content-Type": "application/json",
                    "X-GitHub-Api-Version": "2022-11-28",
                },
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                resp.read()
            self.success.emit()
        except Exception as e:
            self.failure.emit(str(e))


class FeedbackDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("反馈 Bug")
        self.resize(480, 320)
        self.setStyleSheet("""
            QDialog, QWidget { background: #FFFFFF; color: #212121; }
            QLabel { background: transparent; color: #212121; font-size: 13px; }
            QTextEdit {
                background: #FFFFFF; color: #212121;
                border: 1px solid #BDBDBD; border-radius: 4px;
                font-size: 13px; padding: 4px;
            }
            QLineEdit {
                background: #FFFFFF; color: #212121;
                border: 1px solid #BDBDBD; border-radius: 4px;
                font-size: 13px; padding: 4px;
                min-height: 28px;
            }
            QPushButton {
                background: #F5F5F5; color: #212121;
                border: 1px solid #BDBDBD; border-radius: 4px;
                font-size: 13px; padding: 4px 12px; min-height: 28px;
            }
            QPushButton:hover { background: #E0E0E0; }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        layout.addWidget(QLabel("描述你发现的问题（必填）："))
        self._desc = QTextEdit()
        self._desc.setPlaceholderText("请描述 bug 的现象、复现步骤等……")
        self._desc.setFixedHeight(140)
        layout.addWidget(self._desc)

        layout.addWidget(QLabel("联系方式（选填，方便我回复你）："))
        self._contact = QLineEdit()
        self._contact.setPlaceholderText("邮箱 / 微信 / 留空也可以")
        layout.addWidget(self._contact)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self._submit_btn = QPushButton("提交")
        self._submit_btn.setFixedWidth(80)
        self._submit_btn.setStyleSheet("""
            QPushButton {
                background: #1565C0; color: white;
                border: none; border-radius: 4px;
                font-size: 13px; padding: 4px 12px; min-height: 28px;
            }
            QPushButton:hover { background: #1976D2; }
            QPushButton:disabled { background: #90CAF9; color: white; }
        """)
        self._submit_btn.clicked.connect(self._submit)
        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(80)
        cancel_btn.clicked.connect(self.reject)
        btn_row.addWidget(cancel_btn)
        btn_row.addWidget(self._submit_btn)
        layout.addLayout(btn_row)

    def _submit(self):
        desc = self._desc.toPlainText().strip()
        if not desc:
            QMessageBox.warning(self, "提示", "请填写 bug 描述。")
            return

        contact = self._contact.text().strip()
        body = desc
        if contact:
            body += f"\n\n---\n**联系方式：** {contact}"

        self._submit_btn.setEnabled(False)
        self._submit_btn.setText("提交中…")

        payload = json.dumps({
            "title": desc[:80] + ("…" if len(desc) > 80 else ""),
            "body": body,
            "labels": ["bug"],
        }).encode("utf-8")

        self._thread = QThread()
        self._worker = _FeedbackWorker(payload)
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.success.connect(self._on_submit_success)
        self._worker.failure.connect(self._on_submit_failure)
        self._worker.success.connect(self._thread.quit)
        self._worker.failure.connect(self._thread.quit)
        self._thread.start()

    def _on_submit_success(self):
        QMessageBox.information(self, "已提交", "感谢反馈！\n\n已成功记录，开发者会尽快处理。")
        self.accept()

    def _on_submit_failure(self, err):
        self._submit_btn.setEnabled(True)
        self._submit_btn.setText("提交")
        QMessageBox.critical(self, "提交失败", f"网络错误，请稍后再试。\n\n{err}")


# CnkiLoginDialog — 让用户在嵌入浏览器中完成知网登录
# ─────────────────────────────────────────────────────────────────────────────

class CnkiLoginDialog(QDialog):
    def __init__(self, cnki_manager, parent=None):
        super().__init__(parent)
        self.setWindowTitle("知网登录")
        self.resize(1000, 700)
        layout = QVBoxLayout(self)
        if WEBENGINE_AVAILABLE and cnki_manager and cnki_manager._profile:
            from PyQt6.QtWebEngineWidgets import QWebEngineView
            from PyQt6.QtWebEngineCore import QWebEnginePage
            from PyQt6.QtCore import QUrl
            self._view = QWebEngineView()
            page = QWebEnginePage(cnki_manager._profile, self._view)
            self._view.setPage(page)
            self._view.load(QUrl("https://www.cnki.net"))
            layout.addWidget(self._view)
        else:
            layout.addWidget(QLabel("需要安装 PyQt6-WebEngine 才能使用此功能"))
        hint = QLabel("请在浏览器中完成知网登录（校园网/机构认证），完成后点击下方按钮。")
        hint.setWordWrap(True)
        layout.addWidget(hint)
        btn = QPushButton("我已登录，继续")
        btn.clicked.connect(self.accept)
        layout.addWidget(btn)


class CnkiJournalDialog(QDialog):
    """多选知网期刊对话框，经济类在左、管理类在右两列布局。"""

    def __init__(self, selected: list, parent=None):
        super().__init__(parent)
        self.setWindowTitle("选择中文期刊")
        self.resize(720, 600)
        self._checkboxes: dict = {}

        _check_svg = _CHECK_SVG_PATH
        self.setStyleSheet(f"""
            QDialog {{ background: #FFFFFF; color: #212121; }}
            QGroupBox {{
                background: #F5F5F5; color: #212121;
                border: 1px solid #BDBDBD; border-radius: 4px;
                margin-top: 8px; font-weight: bold; font-size: 13pt;
            }}
            QGroupBox::title {{ subcontrol-origin: margin; left: 8px; padding: 0 4px; }}
            QCheckBox {{ color: #212121; font-size: 13pt; spacing: 8px; padding: 2px 0; }}
            QCheckBox::indicator {{
                width: 18px; height: 18px;
                border: 2px solid #9E9E9E; border-radius: 3px; background: #FFFFFF;
            }}
            QCheckBox::indicator:checked {{
                background: #2E7D32; border: 2px solid #2E7D32;
                image: url("{_check_svg}");
            }}
            QCheckBox::indicator:hover {{ border-color: #1565C0; }}
            QPushButton {{ color: #212121; background: #EEEEEE; border: 1px solid #BDBDBD; border-radius: 3px; padding: 2px 8px; }}
            QPushButton:hover {{ background: #E0E0E0; }}
        """)

        outer = QVBoxLayout(self)
        outer.setSpacing(8)

        # 左右两列：经济类 | 管理类
        cols_layout = QHBoxLayout()
        cols_layout.setSpacing(10)

        left_scroll  = QScrollArea(); left_scroll.setWidgetResizable(True);  left_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        right_scroll = QScrollArea(); right_scroll.setWidgetResizable(True); right_scroll.setFrameShape(QScrollArea.Shape.NoFrame)
        left_inner  = QWidget(); left_inner.setStyleSheet("background:#FFFFFF;")
        right_inner = QWidget(); right_inner.setStyleSheet("background:#FFFFFF;")
        left_layout  = QVBoxLayout(left_inner);  left_layout.setSpacing(6)
        right_layout = QVBoxLayout(right_inner); right_layout.setSpacing(6)
        left_scroll.setWidget(left_inner)
        right_scroll.setWidget(right_inner)

        # 列标题
        def _col_title(text, color):
            lbl = QLabel(text)
            lbl.setStyleSheet(f"color:{color}; font-weight:bold; font-size:14pt; padding:4px 0;")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            return lbl
        left_layout.addWidget(_col_title("经济类", "#1565C0"))
        right_layout.addWidget(_col_title("管理类", "#1B5E20"))

        for tier, journals in CNKI_JOURNALS.items():
            is_mgmt = "管理类" in tier
            title_color = "#1B5E20" if is_mgmt else "#1565C0"
            target = right_layout if is_mgmt else left_layout

            grp = QGroupBox(f"【{tier}】")
            grp.setStyleSheet(f"QGroupBox::title {{ color: {title_color}; }}")
            grp_layout = QVBoxLayout(grp)
            grp_layout.setSpacing(4)

            sel_row = QHBoxLayout()
            all_btn  = QPushButton("全选");  all_btn.setFixedHeight(24); all_btn.setFixedWidth(72)
            none_btn = QPushButton("全不选"); none_btn.setFixedHeight(24); none_btn.setFixedWidth(72)

            def _make_tier_setter(js, v):
                def _set():
                    for j in js:
                        if j in self._checkboxes:
                            self._checkboxes[j].setChecked(v)
                return _set
            all_btn.clicked.connect(_make_tier_setter(journals, True))
            none_btn.clicked.connect(_make_tier_setter(journals, False))
            sel_row.addWidget(all_btn); sel_row.addWidget(none_btn); sel_row.addStretch()
            grp_layout.addLayout(sel_row)

            for j in journals:
                cb = QCheckBox(j)
                cb.setChecked(j in selected)
                self._checkboxes[j] = cb
                grp_layout.addWidget(cb)
            target.addWidget(grp)

        left_layout.addStretch()
        right_layout.addStretch()
        cols_layout.addWidget(left_scroll, stretch=3)
        cols_layout.addWidget(right_scroll, stretch=2)
        outer.addLayout(cols_layout, stretch=1)

        # 全局按钮行
        global_row = QHBoxLayout()
        g_all  = QPushButton("全部选中"); g_all.setFixedHeight(28);  g_all.setFixedWidth(96)
        g_none = QPushButton("全部取消"); g_none.setFixedHeight(28); g_none.setFixedWidth(96)
        g_all.clicked.connect(lambda: [c.setChecked(True)  for c in self._checkboxes.values()])
        g_none.clicked.connect(lambda: [c.setChecked(False) for c in self._checkboxes.values()])
        global_row.addWidget(g_all); global_row.addWidget(g_none); global_row.addStretch()
        outer.addLayout(global_row)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        outer.addWidget(btns)

    def get_selected(self) -> list:
        return [j for j, cb in self._checkboxes.items() if cb.isChecked()]


# ─────────────────────────────────────────────────────────────────────────────
# fetch_cnki_papers — 供 WorkerThread 调用（在后台线程执行，阻塞等待结果）
# ─────────────────────────────────────────────────────────────────────────────

def fetch_cnki_papers(keyword: str, max_results: int, years_back: int,
                      cnki_manager, timeout: float = 0,
                      journal_filter: str = "", sort_id: str = "PT") -> list:
    if not WEBENGINE_AVAILABLE or cnki_manager is None:
        return []
    # 动态计算超时：每刊搜索约30秒 + 每篇摘要约2秒，再加60秒余量
    import json as _j
    try:
        n_journals = len(_j.loads(journal_filter)) if journal_filter else 1
    except Exception:
        n_journals = 1
    if timeout <= 0:
        timeout = n_journals * 30 + n_journals * max_results * 2 + 60
    event = threading.Event()
    cnki_manager._worker_event = event
    cnki_manager._worker_result = []
    QMetaObject.invokeMethod(
        cnki_manager, "start_search",
        Qt.ConnectionType.QueuedConnection,
        Q_ARG(str, keyword),
        Q_ARG(int, max_results),
        Q_ARG(int, years_back),
        Q_ARG(str, journal_filter),
        Q_ARG(str, sort_id),
    )
    event.wait(timeout=timeout)
    return list(cnki_manager._worker_result)


# ─────────────────────────────────────────────────────────────────────────────
# WorkerThread
# ─────────────────────────────────────────────────────────────────────────────

class WorkerThread(QThread):
    log_message  = pyqtSignal(str)
    progress     = pyqtSignal(int, str)
    finished     = pyqtSignal(bool, str)
    papers_ready = pyqtSignal(list)

    def __init__(self, opts: dict, parent=None, cnki_manager=None):
        super().__init__(parent)
        self.opts = opts
        self._stop_flag = False
        self._cnki_manager = cnki_manager

    def stop(self):
        self._stop_flag = True

    def run(self):
        handler = LogSignalHandler(self.log_message)
        root_logger = logging.getLogger()
        root_logger.setLevel(logging.INFO)
        root_logger.addHandler(handler)
        # 同时写入日志文件，方便排查问题
        log_path = os.path.join(_EXE_DIR, "paper_tracker.log")
        file_handler = logging.FileHandler(log_path, encoding="utf-8", mode="a")
        file_handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s", "%H:%M:%S"))
        root_logger.addHandler(file_handler)
        try:
            self._execute()
        except Exception as e:
            self.log_message.emit(f"[ERROR] 执行出错: {e}")
            self.finished.emit(False, str(e))
        finally:
            root_logger.removeHandler(handler)
            root_logger.removeHandler(file_handler)
            file_handler.close()

    @staticmethod
    def _is_chinese(text: str) -> bool:
        return any('\u4e00' <= c <= '\u9fff' for c in text)

    def _translate_keyword_to_english(self, keyword: str) -> str:
        prompt = (
            f"请将以下经济学研究关键词翻译成英文学术搜索词。\n"
            f"要求：\n"
            f"1. 只输出英文关键词，不要任何解释\n"
            f"2. 如果一个中文词对应多个常用英文表达，用逗号隔开，最多给3个\n"
            f"3. 使用经济学学术论文中最常用的表达\n\n"
            f"中文关键词：{keyword}\n\n"
            f"英文关键词："
        )
        result = _call_llm(prompt, max_tokens=100, kind="fast")
        if not result:
            raise RuntimeError("API 调用失败")
        return result.strip()

    def _execute(self):
        global ANTHROPIC_API_KEY, ANTHROPIC_BASE_URL, MODEL_FAST, MODEL_STRONG, API_TYPE

        opts = self.opts
        keywords = [k for k in opts.get("keywords", []) if k.strip()]
        if not keywords:
            self.finished.emit(False, "请至少输入一个关键词")
            return

        # 动态注入配置
        if opts.get("api_key"):
            ANTHROPIC_API_KEY = opts["api_key"].strip()
        if opts.get("base_url"):
            ANTHROPIC_BASE_URL = opts["base_url"].strip()
        if opts.get("model_fast"):
            MODEL_FAST = opts["model_fast"].strip()
        if opts.get("model_strong"):
            MODEL_STRONG = opts["model_strong"].strip()
        API_TYPE = opts.get("api_type", "anthropic")

        all_saved_paths = []

        for kw_idx, keyword in enumerate(keywords):
            if self._stop_flag:
                self.log_message.emit("[INFO] 用户已停止任务")
                break

            kw_prefix = f"[{kw_idx+1}/{len(keywords)}] " if len(keywords) > 1 else ""
            self.log_message.emit(f"\n{'='*50}")
            self.log_message.emit(f"{kw_prefix}开始搜索关键词：{keyword}")

            all_papers = []
            base_pct = kw_idx * 100 // len(keywords)
            en_keyword = keyword  # 默认值，skip_en 时不翻译直接用原词

            # 抓取英文期刊（period_scope == 0 时跳过）
            if not opts.get("skip_en"):
                # 中文关键词 → 英文
                en_keyword = keyword
                if self._is_chinese(keyword):
                    if ANTHROPIC_API_KEY:
                        self.progress.emit(base_pct + 2, f"{kw_prefix}翻译关键词为英文...")
                        try:
                            en_keyword = self._translate_keyword_to_english(keyword)
                            self.log_message.emit(f"[INFO] 关键词翻译：「{keyword}」→「{en_keyword}」")
                        except Exception as e:
                            self.log_message.emit(f"[WARNING] 关键词翻译失败，使用原词检索: {e}")
                            en_keyword = keyword
                    else:
                        self.log_message.emit("[WARNING] 未配置 API Key，中文关键词将直接用于英文期刊检索")

                # 布尔表达式：提取所有词项用于 CrossRef 搜索
                is_bool = _is_bool_expr(en_keyword)
                if is_bool:
                    search_terms = _extract_all_terms(en_keyword)
                    crossref_query = " ".join(search_terms) if search_terms else en_keyword
                    self.log_message.emit(f"[INFO] 布尔检索模式，提取词项：{search_terms}，CrossRef 查询：「{crossref_query}」")
                else:
                    crossref_query = en_keyword

                self.progress.emit(base_pct + 5, f"{kw_prefix}抓取英文顶刊...")
                try:
                    fetch_kwargs = dict(
                        mode=opts.get("mode", "both"),
                        max_per_journal=opts.get("max", 50),
                        years_back=opts.get("years", 3),
                    )
                    custom_issns = opts.get("en_custom_journals", [])
                    if custom_issns:
                        all_en = {**ENGLISH_JOURNALS, **MANAGEMENT_JOURNALS, **ECONOMICS_JOURNALS}
                        custom_dict = {issn: all_en[issn] for issn in custom_issns if issn in all_en}
                        self.log_message.emit(f"[INFO] 自定义英文期刊：{len(custom_dict)} 本")
                        en_papers = fetch_english_papers(crossref_query, journals=custom_dict, **fetch_kwargs)
                    elif opts.get("all_journals"):
                        en_papers = fetch_all_papers(crossref_query, **fetch_kwargs)
                    elif opts.get("top5"):
                        en_papers = fetch_top5_only(crossref_query, **fetch_kwargs)
                    elif opts.get("management"):
                        en_papers = fetch_management_papers(crossref_query, **fetch_kwargs)
                    elif opts.get("economics"):
                        en_papers = fetch_economics_papers(crossref_query, **fetch_kwargs)
                    else:
                        en_papers = fetch_english_papers(crossref_query, **fetch_kwargs)
                    self.log_message.emit(f"[INFO] 英文期刊抓取完成：{len(en_papers)} 篇")
                    all_papers.extend(en_papers)
                except Exception as e:
                    self.log_message.emit(f"[WARNING] 英文期刊抓取异常: {e}")
            else:
                self.log_message.emit("[INFO] 期刊范围设为「无」，跳过英文期刊检索")

            # 知网中文文献
            if opts.get("search_cnki") and self._cnki_manager is not None:
                cn_max = opts.get("cnki_max", 20)
                self.log_message.emit(f"[INFO] 开始搜索知网中文文献：{keyword}（最多 {cn_max} 篇）")
                self.progress.emit(base_pct + 20, f"{kw_prefix}知网搜索中…")
                try:
                    cn_papers = fetch_cnki_papers(
                        keyword, cn_max, opts.get("years", 3),
                        self._cnki_manager,
                        journal_filter=opts.get("cnki_journal_filter", ""),
                        sort_id=opts.get("cnki_sort_id", "PT"),
                    )
                    self.log_message.emit(f"[INFO] 知网返回 {len(cn_papers)} 篇中文文献")
                    all_papers.extend(cn_papers)
                except Exception as e:
                    self.log_message.emit(f"[WARNING] 知网搜索异常: {e}")

            if self._stop_flag:
                break

            if not all_papers:
                self.log_message.emit(f"[WARNING] 关键词 '{keyword}' 未抓取到任何论文")
                continue

            # 去重（按标题规范化后去重）
            seen_titles = set()
            deduped = []
            for p in all_papers:
                norm = re.sub(r'[^\w]', '', (p.title or '').lower())[:60]
                if norm and norm not in seen_titles:
                    seen_titles.add(norm)
                    deduped.append(p)
            if len(deduped) < len(all_papers):
                self.log_message.emit(f"[INFO] 去重：移除 {len(all_papers) - len(deduped)} 篇重复文献")
            all_papers = deduped

            # 排序
            mode = opts.get("mode", "both")
            if mode == "cited":
                all_papers.sort(key=lambda p: p.citations, reverse=True)
            elif mode == "latest":
                all_papers.sort(key=lambda p: p.year, reverse=True)
            else:
                all_papers.sort(key=lambda p: (p.year, p.citations), reverse=True)

            self.log_message.emit(f"[INFO] 共抓取 {len(all_papers)} 篇论文")

            if self._stop_flag:
                break

            # 阶段2.5：Unpaywall / 网页补全仍缺摘要的论文
            missing_count = sum(1 for p in all_papers if not p.abstract and p.url)
            if missing_count > 0:
                self.progress.emit(base_pct + 55, f"{kw_prefix}补全剩余摘要（{missing_count}篇）...")
                self.log_message.emit(f"[INFO] OpenAlex 后仍缺摘要 {missing_count} 篇，尝试 Unpaywall/网页补全...")
                try:
                    filled = scrape_missing_abstracts(all_papers, log_fn=self.log_message.emit)
                    still = sum(1 for p in all_papers if not p.abstract)
                    self.log_message.emit(f"[INFO] 摘要补全完成，仍缺失 {still} 篇（无公开摘要）")
                except Exception as e:
                    self.log_message.emit(f"[WARNING] 摘要补全出错: {e}")

            if self._stop_flag:
                break

            # 篇关摘匹配位置标注（始终执行，过滤模式下同时移除不匹配论文）
            use_title = opts.get("filter_title", True)
            use_kw    = opts.get("filter_kw",    True)
            use_abs   = opts.get("filter_abs",   True)
            paper_matched = {}  # paper id → bool
            for p in all_papers:
                # 知网文献已在服务端做了主题词匹配（篇名+关键词+摘要），不再做本地过滤
                # 仍标注匹配位置，但始终视为"已匹配"
                if p.source == "cnki":
                    match_kw = keyword
                    _, reason = _match_bool_expr_fields(
                        match_kw,
                        p.title, p.abstract, p.keywords,
                        use_title, use_kw, use_abs,
                    )
                    paper_matched[id(p)] = True  # 知网文献不过滤
                    if reason:
                        p.match_reason = reason
                    elif use_abs and not (p.abstract or "").strip():
                        p.match_reason = "摘要缺失（知网已匹配）"
                    else:
                        p.match_reason = "知网主题词已匹配"
                else:
                    matched, reason = _match_bool_expr_fields(
                        en_keyword,
                        p.title, p.abstract, p.keywords,
                        use_title, use_kw, use_abs,
                    )
                    paper_matched[id(p)] = matched
                    if reason:
                        p.match_reason = reason
                    else:
                        if use_abs and not (p.abstract or "").strip():
                            p.match_reason = "摘要缺失"
                        else:
                            p.match_reason = "词项未出现"

            if opts.get("strict_search"):
                before = len(all_papers)
                all_papers = [p for p in all_papers if paper_matched.get(id(p))]
                removed = before - len(all_papers)
                fields_desc = "、".join(
                    f for f, flag in [("篇名", use_title), ("关键词", use_kw), ("摘要", use_abs)] if flag
                ) or "无"
                self.log_message.emit(f"[INFO] 篇关摘过滤（范围：{fields_desc}）：保留 {len(all_papers)} 篇，移除 {removed} 篇")
                if not all_papers:
                    self.log_message.emit(f"[WARNING] 篇关摘过滤后无剩余论文，跳过后续步骤")
                    continue

            # 阶段3：翻译
            if not opts.get("no_translate") and ANTHROPIC_API_KEY:
                self.progress.emit(base_pct + 60, f"{kw_prefix}翻译英文摘要...")
                try:
                    translate_papers(all_papers)
                    self.log_message.emit("[INFO] 翻译完成")
                except Exception as e:
                    self.log_message.emit(f"[WARNING] 翻译出错: {e}")
            elif not ANTHROPIC_API_KEY:
                self.log_message.emit("[INFO] 未配置 API Key，跳过翻译")

            if self._stop_flag:
                break

            # 阶段4：导出 Excel
            self.progress.emit(base_pct + 80, f"{kw_prefix}导出 Excel...")
            try:
                output_path = opts.get("output") or get_output_path(keyword)
                saved = export_to_excel(all_papers, output_path, keyword=keyword)
                if saved:
                    self.log_message.emit(f"[INFO] Excel 已保存: {saved}")
                    all_saved_paths.append(saved)
            except Exception as e:
                self.log_message.emit(f"[ERROR] Excel 导出失败: {e}")
                saved = ""

            self.papers_ready.emit(list(all_papers))

            # 阶段5：生成综述
            if not opts.get("no_review") and ANTHROPIC_API_KEY:
                n = len(all_papers)
                if n > 40:
                    self.progress.emit(base_pct + 90, f"{kw_prefix}生成文献综述（两阶段，共 {n} 篇）...")
                else:
                    self.progress.emit(base_pct + 90, f"{kw_prefix}生成文献综述...")
                try:
                    review_path = get_review_path(keyword)
                    review_text = generate_review(
                        all_papers, keyword=keyword,
                        output_path=review_path,
                        style=opts.get("review_style", "academic"),
                    )
                    if review_text and not review_text.startswith("[错误]"):
                        self.log_message.emit(f"[INFO] 综述已保存: {review_path}")
                except Exception as e:
                    self.log_message.emit(f"[WARNING] 综述生成失败: {e}")
            elif not ANTHROPIC_API_KEY and not opts.get("no_review"):
                self.log_message.emit("[INFO] 未配置 API Key，跳过综述生成")

        self.progress.emit(100, "完成！")
        first_path = all_saved_paths[0] if all_saved_paths else ""
        self.finished.emit(True, first_path)


# ─────────────────────────────────────────────────────────────────────────────
# 关键词 Chip Widget
# ─────────────────────────────────────────────────────────────────────────────

# 运算符颜色映射（与按钮配色一致）
OP_COLORS = {
    "*": ("#1565C0", "#E3F2FD"),   # 蓝色 AND
    "+": ("#2E7D32", "#E8F5E9"),   # 绿色 OR
    "-": ("#E65100", "#FFF3E0"),   # 橙色 NOT
    "(": ("#6A1B9A", "#F3E5F5"),   # 紫色 括号
    ")": ("#6A1B9A", "#F3E5F5"),
    "'": ("#880E4F", "#FCE4EC"),   # 粉色 引号
}


class OperatorLineEdit(QLineEdit):
    """在普通 QLineEdit 上叠加绘制运算符高亮（自定义 paintEvent）"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._op_chars = set(OP_COLORS.keys())

    def paintEvent(self, event):
        # 先用默认绘制（光标、选区、文字底层）
        super().paintEvent(event)

        txt = self.text()
        if not txt:
            return

        # 找到所有运算符位置
        op_positions = [(i, ch) for i, ch in enumerate(txt) if ch in self._op_chars]
        if not op_positions:
            return

        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        fm = self.fontMetrics()
        # 计算文字起始 x（受 margin/padding 影响，通过 QStyle 获取）
        cr = self.contentsRect()
        text_x = cr.x() + 4  # 与 QLineEdit 内部 margin 对齐

        # 用 fm.horizontalAdvance 计算每个字符的像素位置
        for idx, ch in op_positions:
            fg, bg = OP_COLORS[ch]
            # 字符前缀宽度
            x_start = text_x + fm.horizontalAdvance(txt[:idx])
            ch_w = fm.horizontalAdvance(ch)
            ch_h = fm.height()
            y_center = cr.y() + cr.height() // 2

            # 绘制背景圆角矩形
            rect = QRect(x_start - 1, y_center - ch_h // 2, ch_w + 2, ch_h)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(QColor(bg))
            painter.drawRoundedRect(rect, 3, 3)

            # 绘制彩色文字
            painter.setPen(QColor(fg))
            painter.setFont(self.font())
            painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, ch)

        painter.end()


class KeywordChip(QFrame):
    removed = pyqtSignal(str)

    def __init__(self, keyword: str, color: str = "#E3F2FD", parent=None):
        super().__init__(parent)
        self.keyword = keyword
        self.setStyleSheet(f"""
            QFrame {{
                background: {color};
                border: 1px solid #90CAF9;
                border-radius: 14px;
            }}
        """)
        layout = QHBoxLayout(self)
        layout.setContentsMargins(10, 4, 6, 4)
        layout.setSpacing(4)

        label = QLabel(keyword)
        label.setFont(QFont("SimSun", 13))
        label.setStyleSheet("color: #1565C0; border: none; background: transparent;")

        btn = QPushButton("×")
        btn.setFixedSize(18, 18)
        btn.setStyleSheet("""
            QPushButton {
                color: #1565C0; background: transparent; border: none;
                font-size: 14px; font-weight: bold; padding: 0;
            }
            QPushButton:hover { color: #B71C1C; }
        """)
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.clicked.connect(lambda: self.removed.emit(self.keyword))

        layout.addWidget(label)
        layout.addWidget(btn)
        self.setSizePolicy(SP.Policy.Fixed, SP.Policy.Fixed)


# ─────────────────────────────────────────────────────────────────────────────
# 主窗口
# ─────────────────────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.keywords: list[str] = []
        self.worker: WorkerThread | None = None
        self.last_output_path: str = ""

        # 知网嵌入式浏览器管理器（主线程创建）
        if WEBENGINE_AVAILABLE:
            self.cnki_manager = CnkiManager(self)
            self.cnki_manager.status_update.connect(self._on_cnki_status)
            self.cnki_manager.captcha_detected.connect(self._on_cnki_captcha)
            self.cnki_manager.download_started.connect(self._on_cnki_download_started)
            self.cnki_manager.download_done.connect(self._on_cnki_download_done)
            self.cnki_manager.download_failed.connect(self._on_cnki_download_failed)
        else:
            self.cnki_manager = None

        self.setWindowTitle("经济学论文追踪系统")
        self.setMinimumSize(960, 700)
        self.resize(1100, 800)

        self._init_ui()
        self._load_settings()

    def _init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root_layout = QVBoxLayout(central)
        root_layout.setContentsMargins(12, 8, 12, 8)
        root_layout.setSpacing(8)

        title_row = QHBoxLayout()
        title_label = QLabel("经济学论文追踪系统")
        title_label.setFont(QFont("SimSun", 16, QFont.Weight.Bold))
        title_label.setStyleSheet("color: #1A237E; padding: 4px 0;")
        journal_btn = QPushButton("📋 收录期刊")
        journal_btn.setFixedHeight(28)
        journal_btn.setFixedWidth(100)
        journal_btn.setStyleSheet("""
            QPushButton {
                background: #E3F2FD; color: #1565C0;
                border: 1px solid #90CAF9; border-radius: 4px;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px;
            }
            QPushButton:hover { background: #BBDEFB; }
        """)
        journal_btn.clicked.connect(self._show_journal_info)

        syntax_btn = QPushButton("❓ 语法说明")
        syntax_btn.setFixedHeight(28)
        syntax_btn.setFixedWidth(100)
        syntax_btn.setStyleSheet("""
            QPushButton {
                background: #F3E5F5; color: #6A1B9A;
                border: 1px solid #CE93D8; border-radius: 4px;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px;
            }
            QPushButton:hover { background: #E1BEE7; }
        """)
        syntax_btn.clicked.connect(self._show_syntax_help)

        search_info_btn = QPushButton("📖 搜索说明")
        search_info_btn.setFixedHeight(28)
        search_info_btn.setFixedWidth(100)
        search_info_btn.setStyleSheet("""
            QPushButton {
                background: #E3F2FD; color: #1565C0;
                border: 1px solid #90CAF9; border-radius: 4px;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px;
            }
            QPushButton:hover { background: #BBDEFB; }
        """)
        search_info_btn.clicked.connect(self._show_search_info)

        feedback_btn = QPushButton("🐛 反馈Bug")
        feedback_btn.setFixedHeight(28)
        feedback_btn.setFixedWidth(100)
        feedback_btn.setStyleSheet("""
            QPushButton {
                background: #FFF3E0; color: #E65100;
                border: 1px solid #FFCC80; border-radius: 4px;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px;
            }
            QPushButton:hover { background: #FFE0B2; }
        """)
        feedback_btn.clicked.connect(lambda: FeedbackDialog(self).exec())

        manual_btn = QPushButton("📖 使用说明")
        manual_btn.setFixedHeight(28)
        manual_btn.setFixedWidth(100)
        manual_btn.setStyleSheet("""
            QPushButton {
                background: #E8F5E9; color: #2E7D32;
                border: 1px solid #A5D6A7; border-radius: 4px;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px;
            }
            QPushButton:hover { background: #C8E6C9; }
        """)
        manual_btn.clicked.connect(self._show_manual)

        title_row.addWidget(title_label)
        title_row.addStretch()
        title_row.addWidget(manual_btn)
        title_row.addWidget(feedback_btn)
        title_row.addWidget(search_info_btn)
        title_row.addWidget(syntax_btn)
        title_row.addWidget(journal_btn)
        root_layout.addLayout(title_row)

        splitter = QSplitter(Qt.Orientation.Vertical)
        splitter.setChildrenCollapsible(False)
        root_layout.addWidget(splitter, stretch=1)

        top_widget = QWidget()
        top_layout = QVBoxLayout(top_widget)
        top_layout.setContentsMargins(0, 0, 0, 0)
        top_layout.setSpacing(8)

        panel_layout = QHBoxLayout()
        panel_layout.setSpacing(12)
        panel_layout.addWidget(self._build_keyword_area(), stretch=1)
        panel_layout.addWidget(self._build_params_area(), stretch=1)
        top_layout.addLayout(panel_layout)
        top_layout.addWidget(self._build_run_area())
        splitter.addWidget(top_widget)
        splitter.addWidget(self._build_output_area())
        splitter.setSizes([340, 360])

        self.setStyleSheet("""
            QMainWindow { background: #F5F5F5; }
            QGroupBox {
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px; font-weight: bold;
                color: #37474F; border: 1px solid #CFD8DC; border-radius: 6px;
                margin-top: 8px; padding-top: 4px;
            }
            QGroupBox::title { subcontrol-origin: margin; left: 12px; padding: 0 4px; }
            QComboBox, QLineEdit {
                border: 1px solid #B0BEC5; border-radius: 4px; padding: 4px 8px;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px;
                background: white; color: #212121;
            }
            QComboBox:focus, QLineEdit:focus { border-color: #1976D2; }
            QComboBox QAbstractItemView {
                background: white; color: #212121;
                selection-background-color: #BBDEFB; selection-color: #212121;
            }
            QCheckBox { font-family: 'SimSun', 'Times New Roman'; font-size: 13px; spacing: 6px; color: #212121; }
            QLabel { font-family: 'SimSun', 'Times New Roman'; font-size: 13px; color: #212121; }
        """)

    def _build_keyword_area(self) -> QGroupBox:
        group = QGroupBox("搜索关键词")
        layout = QVBoxLayout(group)
        layout.setSpacing(8)

        hint = QLabel('在下方框内输入关键词，回车添加；支持布尔运算符（点右上角"语法说明"了解用法）')
        hint.setStyleSheet("color: #78909C; font-size: 13px;")
        layout.addWidget(hint)

        # 运算符快捷按钮
        op_row = QHBoxLayout()
        op_row.setSpacing(5)
        op_row.setContentsMargins(0, 0, 0, 0)
        op_label = QLabel("插入运算符：")
        op_label.setStyleSheet("color: #78909C; font-size: 13px; font-family: 'SimSun', 'Times New Roman';")
        op_row.addWidget(op_label)

        OP_STYLE = """
            QPushButton {{
                background: {bg}; color: {fg};
                border: 1px solid {border}; border-radius: 3px;
                font-family: Consolas; font-size: 13pt; font-weight: bold;
                padding: 1px 8px; min-width: 28px; max-height: 22px;
            }}
            QPushButton:hover {{ background: {hover}; }}
            QPushButton:pressed {{ background: {press}; }}
        """
        ops = [
            ("*",  "AND（且）",   "#E3F2FD", "#1565C0", "#90CAF9", "#BBDEFB", "#1976D2"),
            ("+",  "OR（或）",    "#E8F5E9", "#2E7D32", "#A5D6A7", "#C8E6C9", "#388E3C"),
            ("-",  "NOT（非）",   "#FFF3E0", "#E65100", "#FFCC80", "#FFE0B2", "#F57C00"),
            ("()", "括号分组",    "#F3E5F5", "#6A1B9A", "#CE93D8", "#E1BEE7", "#7B1FA2"),
            ("''", "精确短语",    "#FCE4EC", "#880E4F", "#F48FB1", "#F8BBD0", "#C2185B"),
        ]
        for symbol, tip, bg, fg, border, hover, press in ops:
            btn = QPushButton(symbol)
            btn.setToolTip(tip)
            btn.setFixedHeight(22)
            btn.setStyleSheet(OP_STYLE.format(bg=bg, fg=fg, border=border, hover=hover, press=press))
            def _make_inserter(sym):
                def _insert():
                    inp = self.kw_input
                    pos = inp.cursorPosition()
                    txt = inp.text()
                    if sym == "()":
                        inp.setText(txt[:pos] + "()" + txt[pos:])
                        inp.setCursorPosition(pos + 1)
                    elif sym == "''":
                        inp.setText(txt[:pos] + "''" + txt[pos:])
                        inp.setCursorPosition(pos + 1)
                    else:
                        left  = txt[:pos].rstrip(" ")
                        right = txt[pos:].lstrip(" ")
                        insert = f" {sym} "
                        inp.setText(left + insert + right)
                        inp.setCursorPosition(len(left) + len(insert))
                    inp.setFocus()
                return _insert
            btn.clicked.connect(_make_inserter(symbol))
            op_row.addWidget(btn)

        op_row.addStretch()
        layout.addLayout(op_row)

        # chip 区域（内嵌输入框）
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setMinimumHeight(80)
        scroll.setStyleSheet("QScrollArea { border: 1px solid #BBDEFB; border-radius: 4px; background: #FAFAFA; }")

        self.chip_container = QWidget()
        self.chip_container.setStyleSheet("background: #FAFAFA;")
        self.chip_layout = FlowLayout(self.chip_container, h_spacing=6, v_spacing=6)
        self.chip_layout.setContentsMargins(8, 8, 8, 8)

        # 内联输入框——嵌在 chip 区域末尾
        self.kw_input = OperatorLineEdit()
        self.kw_input.setPlaceholderText("输入关键词，按回车添加…")
        self.kw_input.setMinimumWidth(180)
        self.kw_input.setFixedHeight(30)
        self.kw_input.setStyleSheet("""
            QLineEdit {
                border: none; border-bottom: 2px solid #90CAF9;
                background: transparent; font-family: 'SimSun', 'Times New Roman';
                font-size: 13px; color: #212121; padding: 0 4px;
            }
            QLineEdit:focus { border-bottom: 2px solid #1976D2; }
        """)
        self.kw_input.returnPressed.connect(self._add_keyword)
        self.kw_input.textChanged.connect(lambda: self.kw_input.update())
        self.chip_layout.addWidget(self.kw_input)

        scroll.setWidget(self.chip_container)
        layout.addWidget(scroll, stretch=1)
        return group

    def _build_params_area(self) -> QWidget:
        # 外层用 QWidget + QVBoxLayout，内部三个 QGroupBox 上下排列
        outer = QWidget()
        outer_layout = QVBoxLayout(outer)
        outer_layout.setContentsMargins(0, 0, 0, 0)
        outer_layout.setSpacing(6)

        # ── 帮助函数：在 QGridLayout 里添加 label+widget 一行 ──────────
        def add_row(lyt, label_text, widget, r):
            lbl = QLabel(label_text)
            lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            lyt.addWidget(lbl, r, 0)
            lyt.addWidget(widget, r, 1)

        # ══════════════════════════════════════════════════════════════════
        # 检索源切换栏（英文文献 / 中文文献 toggle tabs）
        # ══════════════════════════════════════════════════════════════════
        src_bar = QWidget()
        src_bar.setObjectName("srcBar")
        src_bar.setStyleSheet("""
            QWidget#srcBar {
                background: #F5F5F5;
                border: 1px solid #E0E0E0;
                border-radius: 6px;
            }
        """)
        src_bar_layout = QHBoxLayout(src_bar)
        src_bar_layout.setContentsMargins(6, 4, 6, 4)
        src_bar_layout.setSpacing(0)

        lbl_src = QLabel("检索源：")
        lbl_src.setStyleSheet("font-family: 'SimSun','Times New Roman'; font-size: 13pt; color: #424242; background: transparent;")
        src_bar_layout.addWidget(lbl_src)
        src_bar_layout.addSpacing(4)

        def _make_src_btn(text, icon_char, active_color, active_text_color="#FFFFFF"):
            btn = QPushButton(f"  {icon_char}  {text}")
            btn.setCheckable(True)
            btn.setFixedHeight(28)
            btn.setStyleSheet(f"""
                QPushButton {{
                    font-family: 'SimSun','Times New Roman'; font-size: 13pt;
                    border: 1px solid #BDBDBD; border-radius: 5px;
                    padding: 0 12px; background: #FFFFFF; color: #424242;
                }}
                QPushButton:checked {{
                    background: {active_color}; color: {active_text_color};
                    border: 1px solid {active_color};
                }}
                QPushButton:hover:!checked {{
                    background: #EEEEEE;
                }}
            """)
            return btn

        self.src_en_btn = _make_src_btn("英文文献", "⊕", "#1565C0")
        self.src_cn_btn = _make_src_btn("中文文献", "🇨🇳", "#C62828")
        self.src_en_btn.setChecked(True)

        src_bar_layout.addWidget(self.src_en_btn)
        src_bar_layout.addSpacing(4)
        src_bar_layout.addWidget(self.src_cn_btn)
        src_bar_layout.addStretch()
        outer_layout.addWidget(src_bar)

        # QStackedWidget：page 0 = 英文，page 1 = 中文
        self._src_stack = QStackedWidget()
        outer_layout.addWidget(self._src_stack)

        def _on_src_en():
            self.src_en_btn.setChecked(True)
            self.src_cn_btn.setChecked(False)
            self._src_stack.setCurrentIndex(0)

        def _on_src_cn():
            self.src_cn_btn.setChecked(True)
            self.src_en_btn.setChecked(False)
            self._src_stack.setCurrentIndex(1)

        self.src_en_btn.clicked.connect(lambda: _on_src_en())
        self.src_cn_btn.clicked.connect(lambda: _on_src_cn())

        # ══════════════════════════════════════════════════════════════════
        # 组 1：英文期刊搜索设置（放入 stack page 0）
        # ══════════════════════════════════════════════════════════════════
        en_page = QWidget()
        en_page_layout = QVBoxLayout(en_page)
        en_page_layout.setContentsMargins(0, 0, 0, 0)
        en_page_layout.setSpacing(0)
        en_group = QGroupBox("英文期刊搜索设置")
        en_layout = QGridLayout(en_group)
        en_layout.setSpacing(6)
        en_layout.setColumnStretch(1, 1)
        en_row = 0

        # 期刊范围行：下拉 + 自定义期刊按钮
        self._en_selected_issns: set = set()   # 空 = 未自定义，按 scope_combo 决定
        self.scope_combo = QComboBox()
        self.scope_combo.addItems(SCOPE_LABELS)
        scope_row = QHBoxLayout()
        scope_row.setContentsMargins(0, 0, 0, 0)
        scope_row.addWidget(self.scope_combo, stretch=1)
        scope_row.addSpacing(8)
        self.en_journal_btn = QPushButton("自定义期刊…")
        self.en_journal_btn.setFixedHeight(26)
        self.en_journal_btn.setToolTip("手动勾选要检索的具体英文期刊（覆盖上方期刊范围设置）")
        self.en_journal_btn.clicked.connect(self._on_en_journal_select)
        scope_row.addWidget(self.en_journal_btn)
        lbl_scope = QLabel("期刊范围：")
        lbl_scope.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        en_layout.addWidget(lbl_scope, en_row, 0)
        scope_widget = QWidget(); scope_widget.setLayout(scope_row)
        en_layout.addWidget(scope_widget, en_row, 1); en_row += 1

        # 搜索年限 + 搜索模式 同行
        years_mode_row = QHBoxLayout()
        years_mode_row.setContentsMargins(0, 0, 0, 0)
        self.years_combo = QComboBox()
        self.years_combo.addItems(YEARS_LABELS)
        self.years_combo.setCurrentIndex(1)
        years_mode_row.addWidget(self.years_combo, stretch=1)
        years_mode_row.addSpacing(12)
        years_mode_row.addWidget(QLabel("搜索模式："))
        self.mode_combo = QComboBox()
        self.mode_combo.addItems(MODE_LABELS)
        years_mode_row.addWidget(self.mode_combo, stretch=1)
        lbl_years = QLabel("检索年限：")
        lbl_years.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        en_layout.addWidget(lbl_years, en_row, 0)
        years_mode_widget = QWidget(); years_mode_widget.setLayout(years_mode_row)
        en_layout.addWidget(years_mode_widget, en_row, 1); en_row += 1

        # 通用 checkbox 样式：未选灰框，已选绿底白钩
        CB_STYLE = f"""
            QCheckBox {{
                color: #212121; font-family: 'SimSun', 'Times New Roman'; font-size: 13pt;
                spacing: 6px;
            }}
            QCheckBox::indicator {{
                width: 16px; height: 16px;
                border: 2px solid #9E9E9E; border-radius: 3px; background: #FFFFFF;
            }}
            QCheckBox::indicator:checked {{
                background: #2E7D32; border: 2px solid #2E7D32;
                image: url("{_CHECK_SVG_PATH}");
            }}
            QCheckBox::indicator:hover {{ border-color: #1565C0; }}
        """

        # 每刊抓取 + 选项 同行
        max_opts_row = QHBoxLayout()
        max_opts_row.setContentsMargins(0, 0, 0, 0)
        max_opts_row.addWidget(QLabel("每刊抓取："))
        self.max_spin = QSpinBox()
        self.max_spin.setRange(1, 9999)
        self.max_spin.setValue(50)
        self.max_spin.setSuffix(" 篇")
        self.max_spin.setFixedWidth(80)
        max_opts_row.addWidget(self.max_spin)
        max_opts_row.addSpacing(16)
        self.translate_cb = QCheckBox("翻译摘要")
        self.translate_cb.setChecked(True)
        self.translate_cb.setStyleSheet(CB_STYLE)
        # review_cb 和 brief_cb 创建在此处供 _build_opts 引用，但实际添加到运行区进度条右侧
        self.review_cb = QCheckBox("生成综述")
        self.review_cb.setChecked(True)
        self.review_cb.setStyleSheet(CB_STYLE)
        self.brief_cb = QCheckBox("简洁综述")
        self.brief_cb.setStyleSheet(CB_STYLE)
        max_opts_row.addWidget(self.translate_cb)
        max_opts_row.addStretch()
        lbl_opt = QLabel("选项：")
        lbl_opt.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        en_layout.addWidget(lbl_opt, en_row, 0)
        max_opts_widget = QWidget(); max_opts_widget.setLayout(max_opts_row)
        en_layout.addWidget(max_opts_widget, en_row, 1)

        en_page_layout.addWidget(en_group)
        self._src_stack.addWidget(en_page)   # stack index 0

        # ══════════════════════════════════════════════════════════════════
        # 组 2：中文期刊搜索设置（放入 stack page 1）
        # ══════════════════════════════════════════════════════════════════
        cn_page = QWidget()
        cn_page_layout = QVBoxLayout(cn_page)
        cn_page_layout.setContentsMargins(0, 0, 0, 0)
        cn_page_layout.setSpacing(0)
        cn_group = QGroupBox("中文期刊搜索设置")
        cn_layout = QGridLayout(cn_group)
        cn_layout.setSpacing(6)
        cn_layout.setColumnStretch(1, 1)
        cn_row = 0

        # 行1：启用开关 + 抓取数量
        cn_enable_row = QHBoxLayout()
        self.cnki_cb = QCheckBox("搜索中文文献（知网）")
        self.cnki_cb.setChecked(False)
        self.cnki_cb.setStyleSheet(CB_STYLE)
        cn_enable_row.addWidget(self.cnki_cb)
        cn_enable_row.addSpacing(16)
        cn_enable_row.addWidget(QLabel("每刊抓取："))
        self.cnki_max_spin = QSpinBox()
        self.cnki_max_spin.setRange(1, 9999)
        self.cnki_max_spin.setValue(20)
        self.cnki_max_spin.setSuffix(" 篇/刊")
        self.cnki_max_spin.setFixedWidth(90)
        cn_enable_row.addWidget(self.cnki_max_spin)
        cn_enable_row.addStretch()
        lbl_cn1 = QLabel("启用：")
        lbl_cn1.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        cn_layout.addWidget(lbl_cn1, cn_row, 0)
        cn_enable_widget = QWidget(); cn_enable_widget.setLayout(cn_enable_row)
        cn_layout.addWidget(cn_enable_widget, cn_row, 1); cn_row += 1

        # 行2：期刊选择器按钮 + 登录按钮
        cn_journal_row = QHBoxLayout()
        self._cnki_selected_journals = list(CNKI_ALL_JOURNALS)  # 默认全选
        self.cnki_journal_btn = QPushButton(
            f"选择期刊 ({len(self._cnki_selected_journals)}/{len(CNKI_ALL_JOURNALS)})"
        )
        self.cnki_journal_btn.setFixedHeight(26)
        self.cnki_journal_btn.clicked.connect(self._on_cnki_journal_select)
        cn_journal_row.addWidget(self.cnki_journal_btn)
        cn_journal_row.addSpacing(16)
        self.cnki_login_btn = QPushButton("知网登录")
        self.cnki_login_btn.setFixedHeight(26)
        self.cnki_login_btn.clicked.connect(self._on_cnki_login)
        cn_journal_row.addWidget(self.cnki_login_btn)
        cn_journal_row.addStretch()
        lbl_cn2 = QLabel("期刊/登录：")
        lbl_cn2.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        cn_layout.addWidget(lbl_cn2, cn_row, 0)
        cn_journal_widget = QWidget(); cn_journal_widget.setLayout(cn_journal_row)
        cn_layout.addWidget(cn_journal_widget, cn_row, 1); cn_row += 1

        # 行3：知网排序方式 + 自动下载 PDF
        cn_sort_row = QHBoxLayout()
        self.cnki_sort_combo = QComboBox()
        for label, val in [("发表时间", "PT"), ("被引次数", "CF"), ("相关度", "FFD"), ("下载次数", "DFR"), ("综合", "ZH")]:
            self.cnki_sort_combo.addItem(label, val)
        self.cnki_sort_combo.setCurrentIndex(0)  # 默认发表时间
        cn_sort_row.addWidget(self.cnki_sort_combo)
        cn_sort_row.addSpacing(20)
        self.cnki_auto_dl_cb = QCheckBox("搜索完成后自动下载全部 PDF")
        self.cnki_auto_dl_cb.setChecked(False)
        self.cnki_auto_dl_cb.setStyleSheet(CB_STYLE)
        self.cnki_auto_dl_cb.setToolTip(
            "勾选后，程序在完成篇关摘筛选之后，自动下载全部中文文献的 PDF 全文。\n"
            "需已完成知网登录，且所在机构对对应期刊有下载权限。"
        )
        cn_sort_row.addWidget(self.cnki_auto_dl_cb)
        cn_sort_row.addStretch()
        lbl_cn3 = QLabel("知网排序：")
        lbl_cn3.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        cn_layout.addWidget(lbl_cn3, cn_row, 0)
        cn_sort_widget = QWidget(); cn_sort_widget.setLayout(cn_sort_row)
        cn_layout.addWidget(cn_sort_widget, cn_row, 1)

        if not WEBENGINE_AVAILABLE:
            self.cnki_cb.setEnabled(False)
            self.cnki_max_spin.setEnabled(False)
            self.cnki_journal_btn.setEnabled(False)
            self.cnki_login_btn.setEnabled(False)
            self.cnki_sort_combo.setEnabled(False)
            self.cnki_auto_dl_cb.setEnabled(False)
            self.cnki_cb.setToolTip("需安装 PyQt6-WebEngine 才能使用此功能")

        cn_page_layout.addWidget(cn_group)
        cn_page_layout.addStretch()
        self._src_stack.addWidget(cn_page)   # stack index 1

        # ══════════════════════════════════════════════════════════════════
        # 组 3：模型设置
        # ══════════════════════════════════════════════════════════════════
        model_group = QGroupBox("模型设置")
        model_layout = QGridLayout(model_group)
        model_layout.setSpacing(6)
        model_layout.setColumnStretch(1, 1)
        m_row = 0

        self.api_type_combo = QComboBox()
        self.api_type_combo.addItem("Anthropic（Claude）",   "anthropic")
        self.api_type_combo.addItem("OpenAI 兼容（DeepSeek / Gemini / 豆包等）", "openai")
        add_row(model_layout, "接口类型：", self.api_type_combo, m_row); m_row += 1

        # API Key + 保存按钮
        api_row = QHBoxLayout()
        api_row.setContentsMargins(0, 0, 0, 0)
        self.api_key_input = QLineEdit()
        self.api_key_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.api_key_input.setPlaceholderText("sk-ant-... （翻译/综述必填）")
        self.api_key_input.setFixedHeight(28)
        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(52); save_btn.setFixedHeight(28)
        save_btn.setStyleSheet("""
            QPushButton { background: #43A047; color: white; border-radius: 4px; font-size: 13px; }
            QPushButton:hover { background: #388E3C; }
        """)
        save_btn.clicked.connect(self._save_settings)
        api_row.addWidget(self.api_key_input); api_row.addWidget(save_btn)
        lbl_key = QLabel("API Key：")
        lbl_key.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        model_layout.addWidget(lbl_key, m_row, 0)
        api_widget = QWidget(); api_widget.setLayout(api_row)
        api_widget.setMinimumHeight(34)
        model_layout.addWidget(api_widget, m_row, 1); m_row += 1

        self.base_url_input = QLineEdit()
        add_row(model_layout, "API 地址：", self.base_url_input, m_row); m_row += 1

        self.model_fast_combo = QComboBox()
        self.model_fast_combo.setEditable(True)
        self.model_fast_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.model_fast_combo.lineEdit().setPlaceholderText("翻译模型（留空使用默认）")
        add_row(model_layout, "翻译模型：", self.model_fast_combo, m_row); m_row += 1

        self.model_strong_combo = QComboBox()
        self.model_strong_combo.setEditable(True)
        self.model_strong_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.model_strong_combo.lineEdit().setPlaceholderText("综述模型（留空使用默认）")
        add_row(model_layout, "综述模型：", self.model_strong_combo, m_row)

        # 各接口类型的预设模型列表：(显示名, model_id)
        _MODEL_OPTIONS = {
            "anthropic": [
                ("（默认）claude-haiku-4-5",          "claude-haiku-4-5-20251001"),
                ("claude-sonnet-4-6（综述推荐）",      "claude-sonnet-4-6"),
                ("claude-opus-4-6（最强）",            "claude-opus-4-6"),
                ("claude-haiku-3-5",                   "claude-haiku-3-5-20241022"),
                ("claude-sonnet-3-7",                  "claude-sonnet-3-7-20250219"),
            ],
            "openai": [
                ("deepseek-chat（V3，翻译推荐）",      "deepseek-chat"),
                ("deepseek-reasoner（R1，综述推荐）",  "deepseek-reasoner"),
                ("gemini-2.0-flash",                   "gemini-2.0-flash"),
                ("gemini-1.5-pro",                     "gemini-1.5-pro"),
                ("gpt-4o-mini",                        "gpt-4o-mini"),
                ("gpt-4o",                             "gpt-4o"),
                ("doubao-pro-4k",                      "doubao-pro-4k"),
            ],
        }

        def _update_model_combos(idx=None):
            atype = self.api_type_combo.currentData()
            options = _MODEL_OPTIONS.get(atype, [])
            for combo in (self.model_fast_combo, self.model_strong_combo):
                cur = combo.currentText().strip()
                combo.blockSignals(True)
                combo.clear()
                combo.addItem("（留空使用默认）", "")
                for label, val in options:
                    combo.addItem(label, val)
                # 恢复之前的值：优先按 data 匹配，其次直接填入自定义文本
                matched = False
                for i in range(combo.count()):
                    if combo.itemData(i) == cur or combo.itemText(i) == cur:
                        combo.setCurrentIndex(i)
                        matched = True
                        break
                if not matched and cur:
                    combo.setCurrentText(cur)
                combo.blockSignals(False)
            # 更新占位符
            if atype == "openai":
                self.api_key_input.setPlaceholderText("API Key（如 sk-xxx）")
                self.base_url_input.setPlaceholderText(
                    "API 地址，如 https://api.deepseek.com  /  https://api.doubao.com"
                )
            else:
                self.api_key_input.setPlaceholderText("sk-ant-... （翻译/综述必填）")
                self.base_url_input.setPlaceholderText("留空=原生Anthropic，中转站填入如 https://xxx.com")

        self.api_type_combo.currentIndexChanged.connect(_update_model_combos)
        _update_model_combos()

        outer_layout.addWidget(model_group)

        return outer

    def _on_en_journal_select(self):
        # 若还没自定义过，默认按当前 scope_combo 全选对应期刊
        if not self._en_selected_issns:
            scope_idx = self.scope_combo.currentIndex()
            if scope_idx == 0:
                default = set({**ENGLISH_JOURNALS, **MANAGEMENT_JOURNALS, **ECONOMICS_JOURNALS}.keys())
            elif scope_idx == 2:
                default = {k for k, v in ENGLISH_JOURNALS.items() if v["tier"] == "Top5"}
            elif scope_idx == 3:
                default = set(MANAGEMENT_JOURNALS.keys())
            elif scope_idx == 4:
                default = set(ECONOMICS_JOURNALS.keys())
            else:
                default = set(ENGLISH_JOURNALS.keys())
        else:
            default = self._en_selected_issns

        dlg = EnJournalDialog(default, self)
        if dlg.exec():
            self._en_selected_issns = dlg.get_selected_issns()
            n = len(self._en_selected_issns)
            total = sum(len(js) for tiers in _EN_JOURNAL_TREE.values() for js in tiers.values())
            if n == 0:
                self.en_journal_btn.setText("自定义期刊…")
                # 清空自定义，scope_combo 恢复原有意义
                if self.scope_combo.count() > len(SCOPE_LABELS):
                    self.scope_combo.removeItem(len(SCOPE_LABELS))
            else:
                self.en_journal_btn.setText(f"自定义期刊 ({n}/{total}) ✓")
                # 切换到/更新"自定义"选项
                custom_label = f"自定义（{n} 本）"
                if self.scope_combo.count() > len(SCOPE_LABELS):
                    self.scope_combo.setItemText(len(SCOPE_LABELS), custom_label)
                else:
                    self.scope_combo.addItem(custom_label)
                self.scope_combo.setCurrentIndex(len(SCOPE_LABELS))

    def _on_cnki_journal_select(self):
        dlg = CnkiJournalDialog(self._cnki_selected_journals, self)
        if dlg.exec():
            self._cnki_selected_journals = dlg.get_selected()
            n = len(self._cnki_selected_journals)
            self.cnki_journal_btn.setText(f"选择期刊 ({n}/{len(CNKI_ALL_JOURNALS)})")

    def _on_cnki_login(self):
        if not WEBENGINE_AVAILABLE or self.cnki_manager is None:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "提示", "需安装 PyQt6-WebEngine 才能使用此功能。\n运行: pip install PyQt6-WebEngine")
            return
        dlg = CnkiLoginDialog(self.cnki_manager, self)
        if dlg.exec():
            self.cnki_manager.login_confirmed = True
            self.cnki_login_btn.setText("已登录 ✓")
            self.cnki_login_btn.setStyleSheet("QPushButton { background: #C8E6C9; color: #1B5E20; border: 1px solid #81C784; border-radius: 3px; }")

    def _build_run_area(self) -> QWidget:
        w = QWidget()
        layout = QHBoxLayout(w)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(12)

        self.start_btn = QPushButton("▶  开始抓取")
        self.start_btn.setFixedHeight(40)
        self.start_btn.setMinimumWidth(140)
        self.start_btn.setFont(QFont("SimSun", 13, QFont.Weight.Bold))
        self._set_start_btn_style(running=False)
        self.start_btn.clicked.connect(self._on_start_stop)

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(20)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat("等待开始")
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 1px solid #B0BEC5; border-radius: 12px; background: #ECEFF1;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px; color: #37474F; text-align: center;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #42A5F5, stop:1 #1565C0);
                border-radius: 12px;
            }
        """)

        FILTER_SWITCH_STYLE = """
            QCheckBox {
                color: #757575;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13pt; font-weight: bold;
                spacing: 4px;
                padding: 3px 8px;
                border: 1.5px solid #BDBDBD;
                border-radius: 4px;
                background: #F5F5F5;
            }
            QCheckBox:checked {
                color: #FFFFFF;
                border: 1.5px solid #1565C0;
                background: #1976D2;
            }
            QCheckBox::indicator { width: 0px; height: 0px; }
        """
        FILTER_FIELD_STYLE = """
            QCheckBox {
                color: #9E9E9E;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13pt;
                spacing: 4px;
                padding: 3px 8px;
                border: 1.5px solid #E0E0E0;
                border-radius: 4px;
                background: #FAFAFA;
            }
            QCheckBox:checked {
                color: #FFFFFF;
                border: 1.5px solid #2E7D32;
                background: #43A047;
            }
            QCheckBox:disabled {
                color: #BDBDBD;
                border: 1.5px solid #EEEEEE;
                background: #F5F5F5;
            }
            QCheckBox::indicator { width: 0px; height: 0px; }
        """

        self.strict_search_cb = QCheckBox("篇关摘过滤")
        self.strict_search_cb.setToolTip("启用后仅保留在勾选字段中出现检索词的论文")
        self.strict_search_cb.setStyleSheet(FILTER_SWITCH_STYLE)

        self.filter_title_cb = QCheckBox("篇名")
        self.filter_title_cb.setChecked(True)
        self.filter_kw_cb    = QCheckBox("关键词")
        self.filter_kw_cb.setChecked(True)
        self.filter_abs_cb   = QCheckBox("摘要")
        self.filter_abs_cb.setChecked(True)
        for cb in (self.filter_title_cb, self.filter_kw_cb, self.filter_abs_cb):
            cb.setStyleSheet(FILTER_FIELD_STYLE)

        def _toggle_filter_fields(state):
            enabled = bool(state)
            for cb in (self.filter_title_cb, self.filter_kw_cb, self.filter_abs_cb):
                cb.setEnabled(enabled)
        self.strict_search_cb.stateChanged.connect(_toggle_filter_fields)
        _toggle_filter_fields(self.strict_search_cb.checkState())

        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(6)
        filter_layout.setContentsMargins(0, 0, 0, 0)
        filter_layout.addWidget(self.strict_search_cb)
        filter_layout.addWidget(self.filter_title_cb)
        filter_layout.addWidget(self.filter_kw_cb)
        filter_layout.addWidget(self.filter_abs_cb)
        filter_widget = QWidget()
        filter_widget.setLayout(filter_layout)

        layout.addWidget(self.start_btn)
        layout.addWidget(filter_widget)
        layout.addWidget(self.progress_bar, stretch=1)
        layout.addWidget(self.review_cb)
        layout.addWidget(self.brief_cb)
        return w

    def _build_output_area(self) -> QTabWidget:
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #CFD8DC; border-radius: 4px; background: white; }
            QTabBar::tab {
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px; padding: 6px 16px;
                background: #ECEFF1; border: 1px solid #CFD8DC; border-bottom: none;
                border-radius: 4px 4px 0 0; margin-right: 2px;
            }
            QTabBar::tab:selected { background: white; color: #1565C0; font-weight: bold; }
        """)
        self.tab_widget.addTab(self._build_log_tab(), "📋  运行日志")
        self.tab_widget.addTab(self._build_preview_tab(), "📄  结果预览")
        return self.tab_widget

    def _build_log_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 11))
        self.log_text.setStyleSheet("""
            QTextEdit { background: #1E2A3A; color: #CFD8DC; border: none; border-radius: 4px; padding: 4px; }
        """)
        layout.addWidget(self.log_text, stretch=1)

        btn_row = QHBoxLayout()
        clear_btn = QPushButton("清空日志")
        clear_btn.setFixedHeight(28)
        clear_btn.setStyleSheet("QPushButton { font-family:'SimSun', 'Times New Roman'; font-size:13px; }")
        clear_btn.clicked.connect(self.log_text.clear)

        self.open_folder_btn = QPushButton("📂  打开输出文件夹")
        self.open_folder_btn.setFixedHeight(28)
        self.open_folder_btn.setEnabled(False)
        self.open_folder_btn.setStyleSheet("""
            QPushButton {
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px;
                background: #E8F5E9; border: 1px solid #A5D6A7; border-radius: 4px;
            }
            QPushButton:hover { background: #C8E6C9; }
            QPushButton:disabled { background: #ECEFF1; color: #90A4AE; }
        """)
        self.open_folder_btn.clicked.connect(self._open_output_folder)

        btn_row.addWidget(clear_btn)
        btn_row.addStretch()
        btn_row.addWidget(self.open_folder_btn)
        layout.addLayout(btn_row)
        return w

    def _build_preview_tab(self) -> QWidget:
        w = QWidget()
        layout = QVBoxLayout(w)
        layout.setContentsMargins(8, 8, 8, 8)
        info = QLabel("完成抓取后自动显示全部结果预览")
        info.setStyleSheet("color: #78909C; font-size: 13px;")
        layout.addWidget(info)

        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.preview_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.preview_table.setStyleSheet("""
            QTableWidget { font-family: 'SimSun', 'Times New Roman'; font-size: 13px; gridline-color: #ECEFF1; }
            QHeaderView::section {
                background: #E3F2FD; font-family: 'SimSun', 'Times New Roman'; font-size: 13px;
                font-weight: bold; padding: 4px; border: 1px solid #BBDEFB;
            }
        """)
        layout.addWidget(self.preview_table, stretch=1)

        # ── 知网 PDF 下载工具栏 ──────────────────────────────────────────────────
        dl_row = QHBoxLayout()
        dl_row.setContentsMargins(0, 4, 0, 0)
        self.cnki_dl_btn = QPushButton("⬇  下载选中论文 PDF（知网）")
        self.cnki_dl_btn.setFixedHeight(28)
        self.cnki_dl_btn.setToolTip("选中结果预览表格中的行（知网来源），下载 PDF 全文\n需已完成知网登录，且所在机构有下载权限")
        self.cnki_dl_btn.setEnabled(False)
        self.cnki_dl_btn.clicked.connect(self._on_download_selected_pdfs)
        self.cnki_dl_btn.setStyleSheet("""
            QPushButton {
                background: #1565C0; color: white; border-radius: 4px;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px; padding: 2px 12px;
            }
            QPushButton:hover  { background: #1976D2; }
            QPushButton:pressed{ background: #0D47A1; }
            QPushButton:disabled{ background: #90A4AE; color: #ECEFF1; }
        """)
        dl_hint = QLabel("仅知网来源文章可下载，需已登录且有机构权限")
        dl_hint.setStyleSheet("color: #78909C; font-size: 11px;")
        dl_row.addWidget(self.cnki_dl_btn)
        dl_row.addWidget(dl_hint)
        dl_row.addStretch()
        layout.addLayout(dl_row)

        return w

    # ── 关键词管理 ─────────────────────────────────────────────────────────────

    def _show_manual(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("使用说明")
        dlg.resize(980, 750)
        dlg.setStyleSheet("QDialog { background: #FAFAFA; }")
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(0, 0, 0, 0)
        if WEBENGINE_AVAILABLE:
            from PyQt6.QtWebEngineWidgets import QWebEngineView
            from PyQt6.QtCore import QUrl
            view = QWebEngineView()
            view.setHtml(_MANUAL_HTML, QUrl("about:blank"))
            layout.addWidget(view)
        else:
            import tempfile, webbrowser
            tmp = os.path.join(tempfile.gettempdir(), "pt_manual.html")
            with open(tmp, "w", encoding="utf-8") as f:
                f.write(_MANUAL_HTML)
            webbrowser.open(f"file:///{tmp}")
            return
        close_btn = QPushButton("关闭")
        close_btn.setFixedHeight(28)
        close_btn.setFixedWidth(80)
        close_btn.clicked.connect(dlg.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        btn_row.setContentsMargins(8, 6, 8, 8)
        layout.addLayout(btn_row)
        dlg.exec()

    def _show_search_info(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("搜索说明")
        dlg.setMinimumWidth(680)
        dlg.setMinimumHeight(700)
        dlg.setStyleSheet("QDialog { background: #FFFFFF; }")

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(0)

        text = QTextEdit()
        text.setReadOnly(True)
        text.setStyleSheet("""
            QTextEdit {
                background: #FFFFFF; border: none;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13pt; color: #212121;
            }
        """)
        text.setHtml("""
<style>
body  { font-family: 'SimSun', 'Times New Roman', serif; font-size: 13pt; color: #212121; line-height: 1.8; }
h2   { color: #1565C0; margin: 0 0 12px 0; font-size: 17pt; }
h3   { color: #1565C0; margin: 18px 0 6px 0; font-size: 15pt;
       border-left: 4px solid #1565C0; padding-left: 8px; }
h4   { color: #37474F; margin: 12px 0 4px 0; font-size: 14pt; }
.note{ background: #FFF8E1; border-left: 4px solid #F9A825;
       padding: 8px 12px; margin: 12px 0; border-radius: 2px; font-size: 13pt; color: #5D4037; }
.tip { background: #E8F5E9; border-left: 4px solid #43A047;
       padding: 8px 12px; margin: 12px 0; border-radius: 2px; font-size: 13pt; color: #1B5E20; }
.warn{ background: #FBE9E7; border-left: 4px solid #E64A19;
       padding: 8px 12px; margin: 12px 0; border-radius: 2px; font-size: 13pt; color: #BF360C; }
.ex  { background: #F3E5F5; border-left: 4px solid #7B1FA2;
       padding: 8px 12px; margin: 12px 0; border-radius: 2px; font-size: 13pt; color: #4A148C; }
code { background: #EDE7F6; color: #6A1B9A; padding: 1px 5px;
       border-radius: 3px; font-family: Consolas, monospace; font-size: 11pt; }
table{ border-collapse: collapse; width: 100%; font-size: 13pt; margin: 6px 0; }
td,th{ border: 1px solid #BDBDBD; padding: 5px 8px; }
th   { background: #1565C0; color: #FFFFFF; font-weight: bold; }
tr:nth-child(even) td { background: #F5F5F5; }
</style>

<h2>搜索说明</h2>
<p>本程序分为两条独立的检索链路：<b>英文期刊</b>通过 CrossRef 数据库抓取，<b>中文期刊</b>通过内嵌浏览器操作知网高级检索。两条链路各自工作，最终结果合并输出。</p>

<h3>一、英文期刊检索流程</h3>

<h4>第 1 步：向 CrossRef 发送检索请求</h4>
<p>CrossRef 是全球最大的学术文献元数据注册机构，几乎所有主流英文期刊的论文都在此注册了 DOI。程序对你选择的每一本期刊，用该期刊的 <b>ISSN</b> 作为范围限定，再将你输入的<b>关键词</b>作为 <code>query</code> 参数，调用 CrossRef 的检索接口。</p>
<p>CrossRef 检索的工作原理类似搜索引擎：它会在论文的<b>标题、摘要、作者、参考文献列表、期刊名</b>等全部元数据字段中综合打分，按<b>相关性</b>排序返回结果。<b>它不要求关键词一定出现在某个特定字段，也不区分大小写。</b></p>

<h4>CrossRef 是如何打分的？</h4>
<p>CrossRef 的相关性打分类似 Google 搜索，核心逻辑是：<b>关键词出现在越重要的字段，得分越高；出现次数越多，得分越高；越罕见的词匹配到，得分越高。</b>具体规则如下：</p>
<table>
<tr><th>打分维度</th><th>规则</th><th>举例</th></tr>
<tr><td>字段权重</td><td>标题字段权重最高，其次是摘要，再次是参考文献列表、关键词等</td><td>关键词出现在标题中，得分远高于只出现在参考文献里</td></tr>
<tr><td>词频</td><td>同一字段中关键词出现次数越多，得分越高</td><td>摘要中出现 5 次"robot"比只出现 1 次得分更高</td></tr>
<tr><td>词的稀有度</td><td>越罕见的词匹配上，加分越多；常见词（如 the、and）几乎不加分</td><td>匹配到"monopsony"比匹配到"labor"加分更多，因为前者更稀有</td></tr>
<tr><td>词根匹配</td><td>自动识别词根变形，不需要完全一致</td><td>输入"innovate"，含"innovation""innovative""innovating"的论文也能匹配</td></tr>
<tr><td>多词组合</td><td>输入多个词时，每个词独立打分后相加；所有词都出现得分更高</td><td>输入"minimum wage"，标题含"minimum wage"的论文得分 > 只含"wage"的论文</td></tr>
</table>

<div class="ex">
<b>打分举例：</b>你输入关键词 <code>robot automation</code>，CrossRef 对以下三篇论文打分——<br><br>
&nbsp;&nbsp;论文 A：标题含"robot"和"automation"，摘要也多次提到两词 → <b>高分</b>（标题全中 + 词频高）<br>
&nbsp;&nbsp;论文 B：标题只含"automation"，摘要未提"robot"，但参考文献列出了大量机器人相关论文 → <b>中等分</b>（部分字段命中）<br>
&nbsp;&nbsp;论文 C：标题和摘要均不含这两个词，但引用了多篇机器人论文 → <b>低分</b>（只靠参考文献勉强匹配）<br><br>
CrossRef 会将 A、B、C 按得分从高到低返回。论文 C 虽然得分低，但仍可能出现在结果中——这就是为什么 CrossRef 召回的文章需要用"篇关摘过滤"做二次精确筛选。
</div>

<div class="warn">
<b>这意味着</b>：CrossRef 返回的文章里，有些标题或摘要中未必含有你的关键词原文。这是正常现象——CrossRef 认为它们与主题相关，但匹配依据可能来自参考文献、词根变形或其他字段。这就是为什么建议配合"篇关摘过滤"做二次精确筛选。
</div>

<h4>第 2 步：OpenAlex 补全数据</h4>
<p>CrossRef 返回的摘要有时为空（尤其是较早年份的文章）。程序随后调用 <b>OpenAlex</b> 接口，按 DOI 批量查询，补充：被引次数、摘要（从倒排索引重建）、关键词字段。</p>

<div class="ex">
<b>举例：</b>你选择"全部英文顶刊（23 本）"、检索年限"近 5 年"、关键词输入 <code>minimum wage</code>：<br><br>
① 程序对每本期刊（AER、QJE、JPE……共 23 本）分别发送请求：<br>
&nbsp;&nbsp;&nbsp;<code>CrossRef API → issn:0002-8282 &amp; query:"minimum wage" &amp; from-pub-date:2020</code><br>
② 每本期刊按相关性返回最多 N 篇（由"抓取数量"参数决定）<br>
③ 所有期刊结果合并后，调用 OpenAlex 批量补充被引次数和摘要<br>
④ 去除重复文章（标题相似度判断），得到最终英文候选列表<br><br>
整个过程通常需要 30 秒至数分钟，取决于期刊数量和抓取数量。
</div>

<h4>第 3 步（可选）：篇关摘精确过滤</h4>
<p>启用"篇关摘过滤"后，程序对 CrossRef 返回的每篇文章在<b>本地</b>做精确判断。判断逻辑如下：</p>
<table>
<tr><th>勾选字段</th><th>判断方式</th></tr>
<tr><td>篇名</td><td>检查论文英文标题是否包含关键词</td></tr>
<tr><td>关键词</td><td>检查论文的 subject / keywords 字段是否包含关键词</td></tr>
<tr><td>摘要</td><td>检查英文摘要是否包含关键词</td></tr>
</table>
<p>多个字段的文本会<b>拼合成一段</b>再整体判断，因此 AND 表达式（如 <code>minimum wage * employment</code>）中的两个词可以分别出现在不同字段，依然算作匹配。不满足条件的文章将被剔除，导出 Excel 的"匹配位置"列会标注词项实际出现在哪些字段。</p>

<h4>布尔关键词与 CrossRef 的关系</h4>
<p>CrossRef 接口本身<b>不支持</b>布尔语法。当你输入布尔表达式（如 <code>minimum wage * employment * china</code>）时，程序会先提取所有词项（<code>minimum wage</code>、<code>employment</code>、<code>china</code>），拼成普通关键词发给 CrossRef；布尔逻辑的精确过滤仅在本地篇关摘阶段执行。</p>

<div class="tip">
<b>推荐使用方式：</b><br>
① 先用关键词联网搜回尽量多的候选文章（适当放宽抓取数量）<br>
② 再开启篇关摘过滤，勾选"篇名"或"摘要"做精确筛选<br>
③ 导出 Excel 查看结果，或直接生成综述
</div>

<h3>二、中文期刊检索流程（知网）</h3>

<h4>第 1 步：自动操作知网高级检索</h4>
<p>程序通过内嵌浏览器（Chromium 内核）自动打开知网<b>高级检索</b>页面，在"主题"字段中填入你输入的关键词，依次对每本你选定的中文期刊发起一次独立检索。</p>
<p>知网"主题"检索相当于篇名 + 关键词 + 摘要的联合匹配——只要关键词出现在任意一个字段即可命中，按知网内部相关性算法排序。</p>

<div class="note">
<b>注意：</b>中文关键词<b>直接</b>送入知网，不会被翻译成英文。你输入什么词，知网就检索什么词。
</div>

<h4>第 2 步：逐刊检索，按年份过滤</h4>
<p>程序对你选定的每本期刊（最多 20 本）各自发起一次检索，利用知网左侧"年度"筛选器，只保留检索年限范围内的文章，每本期刊最多抓取你设定的篇数（"每刊抓取"参数）。</p>

<div class="ex">
<b>举例：</b>关键词 <code>最低工资</code>，选择"经济研究、管理世界、中国工业经济"3 本期刊，每刊抓取 20 篇，检索年限近 5 年：<br><br>
① 打开知网高级检索 → 主题填"最低工资"→ 来源期刊填"经济研究"→ 提交<br>
② 应用年份过滤（≥ 2020）→ 翻页采集，最多采 20 条<br>
③ 重复上述步骤，依次检索"管理世界"、"中国工业经济"<br>
④ 三本期刊结果合并（最多 60 篇），逐篇访问详情页抓取摘要<br>
⑤ 与英文文献合并，去重输出
</div>

<h4>第 3 步：逐篇抓取摘要</h4>
<p>知网列表页不显示摘要，程序在采集完文章列表后，会逐篇访问详情页提取摘要和关键词。若文章总数超过 1000 篇，则跳过摘要抓取直接输出（避免等待时间过长）。</p>

<h4>篇关摘过滤对中文文献的作用</h4>
<p>启用篇关摘过滤后，中文文献同样会经过本地精确过滤，判断逻辑与英文相同：</p>
<table>
<tr><th>勾选字段</th><th>对中文文献的判断内容</th></tr>
<tr><td>篇名</td><td>检查中文标题是否包含关键词</td></tr>
<tr><td>关键词</td><td>检查知网关键词字段是否包含关键词</td></tr>
<tr><td>摘要</td><td>检查中文摘要是否包含关键词（需摘要已成功抓取）</td></tr>
</table>
<p>由于知网检索本身已限定在"主题"字段（篇名+关键词+摘要），理论上大部分结果都能通过篇关摘过滤。若摘要未能成功抓取，该篇文章会被标注"摘要缺失"，勾选"摘要"字段的过滤条件对其不生效（不剔除）。</p>

<div class="tip">
<b>使用前需先登录知网：</b>点击"知网登录"按钮，在弹出的内嵌浏览器中用校园网（机构）账号完成登录后点击"我已登录，继续"。登录状态会持久保存，下次启动无需重复操作。
</div>

<h3>三、整体流程汇总</h3>
<table>
<tr><th>阶段</th><th>英文期刊</th><th>中文期刊（知网）</th></tr>
<tr><td>检索接口</td><td>CrossRef API（按 ISSN）</td><td>知网高级检索（内嵌浏览器）</td></tr>
<tr><td>检索字段</td><td>全字段相关性（标题、摘要、参考文献等）</td><td>主题字段（篇名+关键词+摘要）</td></tr>
<tr><td>关键词语言</td><td>英文（中文关键词由 AI 翻译后检索）</td><td>直接使用输入的中文关键词</td></tr>
<tr><td>摘要来源</td><td>CrossRef 原始 + OpenAlex 补全</td><td>逐篇访问详情页提取</td></tr>
<tr><td>精确过滤</td><td>篇关摘过滤（本地子串匹配）</td><td>篇关摘过滤（本地子串匹配，同上）</td></tr>
<tr><td>结果输出</td><td colspan="2">两者合并 → 翻译 → 导出 Excel / 生成综述</td></tr>
</table>

<h3>四、文献综述生成机制</h3>

<p>勾选"生成综述"后，程序在导出 Excel 完成后调用 AI 模型（综述模型）生成文献综述，保存为 <code>.txt</code> 文件。综述根据文献数量自动选择生成方式：</p>

<h4>≤ 40 篇：直接生成</h4>
<p>将所有文献一次性送入 AI，生成包含引言、研究脉络、主要发现、方法综述、研究展望、参考文献六个部分的完整学术综述（约 1500-2500 字）。</p>

<h4>&gt; 40 篇：两阶段生成</h4>
<p>单次 prompt 过长会导致 API 超时断连，因此超过 40 篇时自动启用两阶段流程：</p>
<table>
<tr><th>阶段</th><th>操作</th><th>每次调用规模</th></tr>
<tr><td>阶段一：分批提炼</td><td>每批 40 篇，提炼核心研究问题、主要发现、方法和争议（约 600-900 字/批）</td><td>≤ 40 篇/次，安全</td></tr>
<tr><td>阶段二：整合综述</td><td>将所有批次的提炼结果合并，生成一篇完整综述</td><td>仅传摘要，token 少</td></tr>
</table>
<p>例如共 76 篇文献，阶段一分 2 批（第 1-40 篇、第 41-76 篇）各提炼一次，阶段二基于两批提炼结果生成最终综述。两阶段的综述质量与直接传入全部文献相当，且不会因超时失败。</p>

<div class="tip">
<b>简洁综述</b>：勾选"简洁综述"后改为输出约 500-800 字的简要概述，适合快速了解领域全貌。两阶段机制对简洁综述同样生效。
</div>
""")
        layout.addWidget(text)

        close_btn = QPushButton("关闭")
        close_btn.setFixedHeight(32)
        close_btn.setFixedWidth(80)
        close_btn.setStyleSheet("""
            QPushButton { background: #1565C0; color: white; border-radius: 4px;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background: #0D47A1; }
        """)
        close_btn.clicked.connect(dlg.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        layout.addSpacing(12)
        layout.addLayout(btn_row)
        dlg.exec()

    def _show_syntax_help(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("检索语法说明")
        dlg.setMinimumWidth(580)
        dlg.setMinimumHeight(480)
        dlg.setStyleSheet("QDialog { background: #FFFFFF; }")

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(0)

        text = QTextEdit()
        text.setReadOnly(True)
        text.setStyleSheet("""
            QTextEdit {
                background: #FFFFFF; border: none;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13pt; color: #212121;
            }
        """)
        text.setHtml("""
<style>
body { font-family: 'SimSun', 'Times New Roman', sans-serif; font-size: 13pt; color: #212121; }
h2 { color: #6A1B9A; margin: 0 0 12px 0; font-size: 17pt; }
h3 { color: #1565C0; margin: 14px 0 6px 0; font-size: 15pt; }
table { border-collapse: collapse; width: 100%; margin: 6px 0 12px 0; }
th { background: #F3E5F5; color: #6A1B9A; padding: 6px 12px; text-align: left; border: 1px solid #E1BEE7; }
td { padding: 5px 12px; border: 1px solid #E8EAF6; vertical-align: top; }
td:first-child { font-family: Consolas, monospace; color: #C62828; font-weight: bold; white-space: nowrap; }
tr:nth-child(even) { background: #FAFAFA; }
.note { background: #FFF8E1; border-left: 4px solid #F9A825; padding: 8px 12px; margin: 12px 0; border-radius: 2px; font-size: 13pt; color: #5D4037; }
.example { background: #F3E5F5; border-left: 4px solid #7B1FA2; padding: 6px 12px; margin: 4px 0 8px 0; border-radius: 2px; font-family: Consolas, monospace; font-size: 11pt; color: #4A148C; }
.desc { color: #546E7A; font-size: 13pt; margin: 0 0 4px 12px; }
</style>

<h2>检索语法说明</h2>

<h3>运算符一览</h3>
<table>
<tr><th>运算符</th><th>含义</th><th>说明</th></tr>
<tr><td>A * B</td><td>AND（且）</td><td>篇关摘中同时包含 A 和 B</td></tr>
<tr><td>A + B</td><td>OR（或）</td><td>篇关摘中包含 A 或 B（至少一个）</td></tr>
<tr><td>- A</td><td>NOT（非）</td><td>篇关摘中不包含 A</td></tr>
<tr><td>( )</td><td>括号分组</td><td>改变运算优先级，先计算括号内</td></tr>
<tr><td>'...'</td><td>精确短语</td><td>将引号内视为一个整体（含空格、加号等特殊字符）</td></tr>
</table>

<div class="note">
<b>运算优先级（从高到低）：</b> 括号 &gt; NOT(-) &gt; AND(*) &gt; OR(+)<br>
即 <code>A + B * C</code> 等价于 <code>A + (B * C)</code>，如需 (A+B)*C 请加括号。
</div>

<h3>使用示例</h3>

<div class="example">innovation * technology</div>
<div class="desc">同时包含"innovation"和"technology"</div>

<div class="example">innovation + technology</div>
<div class="desc">包含"innovation"或"technology"（至少其一）</div>

<div class="example">(innovation + technology) * robot</div>
<div class="desc">包含"innovation"或"technology"，且同时包含"robot"</div>

<div class="example">innovation * - policy</div>
<div class="desc">包含"innovation"，但不包含"policy"</div>

<div class="example">'2+3' * technology</div>
<div class="desc">同时包含字面字符串"2+3"（单引号保护）和"technology"</div>

<div class="example">'minimum wage' * (china + india)</div>
<div class="desc">包含短语"minimum wage"，且含"china"或"india"</div>

<h3>默认搜索是怎么工作的</h3>
<p>程序把你输入的关键词直接发给 <b>CrossRef</b> 的全文检索接口（<code>query=</code> 参数），由 CrossRef 服务器在其数据库中进行全文匹配，返回它认为最相关的文章，并按<b>相关性得分</b>从高到低排序。</p>
<p>CrossRef 的相关性算法类似搜索引擎，会综合考虑关键词出现在标题、摘要、作者、期刊名等各个字段中的频率和位置权重，<b>并不要求关键词一定出现在某个特定字段</b>——只要 CrossRef 认为该文与你的词相关，它就会被返回。因此默认搜索结果中有些文章的篇名或摘要里未必含有你的关键词原文。</p>
<p>如果你输入的是布尔表达式（如 <code>innovation * technology</code>），程序会先提取其中所有词项（如 <code>innovation</code>、<code>technology</code>），再将它们拼成一个普通查询词发给 CrossRef，由 CrossRef 按相关性返回结果。</p>

<h3>篇关摘过滤：在本地做精确筛选</h3>
<p>CrossRef 返回的结果有时包含相关性较低的文章。启用"篇关摘过滤"并勾选字段后，程序会在本地对每篇文章再做一次精确判断：</p>
<p>把你勾选的字段（篇名 / 关键词 / 摘要）的文本拼合在一起，检查是否满足你的布尔表达式。<b>不满足的文章直接过滤掉，不出现在结果中。</b></p>

<div class="note">
<b>未启用篇关摘过滤时</b>：CrossRef 返回多少篇就保留多少篇，不做任何二次过滤；布尔表达式只用于在导出的 Excel 表格中标注"匹配位置"列（标注词项出现在哪个字段），不影响文章数量。
</div>
""")
        layout.addWidget(text)

        close_btn = QPushButton("关闭")
        close_btn.setFixedHeight(32)
        close_btn.setFixedWidth(80)
        close_btn.setStyleSheet("""
            QPushButton { background: #7B1FA2; color: white; border-radius: 4px;
                font-family: 'SimSun', 'Times New Roman'; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background: #6A1B9A; }
        """)
        close_btn.clicked.connect(dlg.accept)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(close_btn)
        layout.addSpacing(10)
        layout.addLayout(btn_row)

        dlg.exec()

    def _show_journal_info(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("收录期刊列表")
        dlg.setMinimumWidth(560)
        dlg.setStyleSheet("QDialog { background: #FFFFFF; }")

        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(16, 12, 16, 12)
        layout.setSpacing(10)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: #FFFFFF; }")

        content = QWidget()
        content.setStyleSheet("background: #FFFFFF;")
        vbox = QVBoxLayout(content)
        vbox.setSpacing(0)
        vbox.setContentsMargins(0, 0, 0, 0)

        sections = [
            ("🔵 经济类 — 全部英文顶刊 Top5（5本）", [
                ("AER",     "American Economic Review"),
                ("QJE",     "Quarterly Journal of Economics"),
                ("JPE",     "Journal of Political Economy"),
                ("REStud",  "Review of Economic Studies"),
                ("ECMA",    "Econometrica"),
            ]),
            ("🔷 经济类 — 英文其他顶刊（18本）", [
                ("REStat",  "Review of Economics and Statistics"),
                ("JEL",     "Journal of Economic Literature"),
                ("JEP",     "Journal of Economic Perspectives"),
                ("AEJ:AE",  "American Economic Journal: Applied Economics"),
                ("AEJ:EP",  "American Economic Journal: Economic Policy"),
                ("AEJ:Mac", "American Economic Journal: Macroeconomics"),
                ("AEJ:Mic", "American Economic Journal: Microeconomics"),
                ("JF",      "Journal of Finance"),
                ("JFE",     "Journal of Financial Economics"),
                ("RFS",     "Review of Financial Studies"),
                ("JFQA",    "Journal of Financial and Quantitative Analysis"),
                ("JOLE",    "Journal of Labor Economics"),
                ("JHR",     "Journal of Human Resources"),
                ("JDE",     "Journal of Development Economics"),
                ("JIE",     "Journal of International Economics"),
                ("JTSA",    "Journal of Time Series Analysis"),
                ("JoE",     "Journal of Econometrics"),
                ("NBER WP", "NBER Working Paper"),
            ]),
            ("🟩 管理类 — 交大安泰 A 级（21本）", [
                (v["abbr"], v["name"]) for v in MANAGEMENT_JOURNALS.values() if v["tier"] == "A"
            ]),
            ("🟨 管理类 — 交大安泰 A- 级（25本）", [
                (v["abbr"], v["name"]) for v in MANAGEMENT_JOURNALS.values() if v["tier"] == "A-"
            ]),
            ("🟠 经济类 — 交大安泰 A 级（20本）", [
                (v["abbr"], v["name"]) for v in ECONOMICS_JOURNALS.values() if v["tier"] == "A"
            ]),
            ("🟡 经济类 — 交大安泰 A- 级（26本）", [
                (v["abbr"], v["name"]) for v in ECONOMICS_JOURNALS.values() if v["tier"] == "A-"
            ]),
            ("🔶 经济类 — 上财 First Tier 补充（4本）", [
                (v["abbr"], v["name"]) for v in ECONOMICS_JOURNALS.values() if v["tier"] == "First"
            ]),
            ("🟥 中文·经济类 — 特A/A1 级（3本）", [
                ("经研",   "经济研究"),
                ("经季",   "经济学（季刊）"),
                ("世经",   "世界经济"),
            ]),
            ("🟧 中文·经济类 — A/A2 级（6本）", [
                ("中工经", "中国工业经济"),
                ("金融研", "金融研究"),
                ("数量经", "数量经济技术经济研究"),
                ("统计研", "统计研究"),
                ("经动",   "经济学动态"),
                ("农村经", "中国农村经济"),
            ]),
            ("🟨 中文·经济类 — A- 级（10本）", [
                ("财贸经", "财贸经济"),
                ("国贸问", "国际贸易问题"),
                ("南开经", "南开经济研究"),
                ("经科",   "经济科学"),
                ("财经研", "财经研究"),
                ("经评",   "经济评论"),
                ("经家",   "经济学家"),
                ("国金研", "国际金融研究"),
                ("产经研", "产业经济研究"),
                ("人资环", "中国人口·资源与环境"),
            ]),
            ("🟩 中文·管理类 — 特A/A1 级（1本）", [
                ("管世",   "管理世界"),
            ]),
            ("🟢 中文·管理类 — A/A2 级（4本）", [
                ("会计研", "会计研究"),
                ("管科学", "管理科学学报"),
                ("系工理", "系统工程理论与实践"),
                ("中行管", "中国行政管理"),
            ]),
        ]

        font_sec  = QFont("SimSun", 10, QFont.Weight.Bold)
        font_head = QFont("SimSun", 9,  QFont.Weight.Bold)
        font_body = QFont("SimSun", 9)

        for sec_title, rows in sections:
            # 分组标题颜色：管理类绿色调，经济类蓝色调，中文期刊红色调
            if "管理类" in sec_title:
                sec_bg, sec_fg = "#E8F5E9", "#1B5E20"
            elif "中文·管理" in sec_title:
                sec_bg, sec_fg = "#E8F5E9", "#1B5E20"
            elif "中文·经济" in sec_title:
                sec_bg, sec_fg = "#FFF3E0", "#BF360C"
            else:
                sec_bg, sec_fg = "#DDEEFF", "#1A237E"
            # 分组标题
            sec_lbl = QLabel(sec_title)
            sec_lbl.setFont(font_sec)
            sec_lbl.setStyleSheet(
                f"background:{sec_bg}; color:{sec_fg}; padding:5px 8px; border-radius:3px;"
            )
            vbox.addWidget(sec_lbl)

            # 表头
            head_row = QWidget()
            head_row.setStyleSheet("background:#1F3864;")
            head_layout = QHBoxLayout(head_row)
            head_layout.setContentsMargins(8, 3, 8, 3)
            for txt, w in [("缩写", 90), ("期刊全名 / 期刊名称", 400)]:
                lbl = QLabel(txt)
                lbl.setFont(font_head)
                lbl.setStyleSheet("color:#FFFFFF; background:transparent;")
                lbl.setFixedWidth(w)
                head_layout.addWidget(lbl)
            head_layout.addStretch()
            vbox.addWidget(head_row)

            # 数据行
            for i, (abbr, name) in enumerate(rows):
                row_w = QWidget()
                bg = "#F5F9FF" if i % 2 == 0 else "#FFFFFF"
                row_w.setStyleSheet(f"background:{bg};")
                row_layout = QHBoxLayout(row_w)
                row_layout.setContentsMargins(8, 3, 8, 3)

                abbr_lbl = QLabel(abbr)
                abbr_lbl.setFont(font_body)
                abbr_lbl.setStyleSheet("color:#1565C0; background:transparent;")
                abbr_lbl.setFixedWidth(90)

                name_lbl = QLabel(name)
                name_lbl.setFont(font_body)
                name_lbl.setStyleSheet("color:#212121; background:transparent;")

                row_layout.addWidget(abbr_lbl)
                row_layout.addWidget(name_lbl)
                row_layout.addStretch()
                vbox.addWidget(row_w)

            vbox.addSpacing(8)

        vbox.addStretch()
        scroll.setWidget(content)
        layout.addWidget(scroll)

        ok_btn = QPushButton("关闭")
        ok_btn.setFixedHeight(30)
        ok_btn.setStyleSheet("""
            QPushButton { background:#1976D2; color:white; border-radius:4px;
                font-family:'SimSun', 'Times New Roman'; font-size:12px; }
            QPushButton:hover { background:#1565C0; }
        """)
        ok_btn.clicked.connect(dlg.accept)
        layout.addWidget(ok_btn, alignment=Qt.AlignmentFlag.AlignRight)

        dlg.resize(640, 750)
        dlg.exec()

    def _add_keyword(self):
        kw = self.kw_input.text().strip()
        if not kw or kw in self.keywords:
            self.kw_input.clear()
            return
        self.keywords.append(kw)
        self.kw_input.clear()
        self._refresh_chips()

    def _remove_keyword(self, kw: str):
        if kw in self.keywords:
            self.keywords.remove(kw)
        self._refresh_chips()

    def _refresh_chips(self):
        while self.chip_layout.count():
            item = self.chip_layout.takeAt(0)
            if item and item.widget() and item.widget() is not self.kw_input:
                item.widget().deleteLater()
        for i, kw in enumerate(self.keywords):
            chip = KeywordChip(kw, CHIP_COLORS[i % len(CHIP_COLORS)])
            chip.removed.connect(self._remove_keyword)
            self.chip_layout.addWidget(chip)
        # 输入框始终在末尾
        self.chip_layout.addWidget(self.kw_input)
        self.chip_container.adjustSize()
        self.chip_container.update()
        self.kw_input.setFocus()

    # ── 运行控制 ──────────────────────────────────────────────────────────────

    def _on_start_stop(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.start_btn.setText("▶  开始抓取")
            self._set_start_btn_style(running=False)
            self.progress_bar.setFormat("已停止")
            return

        if not self.keywords:
            QMessageBox.warning(self, "提示", "请至少输入一个关键词")
            return

        opts = self._build_opts()

        # 若勾选了"搜索完成后自动下载 PDF"，提前选好保存目录
        self._cnki_auto_dl_dir = ""
        if opts.get("cnki_auto_dl") and opts.get("search_cnki"):
            save_dir = QFileDialog.getExistingDirectory(
                self, "选择知网 PDF 自动下载保存目录",
                os.path.join("D:\\claude\\research", "paper")
            )
            if not save_dir:
                return  # 用户取消，不启动任务
            self._cnki_auto_dl_dir = save_dir
            self._on_log(f"[INFO] 知网 PDF 将自动保存到：{save_dir}")

        self.worker = WorkerThread(opts, cnki_manager=self.cnki_manager)
        self.worker.log_message.connect(self._on_log)
        self.worker.progress.connect(self._on_progress)
        self.worker.finished.connect(self._on_finished)
        self.worker.papers_ready.connect(self._on_papers_ready)

        self.log_text.clear()
        self.progress_bar.setValue(0)
        self.progress_bar.setFormat("准备中...")
        self.start_btn.setText("⏹  停止")
        self._set_start_btn_style(running=True)
        self.open_folder_btn.setEnabled(False)
        self.worker.start()

    def _set_start_btn_style(self, running: bool):
        if running:
            self.start_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #EF5350,stop:1 #C62828);
                    color: white; border-radius: 6px; padding: 0 16px;
                }
                QPushButton:hover { background: #EF5350; }
            """)
        else:
            self.start_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #1E88E5,stop:1 #1565C0);
                    color: white; border-radius: 6px; padding: 0 16px;
                }
                QPushButton:hover { background: #42A5F5; }
                QPushButton:pressed { background: #0D47A1; }
                QPushButton:disabled { background: #90A4AE; }
            """)

    def _build_opts(self) -> dict:
        scope_idx = self.scope_combo.currentIndex()
        en_custom = getattr(self, '_en_selected_issns', set())
        return {
            "keywords":     list(self.keywords),
            "mode":         MODE_OPTIONS[self.mode_combo.currentIndex()],
            "years":        YEARS_OPTIONS[self.years_combo.currentIndex()],
            "max":          self.max_spin.value(),
            "scope":        scope_idx,
            "skip_en":      scope_idx == 0,
            "all_journals": scope_idx == 1 and not en_custom,
            "top5":         scope_idx == 3 and not en_custom,
            "management":   scope_idx == 4 and not en_custom,
            "economics":    scope_idx == 5 and not en_custom,
            "en_custom_journals": list(en_custom) if en_custom else [],
            "strict_search": self.strict_search_cb.isChecked(),
            "filter_title":  self.filter_title_cb.isChecked(),
            "filter_kw":     self.filter_kw_cb.isChecked(),
            "filter_abs":    self.filter_abs_cb.isChecked(),
            "no_translate": not self.translate_cb.isChecked(),
            "no_review":    not self.review_cb.isChecked(),
            "review_style": "brief" if self.brief_cb.isChecked() else "academic",
            "output":       None,
            "api_key":      self.api_key_input.text().strip(),
            "base_url":     self.base_url_input.text().strip(),
            "model_fast":   (self.model_fast_combo.currentData() or self.model_fast_combo.currentText()).strip(),
            "model_strong": (self.model_strong_combo.currentData() or self.model_strong_combo.currentText()).strip(),
            "api_type":     self.api_type_combo.currentData(),
            "search_cnki":  self.cnki_cb.isChecked() if hasattr(self, 'cnki_cb') else False,
            "cnki_max":     self.cnki_max_spin.value() if hasattr(self, 'cnki_max_spin') else 20,
            "cnki_sort_id": self.cnki_sort_combo.currentData() if hasattr(self, 'cnki_sort_combo') else "PT",
            "cnki_auto_dl": self.cnki_auto_dl_cb.isChecked() if hasattr(self, 'cnki_auto_dl_cb') else False,
            "cnki_journal_filter": (
                json.dumps(self._cnki_selected_journals, ensure_ascii=False)
                if hasattr(self, '_cnki_selected_journals') and self._cnki_selected_journals
                else ""
            ),
        }

    # ── 槽函数 ────────────────────────────────────────────────────────────────

    def _on_cnki_status(self, msg: str):
        self._on_log(f"[INFO] {msg}")
        logging.info(msg)

    def _on_cnki_captcha(self):
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.warning(self, "知网验证码",
            "知网检测到自动访问，请点击「知网登录」按钮，在弹出窗口中手动完成验证后点击「我已登录，继续」。")

    def _on_cnki_download_started(self, filename: str):
        self._on_log(f"[INFO] ⬇ 开始下载：{filename}")

    def _on_cnki_download_done(self, save_path: str):
        self._on_log(f"[INFO] ✓ 下载完成：{save_path}")
        # 继续下载队列中下一篇
        self._cnki_dl_next()

    def _on_cnki_download_failed(self, msg: str):
        self._on_log(f"[ERROR] 下载失败：{msg}")
        self._cnki_dl_next()

    def _cnki_dl_next(self):
        if not hasattr(self, '_cnki_dl_queue') or not self._cnki_dl_queue:
            self._on_log("[INFO] 所有 PDF 下载任务完成")
            if hasattr(self, 'cnki_dl_btn'):
                self.cnki_dl_btn.setEnabled(True)
            return
        url, save_dir = self._cnki_dl_queue.pop(0)
        if self.cnki_manager:
            from PyQt6.QtCore import QMetaObject, Q_ARG
            QMetaObject.invokeMethod(
                self.cnki_manager, "download_pdf",
                Qt.ConnectionType.QueuedConnection,
                Q_ARG(str, url),
                Q_ARG(str, save_dir),
            )

    def _on_download_selected_pdfs(self):
        if not self.cnki_manager:
            return
        # 收集选中行对应的论文
        selected_rows = sorted(set(idx.row() for idx in self.preview_table.selectedIndexes()))
        if not selected_rows:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "提示", "请先在结果预览表格中选择要下载的行。")
            return

        # 取最近的 papers 列表（存储在 self._last_papers 中）
        papers = getattr(self, '_last_papers', [])
        to_download = []
        for r in selected_rows:
            if r < len(papers):
                p = papers[r]
                if p.source == "cnki" and p.download_url:
                    to_download.append(p)

        if not to_download:
            from PyQt6.QtWidgets import QMessageBox
            QMessageBox.information(self, "提示",
                "选中的行中没有可下载的知网文章。\n"
                "可能原因：①非知网来源；②无下载链接（需先完成搜索）；③没有机构下载权限。")
            return

        # 选择保存目录
        save_dir = QFileDialog.getExistingDirectory(self, "选择 PDF 保存目录",
            os.path.join("D:\\claude\\research", "paper"))
        if not save_dir:
            return

        self._on_log(f"[INFO] 准备下载 {len(to_download)} 篇 PDF，保存到：{save_dir}")
        self.cnki_dl_btn.setEnabled(False)

        self._cnki_dl_queue = [(p.download_url, save_dir) for p in to_download]
        self._cnki_dl_next()

    def _on_log(self, msg: str):
        if "[ERROR]" in msg:
            color = "#EF9A9A"
        elif "[WARNING]" in msg:
            color = "#FFE082"
        elif "[INFO]" in msg:
            color = "#A5D6A7" if ("完成" in msg or "已保存" in msg) else "#CFD8DC"
        else:
            color = "#B0BEC5"
        self.log_text.moveCursor(QTextCursor.MoveOperation.End)
        self.log_text.insertHtml(
            f'<span style="color:{color}; font-family:Consolas; font-size:11pt;">'
            f'{msg.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")}'
            f'</span><br>'
        )
        self.log_text.moveCursor(QTextCursor.MoveOperation.End)

    def _on_progress(self, pct: int, text: str):
        self.progress_bar.setValue(pct)
        self.progress_bar.setFormat(f"{text}  {pct}%")

    def _on_finished(self, ok: bool, path: str):
        self.start_btn.setText("▶  开始抓取")
        self._set_start_btn_style(running=False)
        if ok:
            self.progress_bar.setValue(100)
            self.progress_bar.setFormat("完成！✓")
            if path:
                self.last_output_path = path
                self.open_folder_btn.setEnabled(True)
            self._on_log(f"[INFO] ✓ 所有任务完成！共处理 {len(self.keywords)} 个关键词")
        else:
            self.progress_bar.setFormat("出错")
            self._on_log(f"[ERROR] 任务失败: {path}")

    def _on_papers_ready(self, papers: list):
        self._last_papers = list(papers)  # 保存供 PDF 下载使用
        # 有知网文章且有下载链接时启用下载按钮
        has_cnki_dl = any(p.source == "cnki" and p.download_url for p in papers)
        if hasattr(self, 'cnki_dl_btn'):
            self.cnki_dl_btn.setEnabled(has_cnki_dl and self.cnki_manager is not None)
        cols = ["#", "期刊", "年份", "引用数", "英文标题", "中文标题", "作者"]
        max_rows = len(papers)
        self.preview_table.setRowCount(max_rows)
        self.preview_table.setColumnCount(len(cols))
        self.preview_table.setHorizontalHeaderLabels(cols)
        for r, p in enumerate(papers):
            vals = [
                str(r + 1),
                p.journal_abbr or p.journal,
                str(p.year) if p.year else "",
                str(p.citations) if p.citations >= 0 else "",
                p.title,
                p.title_zh or "",
                "; ".join(p.authors[:2]) + (" 等" if len(p.authors) > 2 else ""),
            ]
            for c, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if c in (0, 2, 3):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.preview_table.setItem(r, c, item)
        self.preview_table.resizeColumnsToContents()
        self.preview_table.setColumnWidth(4, 280)
        self.preview_table.setColumnWidth(5, 240)
        self.tab_widget.setCurrentIndex(1)

        # 自动下载知网 PDF（若勾选且预先选定了保存目录）
        auto_dl_dir = getattr(self, '_cnki_auto_dl_dir', '')
        if auto_dl_dir and self.cnki_manager:
            cnki_papers = [p for p in papers if p.source == "cnki" and p.download_url]
            if cnki_papers:
                self._on_log(f"[INFO] 自动下载：共 {len(cnki_papers)} 篇知网 PDF → {auto_dl_dir}")
                if hasattr(self, 'cnki_dl_btn'):
                    self.cnki_dl_btn.setEnabled(False)
                self._cnki_dl_queue = [(p.download_url, auto_dl_dir) for p in cnki_papers]
                self._cnki_dl_next()
            else:
                self._on_log("[INFO] 自动下载：本批无可下载的知网文章（无下载链接）")

    def _open_output_folder(self):
        folder = os.path.dirname(self.last_output_path) if self.last_output_path else _get_output_root()
        if os.path.exists(folder):
            if sys.platform == "win32":
                os.startfile(folder)
            elif sys.platform == "darwin":
                subprocess.run(["open", folder])
            else:
                subprocess.run(["xdg-open", folder])

    # ── 设置持久化 ────────────────────────────────────────────────────────────

    def _save_settings(self):
        s = {
            "api_key":         self.api_key_input.text(),
            "base_url":        self.base_url_input.text(),
            "model_fast":      (self.model_fast_combo.currentData() or self.model_fast_combo.currentText()).strip(),
            "model_strong":    (self.model_strong_combo.currentData() or self.model_strong_combo.currentText()).strip(),
            "api_type":        self.api_type_combo.currentData(),
            "scope":           self.scope_combo.currentIndex(),
            "scope_v2":        True,
            "en_custom_journals": list(getattr(self, '_en_selected_issns', set())),
            "years":           self.years_combo.currentIndex(),
            "mode":            self.mode_combo.currentIndex(),
            "max":             self.max_spin.value(),
            "translate":       self.translate_cb.isChecked(),
            "review":          self.review_cb.isChecked(),
            "brief":           self.brief_cb.isChecked(),
            "strict_search":   self.strict_search_cb.isChecked(),
            "filter_title":    self.filter_title_cb.isChecked(),
            "filter_kw":       self.filter_kw_cb.isChecked(),
            "filter_abs":      self.filter_abs_cb.isChecked(),
            "search_cnki":     self.cnki_cb.isChecked() if hasattr(self, 'cnki_cb') else False,
            "cnki_max":        self.cnki_max_spin.value() if hasattr(self, 'cnki_max_spin') else 20,
            "cnki_sort_id":    self.cnki_sort_combo.currentData() if hasattr(self, 'cnki_sort_combo') else "PT",
            "cnki_auto_dl":    self.cnki_auto_dl_cb.isChecked() if hasattr(self, 'cnki_auto_dl_cb') else False,
            "cnki_journals":   (self._cnki_selected_journals if hasattr(self, '_cnki_selected_journals') else CNKI_ALL_JOURNALS),
            "window_geometry": [self.x(), self.y(), self.width(), self.height()],
        }
        try:
            with open(SETTINGS_PATH, "w", encoding="utf-8") as f:
                json.dump(s, f, ensure_ascii=False, indent=2)
            self._on_log("[INFO] 设置已保存")
        except Exception as e:
            self._on_log(f"[WARNING] 设置保存失败: {e}")

    def _load_settings(self):
        try:
            with open(SETTINGS_PATH, encoding="utf-8") as f:
                s = json.load(f)
            self.api_key_input.setText(s.get("api_key", ""))
            self.base_url_input.setText(s.get("base_url", ""))
            if hasattr(self, 'model_fast_combo'):
                saved_fast = s.get("model_fast", "")
                matched = False
                for i in range(self.model_fast_combo.count()):
                    if self.model_fast_combo.itemData(i) == saved_fast:
                        self.model_fast_combo.setCurrentIndex(i)
                        matched = True
                        break
                if not matched:
                    self.model_fast_combo.setCurrentText(saved_fast)
            if hasattr(self, 'model_strong_combo'):
                saved_strong = s.get("model_strong", "")
                matched = False
                for i in range(self.model_strong_combo.count()):
                    if self.model_strong_combo.itemData(i) == saved_strong:
                        self.model_strong_combo.setCurrentIndex(i)
                        matched = True
                        break
                if not matched:
                    self.model_strong_combo.setCurrentText(saved_strong)
            saved_type = s.get("api_type", "anthropic")
            idx = self.api_type_combo.findData(saved_type)
            if idx >= 0:
                self.api_type_combo.setCurrentIndex(idx)
            # 恢复自定义英文期刊
            saved_en = s.get("en_custom_journals", [])
            if saved_en and hasattr(self, '_en_selected_issns'):
                self._en_selected_issns = set(saved_en)
                n = len(self._en_selected_issns)
                total = sum(len(js) for tiers in _EN_JOURNAL_TREE.values() for js in tiers.values())
                self.en_journal_btn.setText(f"自定义期刊 ({n}/{total}) ✓")
                custom_label = f"自定义（{n} 本）"
                self.scope_combo.addItem(custom_label)
                self.scope_combo.setCurrentIndex(len(SCOPE_LABELS))
            else:
                saved_scope = s.get("scope", 1)
                # 兼容旧存档（无 scope_v2 标记）：旧值 0-4 对应新值 1-5
                if not s.get("scope_v2") and saved_scope < len(SCOPE_LABELS):
                    saved_scope += 1
                self.scope_combo.setCurrentIndex(saved_scope)
            self.years_combo.setCurrentIndex(s.get("years", 1))
            self.mode_combo.setCurrentIndex(s.get("mode", 0))
            if hasattr(self, 'max_spin'):
                self.max_spin.setValue(s.get("max", 50))
            self.translate_cb.setChecked(s.get("translate", True))
            self.review_cb.setChecked(s.get("review", True))
            self.brief_cb.setChecked(s.get("brief", False))
            self.strict_search_cb.setChecked(s.get("strict_search", False))
            self.filter_title_cb.setChecked(s.get("filter_title", True))
            self.filter_kw_cb.setChecked(s.get("filter_kw", True))
            self.filter_abs_cb.setChecked(s.get("filter_abs", True))
            if hasattr(self, 'cnki_cb'):
                self.cnki_cb.setChecked(s.get("search_cnki", False))
            if hasattr(self, 'cnki_max_spin'):
                self.cnki_max_spin.setValue(s.get("cnki_max", 20))
            if hasattr(self, 'cnki_sort_combo'):
                saved_sort = s.get("cnki_sort_id", "PT")
                for i in range(self.cnki_sort_combo.count()):
                    if self.cnki_sort_combo.itemData(i) == saved_sort:
                        self.cnki_sort_combo.setCurrentIndex(i)
                        break
            if hasattr(self, 'cnki_auto_dl_cb'):
                self.cnki_auto_dl_cb.setChecked(s.get("cnki_auto_dl", False))
            if hasattr(self, '_cnki_selected_journals'):
                saved = s.get("cnki_journals", CNKI_ALL_JOURNALS)
                self._cnki_selected_journals = [j for j in saved if j in CNKI_ALL_JOURNALS]
                if not self._cnki_selected_journals:
                    self._cnki_selected_journals = list(CNKI_ALL_JOURNALS)
                if hasattr(self, 'cnki_journal_btn'):
                    n = len(self._cnki_selected_journals)
                    self.cnki_journal_btn.setText(f"选择期刊 ({n}/{len(CNKI_ALL_JOURNALS)})")
            geo = s.get("window_geometry")
            if geo and len(geo) == 4:
                x, y, w, h = geo
                screen = QApplication.primaryScreen().availableGeometry()
                # 忽略异常大的窗口（全屏保存值），保持默认 1100×800
                if w < screen.width() * 0.95 and h < screen.height() * 0.95:
                    self.setGeometry(x, y, w, h)
        except (FileNotFoundError, json.JSONDecodeError, KeyError):
            pass

    def closeEvent(self, event):
        if self.worker and self.worker.isRunning():
            reply = QMessageBox.question(
                self, "确认退出", "任务正在运行，确定要退出吗？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply == QMessageBox.StandardButton.No:
                event.ignore()
                return
            self.worker.stop()
            self.worker.wait(3000)
        self._save_settings()
        super().closeEvent(event)


# =============================================================================
# 入口
# =============================================================================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setFont(QFont("SimSun", 13))
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
